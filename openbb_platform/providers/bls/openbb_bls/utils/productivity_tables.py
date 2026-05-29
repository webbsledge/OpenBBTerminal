"""BLS Productivity supplemental table fetchers and parsers."""

from __future__ import annotations

import io
import re
from datetime import date as dateType
from functools import lru_cache
from typing import Any, Literal

from openbb_core.app.model.abstract.error import OpenBBError

from openbb_bls.utils.constants import BLS_USER_AGENT

_BASE = "https://www.bls.gov/web/prod2"
_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
_XLSX_MAGIC = b"PK\x03\x04"

ProductivityDataset = Literal[
    "major-sectors-quarterly",
    "major-sectors-annual",
    "major-sectors-business-cycles",
    "total-economy-hours-employment",
]

_DATASET_FILE: dict[str, tuple[str, str]] = {
    "major-sectors-quarterly": (
        "labor-productivity-major-sectors.xlsx",
        "MachineReadable",
    ),
    "major-sectors-annual": (
        "labor-productivity-major-sectors.xlsx",
        "MachineReadable",
    ),
    "major-sectors-business-cycles": (
        "labor-productivity-major-sectors.xlsx",
        "BusinessCycles",
    ),
    "total-economy-hours-employment": (
        "total-economy-hours-employment.xlsx",
        "MachineReadable",
    ),
}

_DATASET_LABELS: dict[str, str] = {
    "major-sectors-quarterly": (
        "Quarterly labor productivity and related measures, major sectors."
    ),
    "major-sectors-annual": (
        "Annual labor productivity and related measures, major sectors."
    ),
    "major-sectors-business-cycles": (
        "Labor productivity major sectors, compound annual growth rates by "
        "approximate business cycle period."
    ),
    "total-economy-hours-employment": (
        "Quarterly hours worked and employment, total U.S. economy and subsectors."
    ),
}

# Distinct Sector / Measure / Units values published across the prod2 datasets,
# used to populate the table widget's filter dropdowns. Sourced from the live
# MachineReadable + BusinessCycles sheets (union across all four datasets).
PRODUCTIVITY_SECTORS: tuple[str, ...] = (
    "Nonfarm business sector",
    "Business sector",
    "Nonfinancial corporate sector",
    "Manufacturing sector",
    "Durable manufacturing sector",
    "Nondurable manufacturing sector",
    "Total economy",
)

PRODUCTIVITY_MEASURES: tuple[str, ...] = (
    "Labor productivity",
    "Output per worker",
    "Sectoral output",
    "Real sectoral output",
    "Value-added output",
    "Real value-added output",
    "Hours worked",
    "Employment",
    "Average weekly hours",
    "Labor compensation",
    "Hourly compensation",
    "Real hourly compensation",
    "Unit labor costs",
    "Unit nonlabor costs",
    "Unit nonlabor payments",
    "Unit combined input costs",
    "Unit profits",
    "Nonlabor costs",
    "Nonlabor payments",
    "Labor share",
    "Profits",
    "Consumer price deflator",
    "Sectoral output price deflator",
    "Value-added output price deflator",
)

PRODUCTIVITY_UNITS: tuple[str, ...] = (
    "Index (2017=100)",
    "Index",
    "% Change from previous quarter",
    "% Change from previous year",
    "% Change same quarter 1 year ago",
    "Compound annual growth rate",
    "Billions of current dollars",
    "Billions of hours",
    "Millions of jobs",
    "Current dollars per hour worked",
    "CPI-adjusted dollars per hour worked",
    "Hours worked per job per week",
    "Percentage",
    "Level - not available",
)


@lru_cache(maxsize=4)
def fetch_xlsx(filename: str) -> bytes:
    """Download one prod2 XLSX workbook.

    Memoised per filename: the prod2 workbooks are multi-megabyte and update
    only quarterly, so repeated widget queries (filter changes, dropdown
    population) reuse the in-process copy instead of re-downloading.
    """
    import requests

    url = f"{_BASE}/{filename}"
    resp = requests.get(url, headers=_HEADERS, timeout=120)
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")
    if not resp.content.startswith(_XLSX_MAGIC):
        raise OpenBBError(
            f"BLS returned non-XLSX content for {url} (length="
            f"{len(resp.content)}, head={resp.content[:8]!r})."
        )
    return resp.content


def _parse_value(cell: Any) -> tuple[float | None, str | None]:
    """Coerce one cell into ``(numeric, raw_string)``."""
    if cell is None:
        return None, None
    if isinstance(cell, bool):
        return float(cell), str(cell)
    if isinstance(cell, (int, float)):
        return float(cell), None
    text = str(cell).strip()
    if not text:
        return None, None
    try:
        return float(text.replace(",", "")), None
    except ValueError:
        return None, text


_QTR_TO_MONTH: dict[int, int] = {1: 1, 2: 4, 3: 7, 4: 10}


def _quarter_to_date(year: Any, qtr: Any) -> tuple[dateType | None, str]:
    """Convert ``(Year, Qtr)`` into a first-of-quarter ``date`` plus period kind."""
    if year is None:
        return None, "unknown"
    try:
        y = int(year)
    except (TypeError, ValueError):
        return None, "unknown"
    if isinstance(qtr, str) and qtr.strip().lower() == "annual":
        return dateType(y, 1, 1), "annual"
    try:
        q = int(qtr)
    except (TypeError, ValueError):
        return dateType(y, 1, 1), "unknown"
    month = _QTR_TO_MONTH.get(q)
    if month is None:
        return dateType(y, 1, 1), "unknown"
    return dateType(y, month, 1), f"Q{q}"


_RELEASE_DATE_RE = re.compile(
    r"Data\s+released\s+(?P<mon>\w+)\s+(?P<day>\d{1,2}),\s+(?P<year>\d{4})",
    re.IGNORECASE,
)
_MONTH_NAMES = {
    name: i
    for i, name in enumerate(
        [
            "january",
            "february",
            "march",
            "april",
            "may",
            "june",
            "july",
            "august",
            "september",
            "october",
            "november",
            "december",
        ],
        start=1,
    )
}


def _detect_release_date(rows: list[tuple[Any, ...]]) -> dateType | None:
    """Extract the 'Data released ...' date from row 1 of a prod2 sheet."""
    for row in rows[:5]:
        for cell in row:
            if not isinstance(cell, str):
                continue
            match = _RELEASE_DATE_RE.search(cell)
            if match is None:
                continue
            month = _MONTH_NAMES.get(match.group("mon").strip().lower())
            if month is None:
                continue
            try:
                return dateType(
                    int(match.group("year")), month, int(match.group("day"))
                )
            except ValueError:
                return None
    return None


def parse_dataset(content: bytes, dataset: str) -> list[dict[str, Any]]:
    """Parse one prod2 dataset into long-form records."""
    import openpyxl

    filename, sheet_name = _DATASET_FILE[dataset]
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise OpenBBError(
            f"Productivity workbook '{filename}' does not contain a "
            f"'{sheet_name}' sheet."
        )

    if dataset == "major-sectors-business-cycles":
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        wb.close()
        return _parse_business_cycles(rows, dataset)

    other_sheets = [s for s in ("Quarterly", "Annual") if s in wb.sheetnames]
    release_date: dateType | None = None
    for sn in other_sheets:
        sheet_rows = list(wb[sn].iter_rows(values_only=True))
        release_date = _detect_release_date(sheet_rows)
        if release_date is not None:
            break

    machine_rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    return _parse_machine_readable(machine_rows, dataset, release_date, filename)


def _parse_machine_readable(
    rows: list[tuple[Any, ...]],
    dataset: str,
    release_date: dateType | None,
    filename: str,
) -> list[dict[str, Any]]:
    """Reshape the MachineReadable sheet into one record per data row."""
    if not rows:
        return []
    header = rows[0]
    cols = [(str(cell).strip() if cell is not None else "") for cell in header]
    try:
        sector_idx = cols.index("Sector")
        basis_idx = cols.index("Basis")
        measure_idx = cols.index("Measure")
        units_idx = cols.index("Units")
        year_idx = cols.index("Year")
        qtr_idx = cols.index("Qtr")
        value_idx = cols.index("Value")
    except ValueError as e:
        raise OpenBBError(
            f"Productivity MachineReadable header is missing an expected column: {e}"
        ) from None
    component_idx = cols.index("Component") if "Component" in cols else None

    keep_qtr: str | None = None
    if dataset == "major-sectors-quarterly":
        keep_qtr = "quarterly"
    elif dataset == "major-sectors-annual":
        keep_qtr = "annual"
    elif dataset == "total-economy-hours-employment":
        keep_qtr = "any"

    out: list[dict[str, Any]] = []
    table_id = f"productivity-{dataset}"
    table_title = _DATASET_LABELS[dataset]
    sort_order = 0
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        qtr_cell = row[qtr_idx]
        year_cell = row[year_idx]
        period_date, period_kind = _quarter_to_date(year_cell, qtr_cell)
        if keep_qtr == "quarterly" and period_kind == "annual":
            continue
        if keep_qtr == "annual" and period_kind != "annual":
            continue
        value, value_string = _parse_value(row[value_idx])
        sort_order += 1
        out.append(
            {
                "date": period_date,
                "period_kind": period_kind,
                "year": int(year_cell) if isinstance(year_cell, (int, float)) else None,
                "quarter": (
                    int(qtr_cell)
                    if isinstance(qtr_cell, (int, float))
                    else (str(qtr_cell).strip() if qtr_cell is not None else None)
                ),
                "sector": (
                    str(row[sector_idx]).strip()
                    if row[sector_idx] is not None
                    else None
                ),
                "basis": (
                    str(row[basis_idx]).strip() if row[basis_idx] is not None else None
                ),
                "component": (
                    str(row[component_idx]).strip()
                    if component_idx is not None
                    and component_idx < len(row)
                    and row[component_idx] is not None
                    else None
                ),
                "measure": (
                    str(row[measure_idx]).strip()
                    if row[measure_idx] is not None
                    else ""
                ),
                "units": (
                    str(row[units_idx]).strip() if row[units_idx] is not None else ""
                ),
                "value": value,
                "value_string": value_string,
                "row_index": sort_order,
                "table_id": table_id,
                "table_title": table_title,
                "source_file": filename,
                "release_date": release_date,
            }
        )
    return out


def _parse_business_cycles(
    rows: list[tuple[Any, ...]], dataset: str
) -> list[dict[str, Any]]:
    """Reshape the BusinessCycles sheet from wide → long."""
    release_date = _detect_release_date(rows)
    header_idx: int | None = None
    for i, row in enumerate(rows):
        first = str(row[0]).strip().lower() if row[0] else ""
        if first == "sector":
            header_idx = i
            break
    if header_idx is None:
        raise OpenBBError(
            "Productivity BusinessCycles sheet is missing the 'Sector' header row."
        )
    header_row = rows[header_idx]
    cycle_headers: list[tuple[int, str]] = []
    for i, cell in enumerate(header_row[4:], start=4):
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        cycle_headers.append((i, text))
    if not cycle_headers:
        raise OpenBBError(
            "Productivity BusinessCycles sheet has no business-cycle column "
            "headers after the Sector / Basis / Measure / Units columns."
        )

    out: list[dict[str, Any]] = []
    table_id = f"productivity-{dataset}"
    table_title = _DATASET_LABELS[dataset]
    sort_order = 0
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None for c in row):
            continue
        sector = str(row[0]).strip() if row[0] is not None else None
        if not sector:
            continue
        basis = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        measure = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        units = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        if measure.lower().startswith("data released") or sector.lower().startswith(
            "data released"
        ):
            continue
        for col_idx, cycle_label in cycle_headers:
            # read_only pads every row to max_column, so col_idx (from the
            # header row) is always in range.
            value, value_string = _parse_value(row[col_idx])
            if value is None and value_string is None:
                continue
            sort_order += 1
            start_date, end_date = _parse_cycle_period(cycle_label)
            out.append(
                {
                    "date": start_date,
                    "period_kind": "business_cycle",
                    "year": None,
                    "quarter": None,
                    "sector": sector,
                    "basis": basis,
                    "component": None,
                    "measure": measure,
                    "units": units,
                    "value": value,
                    "value_string": value_string,
                    "row_index": sort_order,
                    "table_id": table_id,
                    "table_title": table_title,
                    "source_file": "labor-productivity-major-sectors.xlsx",
                    "release_date": release_date,
                    "cycle_period": cycle_label,
                    "cycle_start_date": start_date,
                    "cycle_end_date": end_date,
                }
            )
    return out


_CYCLE_RE = re.compile(r"(?P<sy>\d{4})\s*Q(?P<sq>\d)\s*-\s*(?P<ey>\d{4})\s*Q(?P<eq>\d)")


def _parse_cycle_period(label: str) -> tuple[dateType | None, dateType | None]:
    """Resolve a business-cycle column header into ``(start, end)`` dates."""
    match = _CYCLE_RE.search(label)
    if match is None:
        return None, None
    try:
        sy = int(match.group("sy"))
        sq = int(match.group("sq"))
        ey = int(match.group("ey"))
        eq = int(match.group("eq"))
        start = dateType(sy, _QTR_TO_MONTH.get(sq, 1), 1)
        end = dateType(ey, _QTR_TO_MONTH.get(eq, 1), 1)
        return start, end
    except (ValueError, KeyError):
        return None, None
