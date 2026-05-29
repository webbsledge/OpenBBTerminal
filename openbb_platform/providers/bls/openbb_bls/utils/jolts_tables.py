"""BLS JOLTS supplemental table fetchers and parsers."""

from __future__ import annotations

import io
import re
from datetime import (
    date as dateType,
    datetime as dateTimeType,
)
from typing import Any, Literal

from openbb_core.app.model.abstract.error import OpenBBError

from openbb_bls.utils.constants import BLS_USER_AGENT

_BASE_NATIONAL = "https://www.bls.gov/web/jolts"
_BASE_STATE = "https://www.bls.gov/web/jltst"
_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "text/plain, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*",
}
_XLSX_MAGIC = b"PK\x03\x04"

JoltsScope = Literal["national", "state"]

_NATIONAL_TABLE_MEASURES: dict[int, tuple[str, str]] = {
    1: ("Job openings", "over-the-month"),
    2: ("Hires", "over-the-month"),
    3: ("Total separations", "over-the-month"),
    4: ("Quits", "over-the-month"),
    5: ("Layoffs and discharges", "over-the-month"),
    6: ("Other separations", "over-the-month"),
    7: ("Job openings", "over-the-year"),
    8: ("Hires", "over-the-year"),
    9: ("Total separations", "over-the-year"),
    10: ("Quits", "over-the-year"),
    11: ("Layoffs and discharges", "over-the-year"),
    12: ("Other separations", "over-the-year"),
}

_STATE_TABLE_MEASURES: dict[int, tuple[str, str]] = {
    1: ("Job openings", "over-the-month"),
    2: ("Hires", "over-the-month"),
    3: ("Total separations", "over-the-month"),
    4: ("Quits", "over-the-month"),
    5: ("Layoffs and discharges", "over-the-month"),
    6: ("Job openings", "over-the-year"),
    7: ("Hires", "over-the-year"),
    8: ("Total separations", "over-the-year"),
    9: ("Quits", "over-the-year"),
    10: ("Layoffs and discharges", "over-the-year"),
}


def list_change_analysis_tables() -> list[dict[str, Any]]:
    """Return one entry per published JOLTS change-analysis TXT table."""
    out: list[dict[str, Any]] = []
    for table_number, (measure, period) in _NATIONAL_TABLE_MEASURES.items():
        out.append(
            {
                "scope": "national",
                "table_number": table_number,
                "measure": measure,
                "period": period,
                "url": f"{_BASE_NATIONAL}/jlt_table{table_number}.txt",
            }
        )
    for table_number, (measure, period) in _STATE_TABLE_MEASURES.items():
        out.append(
            {
                "scope": "state",
                "table_number": table_number,
                "measure": measure,
                "period": period,
                "url": f"{_BASE_STATE}/jltst_table{table_number}.txt",
            }
        )
    return out


def fetch_change_analysis_txt(scope: str, table_number: int) -> str:
    """Download one JOLTS change-analysis TXT file as decoded text."""
    import requests

    if scope == "national":
        url = f"{_BASE_NATIONAL}/jlt_table{table_number}.txt"
    elif scope == "state":
        url = f"{_BASE_STATE}/jltst_table{table_number}.txt"
    else:
        raise OpenBBError(f"Unknown JOLTS scope '{scope}'.")
    resp = requests.get(url, headers=_HEADERS, timeout=60)
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")
    try:
        return resp.content.decode("utf-8")
    except UnicodeDecodeError:
        return resp.content.decode("latin-1", errors="replace")


_TITLE_RE = re.compile(
    r"^TABLE\s+(?P<num>\d+):\s*(?P<measure>.+?)\s+estimated.*?between\s+"
    r"(?P<start_mon>\w+)\s+(?P<start_year>\d{4})\s+and\s+"
    r"(?P<end_mon>\w+)\s+(?P<end_year>\d{4})\s*[,.]?",
    re.IGNORECASE,
)

_MONTHS = {
    name: i
    for i, name in enumerate(
        [
            "january",
            "february",
            "march",
            "april",
            "may",
            "june",
            "july",
            "august",
            "september",
            "october",
            "november",
            "december",
        ],
        start=1,
    )
}

_DATA_LINE_RE = re.compile(
    r"^(?P<label>.+?)"
    r"\s{2,}(?P<rate_change>-?\d+(?:\.\d+)?)"
    r"\s+(?P<rate_min>-?\d+(?:\.\d+)?)"
    r"(?:\s+(?P<rate_pass>YES))?"
    r"\s+(?P<level_change>-?\d+(?:\.\d+)?)"
    r"\s+(?P<level_min>-?\d+(?:\.\d+)?)"
    r"(?:\s+(?P<level_pass>YES))?"
    r"\s*$"
)

_SOURCE_RE = re.compile(
    r"SOURCE:.*?,\s*(?P<mon>\w+)\s+(?P<day>\d{1,2}),\s+(?P<year>\d{4})", re.IGNORECASE
)

_SECTION_RE = re.compile(r"^\s+(?P<name>[A-Z][A-Z \-]+)\s*$")


def _month_to_int(name: str) -> int | None:
    """Map a month name (case-insensitive, possibly truncated) to its int."""
    norm = name.strip().lower().rstrip(".")
    if norm in _MONTHS:
        return _MONTHS[norm]
    for key, value in _MONTHS.items():
        if key.startswith(norm):
            return value
    return None


def _parse_period(
    start_mon: str, start_year: str, end_mon: str, end_year: str
) -> tuple[dateType | None, dateType | None]:
    """Coerce raw title month-year strings into ``(start, end)`` first-of-month dates."""
    sm = _month_to_int(start_mon)
    em = _month_to_int(end_mon)
    try:
        s = dateType(int(start_year), sm, 1) if sm else None
        e = dateType(int(end_year), em, 1) if em else None
    except ValueError:
        return None, None
    return s, e


def parse_change_analysis(
    content: str, scope: str, table_number: int
) -> dict[str, Any]:
    """Parse one JOLTS change-analysis TXT into structured records."""
    lines = content.splitlines()
    if not lines:
        raise OpenBBError(f"Empty JOLTS {scope} table {table_number} TXT response.")

    title_match = _TITLE_RE.match(lines[0])
    if title_match is None:
        raise OpenBBError(
            f"JOLTS {scope} table {table_number} TXT header is missing the "
            "expected 'TABLE N: <measure> estimated ... between A YYYY and "
            "B YYYY' line."
        )
    measure = title_match.group("measure").strip()
    start_date, end_date = _parse_period(
        title_match.group("start_mon"),
        title_match.group("start_year"),
        title_match.group("end_mon"),
        title_match.group("end_year"),
    )

    second_line = lines[1] if len(lines) > 1 else ""
    seasonally_adjusted: bool | None
    if "not seasonally adjusted" in second_line.lower():
        seasonally_adjusted = False
    elif "seasonally adjusted" in second_line.lower():
        seasonally_adjusted = True
    else:
        seasonally_adjusted = None

    period_kind = "over-the-month"
    if start_date and end_date and (end_date.year - start_date.year) >= 1:
        period_kind = "over-the-year"

    source_date: dateType | None = None
    for line in lines:
        match = _SOURCE_RE.search(line)
        if match is None:
            continue
        mn = _month_to_int(match.group("mon"))
        if mn is None:
            continue
        try:
            source_date = dateType(
                int(match.group("year")), mn, int(match.group("day"))
            )
        except ValueError:
            source_date = None
        break

    table_id = f"jolts-{scope}-t{table_number}"
    table_title = lines[0].strip()
    if len(lines) > 1 and lines[1].strip():
        table_title = f"{table_title} {lines[1].strip()}"

    out: list[dict[str, Any]] = []
    current_section: str | None = None
    sort_order = 0
    for raw_line in lines[2:]:
        line = raw_line.rstrip()
        if not line.strip():
            continue
        if _SOURCE_RE.search(line) or line.lstrip().startswith("*NOTE"):
            break
        if line.lstrip().startswith("NOTE:") or line.lstrip().startswith("*"):
            continue
        if line.lstrip().lower().startswith("is used in testing"):
            continue
        section_match = _SECTION_RE.match(line)
        if section_match and not any(ch.isdigit() for ch in line):
            current_section = section_match.group("name").strip().title()
            continue
        data_match = _DATA_LINE_RE.match(line)
        if data_match is None:
            continue
        label_raw = data_match.group("label").rstrip()
        label = label_raw.strip()
        if not label:
            continue
        leading_ws = len(label_raw) - len(label_raw.lstrip(" "))
        # The regex groups are all `-?\d+(?:\.\d+)?`, so float() is total.
        rate_change = float(data_match.group("rate_change"))
        rate_min_significant = float(data_match.group("rate_min"))
        level_change = float(data_match.group("level_change"))
        level_min_significant = float(data_match.group("level_min"))
        rate_passes = bool(data_match.group("rate_pass"))
        level_passes = bool(data_match.group("level_pass"))

        sort_order += 1
        out.append(
            {
                "table_id": table_id,
                "table_title": table_title,
                "scope": scope,
                "table_number": table_number,
                "measure": measure,
                "period": period_kind,
                "seasonally_adjusted": seasonally_adjusted,
                "period_start": start_date,
                "period_end": end_date,
                "source_date": source_date,
                "row_index": sort_order,
                "section": current_section,
                "indent_spaces": leading_ws,
                "label": label,
                "rate_change": rate_change,
                "rate_min_significant": rate_min_significant,
                "rate_passes_significance": rate_passes,
                "level_change_thousands": level_change,
                "level_min_significant_thousands": level_min_significant,
                "level_passes_significance": level_passes,
            }
        )
    return {"rows": out, "table_id": table_id, "table_title": table_title}


def fetch_revision_xlsx(seasonally_adjusted: bool) -> bytes:
    """Download the SA or NSA revision XLSX workbook from bls.gov."""
    import requests

    stem = "sa-revision-tables" if seasonally_adjusted else "nsa-revision-tables"
    url = f"{_BASE_NATIONAL}/{stem}.xlsx"
    resp = requests.get(url, headers=_HEADERS, timeout=120)
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")
    if not resp.content.startswith(_XLSX_MAGIC):
        raise OpenBBError(
            f"BLS returned non-XLSX content for {url} (length="
            f"{len(resp.content)}, head={resp.content[:8]!r})."
        )
    return resp.content


_REVISION_MEASURES = (
    "JOB OPENINGS",
    "HIRES",
    "TOTAL SEPARATIONS",
    "QUITS",
    "LAYOFFS & DISCHARGES",
    "OTHER SEPARATIONS",
)

_REVISION_FIELD_KEYS = (
    "level_1st",
    "level_2nd",
    "level_benchmark",
    "revision_1st_to_2nd_level",
    "revision_1st_to_2nd_pct",
    "revision_2nd_to_benchmark_level",
    "revision_2nd_to_benchmark_pct",
)


def _cell_to_float(cell: Any) -> float | None:
    """Coerce a numeric XLSX cell into float; None for empty / non-numeric."""
    if cell is None:
        return None
    if isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    if isinstance(cell, str):
        text = cell.strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def parse_revision_xlsx(
    content: bytes, seasonally_adjusted: bool
) -> list[dict[str, Any]]:
    """Parse a JOLTS SA/NSA revision XLSX into wide rows keyed by (industry, month, measure)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    sort_order = 0
    table_id = "jolts-revisions-sa" if seasonally_adjusted else "jolts-revisions-nsa"
    table_title = (
        "JOLTS revision tables — seasonally adjusted"
        if seasonally_adjusted
        else "JOLTS revision tables — not seasonally adjusted"
    )
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue
        title_cell = rows[0][0] if rows[0] else None
        industry_name = ""
        if isinstance(title_cell, str):
            text = title_cell.strip()
            for marker in (", seasonally adjusted", ", not seasonally adjusted"):
                idx = text.lower().find(marker)
                if idx >= 0:
                    industry_name = text[:idx].strip()
                    break
            if not industry_name:
                industry_name = text
        block_starts = [
            i
            for i in range(len(rows[1]))
            if isinstance(rows[1][i], str)
            and rows[1][i].strip().upper() in _REVISION_MEASURES
        ]
        for col_start in block_starts:
            # block_starts only holds indices where rows[1][i] is a str.
            measure = rows[1][col_start].strip().title()
            for row_idx in range(4, len(rows)):
                row = rows[row_idx]
                if col_start - 1 >= len(row):
                    break
                month_cell = row[col_start - 1]
                if month_cell is None:
                    continue
                if isinstance(month_cell, dateTimeType):
                    month_date = month_cell.date()
                elif isinstance(month_cell, dateType):
                    month_date = month_cell
                else:
                    continue
                month_date = dateType(month_date.year, month_date.month, 1)
                values: dict[str, float | None] = {}
                for offset, key in enumerate(_REVISION_FIELD_KEYS):
                    cell_col = col_start + offset
                    if cell_col >= len(row):
                        values[key] = None
                        continue
                    values[key] = _cell_to_float(row[cell_col])
                if all(v is None for v in values.values()):
                    continue
                sort_order += 1
                record: dict[str, Any] = {
                    "date": month_date,
                    "industry_code": sheet_name,
                    "industry_name": industry_name,
                    "seasonally_adjusted": seasonally_adjusted,
                    "measure": measure,
                    "row_index": sort_order,
                    "table_id": table_id,
                    "table_title": table_title,
                }
                record.update(values)
                out.append(record)
    wb.close()
    out.sort(
        key=lambda r: (
            -r["date"].toordinal(),
            r["industry_code"],
            r["measure"],
        )
    )
    return out
