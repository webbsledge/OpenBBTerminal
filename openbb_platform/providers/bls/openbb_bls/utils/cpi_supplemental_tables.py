"""BLS CPI supplemental table registry, fetch, and parse helpers."""

from __future__ import annotations

import calendar
import io
import re
from datetime import date as dateType
from functools import lru_cache
from typing import Any

from openbb_core.app.model.abstract.error import OpenBBError

from openbb_bls.utils.constants import BLS_USER_AGENT

_INDEX_URL = "https://www.bls.gov/cpi/tables/supplemental-files/home.htm"
_BASE = "https://www.bls.gov/cpi/tables/supplemental-files"
_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"),
}
_XLSX_MAGIC = b"PK\x03\x04"
_XLSX_HREF_RE = re.compile(
    r"^/cpi/tables/supplemental-files/(?P<stem>[a-z0-9\-]+)"
    r"-(?P<y>\d{4})(?P<m>\d{2})\.xlsx$",
    re.IGNORECASE,
)


class _TableSpec:
    """Description of one CPI supplemental table family."""

    __slots__ = ("key", "stem", "sheet", "label")

    def __init__(
        self,
        key: str,
        stem: str,
        sheet: str | None,
        label: str,
    ) -> None:
        self.key = key
        self.stem = stem
        self.sheet = sheet
        self.label = label


TABLE_REGISTRY: dict[str, _TableSpec] = {
    "c-cpi-u": _TableSpec(
        key="c-cpi-u",
        stem="c-cpi-u",
        sheet=None,
        label="Chained CPI-U U.S. city average, by item.",
    ),
    "cpi-u-us": _TableSpec(
        key="cpi-u-us",
        stem="cpi-u",
        sheet="US",
        label="CPI-U U.S. city average, by expenditure category.",
    ),
    "cpi-u-regional": _TableSpec(
        key="cpi-u-regional",
        stem="cpi-u",
        sheet="Regional",
        label="CPI-U selected areas, all items index.",
    ),
    "cpi-w": _TableSpec(
        key="cpi-w",
        stem="cpi-w",
        sheet=None,
        label="CPI-W U.S. city average, by expenditure category.",
    ),
    "historical-cpi-u-index": _TableSpec(
        key="historical-cpi-u-index",
        stem="historical-cpi-u",
        sheet="Index values",
        label="Historical CPI-U U.S. city average, all items, monthly indexes.",
    ),
    "historical-cpi-u-averages": _TableSpec(
        key="historical-cpi-u-averages",
        stem="historical-cpi-u",
        sheet="Index averages",
        label="Historical CPI-U U.S. city average, all items, semiannual and annual averages.",
    ),
}

_MONTH_LOOKUP: dict[str, int] = {}
for _m in range(1, 13):
    _MONTH_LOOKUP[calendar.month_abbr[_m].lower()] = _m
    _MONTH_LOOKUP[calendar.month_name[_m].lower()] = _m
_MONTH_LOOKUP["sept"] = 9

_SINGLE_PERIOD_RE = re.compile(r"([A-Za-z]+)\.?\s+(\d{4})", re.IGNORECASE)
_RANGE_PERIOD_RE = re.compile(
    r"([A-Za-z]+)\.?\s*(\d{4})?\s*[-–]\s*([A-Za-z]+)\.?\s*(\d{4})", re.IGNORECASE
)
_PCT_TO_FROM_RE = re.compile(
    r"percent change to\s+([A-Za-z]+)\.?\s*(\d{4})\s*from",
    re.IGNORECASE,
)


def _decode_html(content: bytes) -> str:
    """Decode HTML bytes, falling back to latin-1."""
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("latin-1", errors="replace")


@lru_cache(maxsize=1)
def list_supp_index() -> dict[str, tuple[dateType, ...]]:
    """Return mapping of file stem to published snapshot dates."""
    import requests
    from bs4 import BeautifulSoup

    resp = requests.get(_INDEX_URL, headers=_HEADERS, timeout=30)
    if resp.status_code != 200:
        raise OpenBBError(
            f"BLS returned HTTP {resp.status_code} fetching {_INDEX_URL}."
        )
    soup = BeautifulSoup(_decode_html(resp.content), "lxml")
    by_stem: dict[str, set[dateType]] = {}
    for anchor in soup.find_all("a", href=True):
        href = str(anchor["href"]).strip()
        match = _XLSX_HREF_RE.match(href)
        if not match:
            continue
        stem = match.group("stem").lower()
        year = int(match.group("y"))
        month = int(match.group("m"))
        try:
            d = dateType(year, month, 1)
        except ValueError:
            continue
        by_stem.setdefault(stem, set()).add(d)
    return {stem: tuple(sorted(dates)) for stem, dates in by_stem.items()}


def _stem_url(stem: str, year: int, month: int) -> str:
    """Build the canonical XLSX URL for one stem and snapshot month."""
    return f"{_BASE}/{stem}-{year:04d}{month:02d}.xlsx"


def fetch_xlsx(stem: str, year: int, month: int) -> bytes | None:
    """Download one supplemental XLSX; None when missing."""
    import requests

    url = _stem_url(stem, year, month)
    resp = requests.get(url, headers=_HEADERS, timeout=60, allow_redirects=False)
    if resp.status_code in (301, 302, 303, 307, 308, 404):
        return None
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")
    if not resp.content.startswith(_XLSX_MAGIC):
        return None
    return resp.content


def discover_latest(stem: str) -> tuple[int, int, bytes]:
    """Locate the most recent published XLSX for one stem."""
    index = list_supp_index().get(stem)
    candidates: list[tuple[int, int]] = []
    if index:
        candidates.extend((d.year, d.month) for d in reversed(index))
    today = dateType.today()
    year, month = today.year, today.month
    for _ in range(6):
        candidates.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    seen: set[tuple[int, int]] = set()
    for y, m in candidates:
        if (y, m) in seen:
            continue
        seen.add((y, m))
        content = fetch_xlsx(stem, y, m)
        if content is not None:
            return y, m, content
    raise OpenBBError(
        f"Could not locate a published CPI supplemental XLSX for stem "
        f"'{stem}' within the last six months."
    )


def _norm(text: Any) -> str:
    """Collapse whitespace and newlines in a header cell."""
    if text is None:
        return ""
    s = str(text).replace("\n", " ").replace("\r", " ")
    return " ".join(s.split()).strip()


def _month_token_to_num(token: str) -> int | None:
    """Map a BLS month token to 1..12."""
    return _MONTH_LOOKUP.get(token.strip().lower().rstrip("."))


def _parse_single_period(text: str) -> dateType | None:
    """Parse a single-period header like 'Apr. 2026' into a first-of-month date."""
    match = _SINGLE_PERIOD_RE.search(text)
    if match is None:
        return None
    month_token, year_token = match.groups()
    month = _month_token_to_num(month_token)
    if month is None:
        return None
    return dateType(int(year_token), month, 1)


def _parse_range_period(text: str) -> tuple[dateType, dateType] | None:
    """Parse a range header like 'Apr. 2025-Apr. 2026' into (start, end)."""
    match = _RANGE_PERIOD_RE.search(text)
    if match is None:
        return None
    start_m, start_y, end_m, end_y = match.groups()
    start_month = _month_token_to_num(start_m)
    end_month = _month_token_to_num(end_m)
    if start_month is None or end_month is None or end_y is None:
        return None
    end_year = int(end_y)
    start_year = int(start_y) if start_y is not None else end_year
    return dateType(start_year, start_month, 1), dateType(end_year, end_month, 1)


def _months_between(start: dateType, end: dateType) -> int:
    """Absolute number of calendar months between two first-of-month dates."""
    return abs((end.year - start.year) * 12 + (end.month - start.month))


def _parse_to_from_super(super_text: str) -> dateType | None:
    """Extract the 'to <date>' anchor from a 'Percent change to X from:' super-header."""
    match = _PCT_TO_FROM_RE.search(super_text)
    if match is None:
        return None
    month_token, year_token = match.group(1), match.group(2)
    month = _month_token_to_num(month_token)
    if month is None:
        return None
    return dateType(int(year_token), month, 1)


def _parse_value(cell: Any) -> tuple[float | None, str | None]:
    """Coerce one cell into (numeric, raw_string)."""
    if cell is None:
        return None, None
    if isinstance(cell, bool):
        return float(cell), str(cell)
    if isinstance(cell, (int, float)):
        return float(cell), None
    text = str(cell).strip()
    if not text:
        return None, None
    try:
        return float(text.replace(",", "")), None
    except ValueError:
        return None, text


_HEADER_TOKENS = {"indent level", "indent\nlevel"}


def _is_header_row(row: tuple[Any, ...]) -> bool:
    """Return True when row[0] is the literal 'Indent Level' marker."""
    if not row or row[0] is None:
        return False
    return str(row[0]).strip().lower() in _HEADER_TOKENS


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    """Return True when every cell is empty / whitespace."""
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


_FOOTNOTE_DEF_RE = re.compile(r"^\(([^)]+)\)\s*(.*)$")
_LABEL_REF_RE = re.compile(r"\(([^)]+)\)")


def _is_note_row(row: tuple[Any, ...]) -> bool:
    """Detect trailing NOTE / footnote-definition rows."""
    text: str | None = None
    for cell in row:
        if isinstance(cell, str) and cell.strip():
            text = cell.strip()
            break
    if text is None:
        return False
    if text.upper().startswith("NOTE:"):
        return True
    if _FOOTNOTE_DEF_RE.match(text):
        return True
    if text.lower().startswith("indexes are issued"):
        return True
    return bool(text.lower().startswith("footnote"))


def _parse_footnotes(rows: list[tuple[Any, ...]]) -> dict[str, str]:
    """Extract (N) footnote definitions from trailing rows."""
    out: dict[str, str] = {}
    for row in rows:
        for cell in row:
            if not isinstance(cell, str):
                continue
            text = cell.strip()
            match = _FOOTNOTE_DEF_RE.match(text)
            if match is None:
                continue
            marker = f"({match.group(1)})"
            body = match.group(2).strip() or text
            out[marker] = body
            break
    return out


def _resolve_footnotes(label: str | None, footnotes: dict[str, str]) -> str | None:
    """Concatenate footnote text for any (N) markers in label."""
    if not label or not footnotes:
        return None
    matched: list[str] = []
    seen: set[str] = set()
    for match in _LABEL_REF_RE.finditer(label):
        marker = match.group(0)
        if marker not in footnotes or marker in seen:
            continue
        seen.add(marker)
        matched.append(footnotes[marker])
    if not matched:
        return None
    return "\n\n".join(matched)


_RELATIVE_IMPORTANCE_MARKER = "relative importance"
_PRICING_SCHEDULE_MARKER = "pricing schedule"


def _classify_columns(
    super_row: tuple[Any, ...], sub_row: tuple[Any, ...] | None
) -> list[tuple[str, Any] | None]:
    """Classify each XLSX column into (kind, payload) for downstream pivoting."""
    n = len(super_row)
    filled_super: list[str] = []
    last = ""
    for i in range(n):
        cell = _norm(super_row[i] if i < len(super_row) else None)
        if cell:
            last = cell
        filled_super.append(last)
    sub = sub_row if sub_row is not None else tuple([None] * n)

    out: list[tuple[str, Any] | None] = []
    for i in range(n):
        s = filled_super[i]
        sub_text = _norm(sub[i] if i < len(sub) else None)
        s_lo = s.lower()
        if i == 0 or s_lo == "indent level":
            out.append(None)
            continue
        if s_lo.startswith(
            ("expenditure category", "special aggregate", "title", "item", "area")
        ):
            out.append(None)
            continue
        if _RELATIVE_IMPORTANCE_MARKER in s_lo and not sub_text:
            out.append(("relative_importance", None))
            continue
        if _PRICING_SCHEDULE_MARKER in s_lo and not sub_text:
            out.append(("pricing_schedule", None))
            continue
        if "unadjusted index" in s_lo or s_lo.endswith("indexes"):
            d = _parse_single_period(sub_text)
            out.append(("index_value", d))
            continue
        if "seasonally adjusted index" in s_lo:
            d = _parse_single_period(sub_text)
            out.append(("sa_index_value", d))
            continue
        if "seasonally adjusted percent change" in s_lo:
            rng = _parse_range_period(sub_text)
            end = rng[1] if rng else None
            out.append(("pct_change_1m_sa", end))
            continue
        if "percent change to" in s_lo:
            end = _parse_to_from_super(s)
            start = _parse_single_period(sub_text)
            if end is None or start is None:
                out.append(None)
                continue
            months = _months_between(start, end)
            if months >= 12:
                out.append(("pct_change_12m", end))
            elif months == 1:
                out.append(("pct_change_1m_nsa", end))
            else:
                out.append(("pct_change_window", (start, end)))
            continue
        if "unadjusted percent change" in s_lo or s_lo == "percent change":
            rng = _parse_range_period(sub_text)
            if rng is None:
                out.append(None)
                continue
            start, end = rng
            if _months_between(start, end) >= 12:
                out.append(("pct_change_12m", end))
            else:
                out.append(("pct_change_1m_nsa", end))
            continue
        out.append(None)
    return out


def _parse_index_values_sheet(
    rows: list[tuple[Any, ...]],
    spec: _TableSpec,
    snapshot_date: dateType,
    table_id: str,
    table_name: str,
) -> list[dict[str, Any]]:
    """Parse the Historical CPI-U Index Values sheet (year row × 12 month columns)."""
    header_idx: int | None = None
    for i, row in enumerate(rows):
        if _is_header_row(row):
            header_idx = i
            break
    if header_idx is None:
        return []
    month_cols: list[tuple[int, int]] = []
    for j, cell in enumerate(rows[header_idx][2:], start=2):
        token = _norm(cell)
        m = _month_token_to_num(token)
        if m is not None:
            month_cols.append((j, m))
    out: list[dict[str, Any]] = []
    row_index = 0
    for source_idx, row in enumerate(rows[header_idx + 1 :], start=1):
        if not row or _is_blank_row(row):
            continue
        if _is_note_row(row):
            break
        year_cell = row[1] if len(row) > 1 else None
        try:
            year = int(str(year_cell).strip())
        except (TypeError, ValueError):
            continue
        for col_idx, month in month_cols:
            # read_only iter_rows pads every row to max_column, so col_idx
            # (taken from the header row) is always within range.
            value, value_string = _parse_value(row[col_idx])
            if value is None and value_string is None:
                continue
            row_index += 1
            out.append(
                {
                    "date": dateType(year, month, 1),
                    "snapshot_date": snapshot_date,
                    "label": "All items",
                    "indent_level": 0,
                    "frequency": "monthly",
                    "index_value": value,
                    "sa_index_value": None,
                    "pct_change_1m_nsa": None,
                    "pct_change_1m_sa": None,
                    "pct_change_12m": None,
                    "pct_change_window": None,
                    "pct_change_window_start_date": None,
                    "relative_importance": None,
                    "pricing_schedule": None,
                    "half": None,
                    "value_string": value_string,
                    "footnote": None,
                    "table_key": spec.key,
                    "table_id": table_id,
                    "table_name": table_name,
                    "sheet": spec.sheet or "",
                    "row_index": row_index,
                }
            )
    return out


def _parse_index_averages_sheet(
    rows: list[tuple[Any, ...]],
    spec: _TableSpec,
    snapshot_date: dateType,
    table_id: str,
    table_name: str,
) -> list[dict[str, Any]]:
    """Parse the Historical CPI-U Index Averages sheet (semiannual + annual aggregates)."""
    header_idx: int | None = None
    for i, row in enumerate(rows):
        if _is_header_row(row):
            header_idx = i
            break
    if header_idx is None:
        return []
    super_row = rows[header_idx]
    sub_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else ()
    col_kinds: list[tuple[str, str | None] | None] = []
    for i in range(len(super_row)):
        top = _norm(super_row[i])
        sub = _norm(sub_row[i] if i < len(sub_row) else None)
        top_lo = top.lower()
        if i <= 1:
            col_kinds.append(None)
            continue
        if "semiannual" in top_lo:
            half = (
                "1st"
                if "1st" in sub.lower()
                else "2nd"
                if "2nd" in sub.lower()
                else None
            )
            col_kinds.append(("semiannual_index", half))
        elif "annual avg" in top_lo and not sub:
            col_kinds.append(("annual_index", None))
        elif "percent change from previous" in top_lo:
            if "dec" in sub.lower():
                col_kinds.append(("annual_pct_change_dec", None))
            elif "annual" in sub.lower():
                col_kinds.append(("annual_pct_change_avg", None))
            else:
                col_kinds.append(None)
        else:
            col_kinds.append(None)

    out: list[dict[str, Any]] = []
    row_index = 0
    for row in rows[header_idx + 2 :]:
        if not row or _is_blank_row(row):
            continue
        if _is_note_row(row):
            break
        year_cell = row[1] if len(row) > 1 else None
        try:
            year = int(str(year_cell).strip())
        except (TypeError, ValueError):
            continue
        for i, classification in enumerate(col_kinds):
            if classification is None or i >= len(row):
                continue
            kind, payload = classification
            value, value_string = _parse_value(row[i])
            if value is None and value_string is None:
                continue
            if kind == "semiannual_index":
                ref_date = dateType(year, 1 if payload == "1st" else 7, 1)
                record_kind = ("semiannual", payload, "index_value")
            elif kind == "annual_index":
                ref_date = dateType(year, 1, 1)
                record_kind = ("annual", None, "index_value")
            elif kind == "annual_pct_change_dec":
                ref_date = dateType(year, 12, 1)
                record_kind = ("monthly", None, "pct_change_12m")
            else:
                # kind is always one of the four classified above.
                ref_date = dateType(year, 1, 1)
                record_kind = ("annual", None, "pct_change_12m")
            row_index += 1
            record: dict[str, Any] = {
                "date": ref_date,
                "snapshot_date": snapshot_date,
                "label": "All items",
                "indent_level": 0,
                "frequency": record_kind[0],
                "index_value": None,
                "sa_index_value": None,
                "pct_change_1m_nsa": None,
                "pct_change_1m_sa": None,
                "pct_change_12m": None,
                "pct_change_window": None,
                "pct_change_window_start_date": None,
                "relative_importance": None,
                "pricing_schedule": None,
                "half": record_kind[1],
                "value_string": value_string,
                "footnote": None,
                "table_key": spec.key,
                "table_id": table_id,
                "table_name": table_name,
                "sheet": spec.sheet or "",
                "row_index": row_index,
            }
            record[record_kind[2]] = value
            out.append(record)
    return out


def parse_table(
    content: bytes,
    spec: _TableSpec,
    year: int,
    month: int,
) -> list[dict[str, Any]]:
    """Parse one CPI supplemental XLSX into long-form records."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_name = spec.sheet if spec.sheet else wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise OpenBBError(
            f"CPI XLSX for {spec.key} {year:04d}-{month:02d} has no sheet "
            f"named '{sheet_name}'."
        )
    raw_rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    snapshot_date = dateType(year, month, 1)
    table_id = f"cpi-{spec.key}-{year:04d}{month:02d}"
    table_name = f"BLS CPI Supplemental — {spec.label} ({year:04d}-{month:02d})."

    if spec.key == "historical-cpi-u-index":
        return _parse_index_values_sheet(
            raw_rows, spec, snapshot_date, table_id, table_name
        )
    if spec.key == "historical-cpi-u-averages":
        return _parse_index_averages_sheet(
            raw_rows, spec, snapshot_date, table_id, table_name
        )

    header_idx: int | None = None
    for i, row in enumerate(raw_rows):
        if _is_header_row(row):
            header_idx = i
            break
    if header_idx is None:
        raise OpenBBError(
            f"CPI XLSX for {spec.key} {year:04d}-{month:02d} has no 'Indent Level' header row."
        )

    header_row1 = raw_rows[header_idx]
    header_row2: tuple[Any, ...] | None = None
    if header_idx + 1 < len(raw_rows):
        nxt = raw_rows[header_idx + 1]
        if any(
            cell is not None and str(cell).strip()
            for j, cell in enumerate(nxt)
            if j >= 2 and (j >= len(header_row1) or header_row1[j] != nxt[j])
        ):
            header_row2 = nxt

    classifications = _classify_columns(header_row1, header_row2)
    data_start = header_idx + (2 if header_row2 is not None else 1)

    body_rows: list[tuple[Any, ...]] = []
    trailing_rows: list[tuple[Any, ...]] = []
    in_trailing = False
    for row in raw_rows[data_start:]:
        if _is_blank_row(row):
            if in_trailing:
                continue
            body_rows.append(row)
            continue
        if _is_note_row(row):
            in_trailing = True
        if in_trailing:
            trailing_rows.append(row)
        else:
            body_rows.append(row)

    footnotes = _parse_footnotes(trailing_rows)

    out: list[dict[str, Any]] = []
    row_index = 0
    for source_row in body_rows:
        if not source_row or all(c is None for c in source_row):
            continue
        indent_cell = source_row[0] if len(source_row) > 0 else None
        label_cell = source_row[1] if len(source_row) > 1 else None
        if isinstance(indent_cell, str) and not indent_cell.strip():
            indent_level: int | None = None
        elif isinstance(indent_cell, (int, float)):
            indent_level = int(indent_cell)
        else:
            indent_level = None
        if label_cell is None:
            continue
        label = _norm(label_cell)
        if not label:
            continue

        rel_imp: float | None = None
        pricing_schedule: str | None = None
        per_date: dict[dateType, dict[str, Any]] = {}

        for i, classification in enumerate(classifications):
            if classification is None or i >= len(source_row):
                continue
            kind, payload = classification
            value, value_string = _parse_value(source_row[i])
            if kind == "relative_importance":
                if value is not None:
                    rel_imp = value
                continue
            if kind == "pricing_schedule":
                if isinstance(source_row[i], str) and source_row[i].strip():
                    pricing_schedule = source_row[i].strip()
                continue
            if value is None and value_string is None:
                continue
            if kind == "pct_change_window":
                start, end = payload
                slot = per_date.setdefault(end, {})
                slot["pct_change_window"] = value
                slot["pct_change_window_start_date"] = start
                continue
            ref_date = payload
            if ref_date is None:
                continue
            slot = per_date.setdefault(ref_date, {})
            slot[kind] = value

        footnote_text = _resolve_footnotes(label, footnotes)
        if not per_date and rel_imp is None and pricing_schedule is None:
            continue

        if not per_date:
            row_index += 1
            out.append(
                {
                    "date": snapshot_date,
                    "snapshot_date": snapshot_date,
                    "label": label,
                    "indent_level": indent_level,
                    "frequency": "monthly",
                    "index_value": None,
                    "sa_index_value": None,
                    "pct_change_1m_nsa": None,
                    "pct_change_1m_sa": None,
                    "pct_change_12m": None,
                    "pct_change_window": None,
                    "pct_change_window_start_date": None,
                    "relative_importance": rel_imp,
                    "pricing_schedule": pricing_schedule,
                    "half": None,
                    "value_string": None,
                    "footnote": footnote_text,
                    "table_key": spec.key,
                    "table_id": table_id,
                    "table_name": table_name,
                    "sheet": sheet_name,
                    "row_index": row_index,
                }
            )
            continue

        for ref_date, measures in per_date.items():
            row_index += 1
            record = {
                "date": ref_date,
                "snapshot_date": snapshot_date,
                "label": label,
                "indent_level": indent_level,
                "frequency": "monthly",
                "index_value": None,
                "sa_index_value": None,
                "pct_change_1m_nsa": None,
                "pct_change_1m_sa": None,
                "pct_change_12m": None,
                "pct_change_window": None,
                "pct_change_window_start_date": None,
                "relative_importance": rel_imp,
                "pricing_schedule": pricing_schedule,
                "half": None,
                "value_string": None,
                "footnote": footnote_text,
                "table_key": spec.key,
                "table_id": table_id,
                "table_name": table_name,
                "sheet": sheet_name,
                "row_index": row_index,
            }
            record.update(measures)
            out.append(record)

    out.sort(key=lambda r: (-r["date"].toordinal(), r["row_index"]))
    return out
