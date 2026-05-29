"""BLS CES (Current Employment Statistics) analytical-table fetchers and parsers.

Source page: https://www.bls.gov/ces/data/analytical-tables.htm
The workbooks live at https://www.bls.gov/web/empsit/cesanatab{N}.xlsx and
``cesconfidenceintervals.xlsx`` and are refreshed in place each release, so the
URLs are stable (no monthly archive). Industry rows carry dot- or
non-breaking-space-indented hierarchy; numeric cells are space-padded strings
with a trailing ``*`` flagging a statistically significant change and ``.`` for
not-available.
"""

from __future__ import annotations

import calendar
import io
import re
from collections.abc import Callable
from datetime import date as dateType
from typing import Any

from openbb_core.app.model.abstract.error import OpenBBError

from openbb_bls.utils.constants import BLS_USER_AGENT

_BASE = "https://www.bls.gov/web/empsit"
_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
_XLSX_MAGIC = b"PK\x03\x04"

_MONTH_LOOKUP: dict[str, int] = {}
for _i, _name in enumerate(calendar.month_abbr):
    if _name:
        _MONTH_LOOKUP[_name.lower()] = _i
for _i, _name in enumerate(calendar.month_name):
    if _name:
        _MONTH_LOOKUP[_name.lower()] = _i
_MONTH_LOOKUP["sept"] = 9

_HDR_DATE_RE = re.compile(r"([A-Za-z]{3,9})\.?\s+(\d{4})")
_MON_YEAR_RE = re.compile(r"([A-Za-z]{3})-(\d{4})")
_CI_TABLE_RE = re.compile(r"TABLE\s+([A-C][12]?)", re.IGNORECASE)

CES_STEMS: dict[str, str] = {
    "t1": "cesanatab1",
    "t2": "cesanatab2",
    "t3a": "cesanatab3a",
    "t3b": "cesanatab3b",
    "t4": "cesanatab4",
    "t5": "cesanatab5",
    "t6": "cesanatab6",
    "t7": "cesanatab7",
    "ci": "cesconfidenceintervals",
}

CES_TITLES: dict[str, str] = {
    "t1": (
        "Table 1. Employment: normal seasonal movements, over-the-month changes, "
        "and tests of significance"
    ),
    "t2": (
        "Table 2. Detailed industry employment ranked by over-the-month changes, "
        "tests of significance, and prior 3-month average"
    ),
    "t3a": "Table 3A. Employment changes and tests of significance, seasonally adjusted",
    "t3b": (
        "Table 3B. Over-the-month employment changes compared with recent averages, "
        "seasonally adjusted"
    ),
    "t4": (
        "Table 4. Over-the-year employment changes and tests of significance, "
        "seasonally adjusted"
    ),
    "t5": (
        "Table 5. Average weekly hours and average hourly earnings of all employees: "
        "normal seasonal movements, over-the-month changes, and tests of significance"
    ),
    "t6": (
        "Table 6. Over-the-month and over-the-year changes in aggregate weekly hours "
        "and payrolls of all employees, seasonally adjusted"
    ),
    "t7": (
        "Table 7. Most recent industry-specific employment peak and trough, and "
        "changes from peak and trough to current employment, seasonally adjusted"
    ),
    "ci": (
        "Tables A-C2. 90 percent confidence intervals for employment, hours, "
        "overtime hours, and earnings"
    ),
}

_CI_SHEET_META: dict[str, tuple[str, str | None]] = {
    "CI Industry Employment Chang": ("Employment", None),
    "CI AWH_AE": ("Average weekly hours", "All employees"),
    "CI AHE_AE": ("Average hourly earnings", "All employees"),
    "CI AWH_PE": ("Average weekly hours", "Production employees"),
    "CI AHE_PE": ("Average hourly earnings", "Production employees"),
    "CI AOT_AE": ("Average overtime hours", "All employees"),
    "CI AOT_PE": ("Average overtime hours", "Production employees"),
}


def fetch_table_xlsx(stem: str) -> bytes:
    """Download one CES analytical-table workbook from bls.gov."""
    import requests

    url = f"{_BASE}/{stem}.xlsx"
    resp = requests.get(url, headers=_HEADERS, timeout=120)
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")
    if not resp.content.startswith(_XLSX_MAGIC):
        raise OpenBBError(
            f"BLS returned non-XLSX content for {url} (length="
            f"{len(resp.content)}, head={resp.content[:8]!r})."
        )
    return resp.content


def _norm(value: Any) -> str:
    """Collapse newlines / non-breaking spaces in a header or label cell."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def _to_value(cell: Any) -> tuple[float | None, bool, str | None]:
    """Coerce a CES numeric cell into ``(value, significant, raw_text)``."""
    if cell is None:
        return None, False, None
    if isinstance(cell, bool):
        return None, False, None
    if isinstance(cell, (int, float)):
        return float(cell), False, None
    text = str(cell).replace("\xa0", " ").strip()
    significant = text.endswith("*")
    if significant:
        text = text[:-1].strip()
    if text in ("", ".", "-", "(NA)", "N.A.", "NA"):
        return None, significant, None
    try:
        return float(text.replace(",", "")), significant, None
    except ValueError:
        return None, significant, text


def _to_mon_year(cell: Any) -> dateType | None:
    """Parse a CES peak/trough ``MON-YYYY`` cell into a first-of-month date."""
    if cell is None:
        return None
    text = str(cell).replace("\xa0", " ").strip().rstrip("*").strip()
    match = _MON_YEAR_RE.match(text)
    if match is None:
        return None
    month = _MONTH_LOOKUP.get(match.group(1).lower())
    if month is None:
        return None
    return dateType(int(match.group(2)), month, 1)


def _clean_label(raw: Any) -> tuple[int, str]:
    """Split a CES industry label into ``(indent_level, clean_label)``.

    Tables 1/3A/3B/4/7 prefix labels with leading dots (one per hierarchy
    level); the confidence-interval sheets use leading non-breaking spaces.
    """
    text = str(raw) if raw is not None else ""
    dots = 0
    while dots < len(text) and text[dots] == ".":
        dots += 1
    if dots:
        return dots, _norm(text[dots:])
    spaces = 0
    while spaces < len(text) and text[spaces] in (" ", "\xa0", "\t"):
        spaces += 1
    return spaces, _norm(text[spaces:])


def _reference_date(header_rows: list[tuple[Any, ...]]) -> dateType | None:
    """Return the latest ``Mon. YYYY`` reference month found in the header block."""
    best: dateType | None = None
    for row in header_rows:
        for cell in row or ():
            if not isinstance(cell, str):
                continue
            for mon, year in _HDR_DATE_RE.findall(cell):
                month = _MONTH_LOOKUP.get(mon.strip().lower().rstrip("."))
                if month is None:
                    continue
                candidate = dateType(int(year), month, 1)
                if best is None or candidate > best:
                    best = candidate
    return best


def _load_rows(content: bytes, sheet: str | None = None) -> list[tuple[Any, ...]]:
    """Load one worksheet's rows (first sheet unless ``sheet`` is given)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_name = sheet if sheet is not None else wb.sheetnames[0]
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    return rows


def _find_data_start(rows: list[tuple[Any, ...]]) -> int:
    """Return the index of the first data row (after the Industry/Rank header)."""
    for i, row in enumerate(rows):
        if row and _norm(row[0]).lower() in ("industry", "rank"):
            return i + 1
    raise OpenBBError(
        "CES analytical table is missing the expected 'Industry'/'Rank' header row."
    )


def _is_footnote(raw0: Any) -> bool:
    """Return True when a row's first cell is a trailing footnote marker."""
    return _norm(raw0).startswith("*")


def _title(rows: list[tuple[Any, ...]], fallback: str) -> str:
    """Return the cleaned workbook title from row 0, else a fallback."""
    if rows and rows[0]:
        text = _norm(rows[0][0])
        if text:
            return text
    return fallback


def parse_table1(content: bytes) -> dict[str, Any]:
    """Parse CES Table 1 — NSA + SA employment changes and significance."""
    rows = _load_rows(content)
    start = _find_data_start(rows)
    ref = _reference_date(rows[:start])
    title = _title(rows, CES_TITLES["t1"])
    out: list[dict[str, Any]] = []
    for row in rows[start:]:
        raw0 = row[0] if row else None
        if not _norm(raw0):
            continue
        if _is_footnote(raw0):
            break
        level, label = _clean_label(raw0)
        nsm, _, _ = _to_value(row[1])
        change_nsa, _, _ = _to_value(row[2])
        change_sa, sig_sa, _ = _to_value(row[3])
        min_sig, _, _ = _to_value(row[4])
        out.append(
            {
                "reference_date": ref,
                "indent_level": level,
                "label": label,
                "normal_seasonal_movement": nsm,
                "change_nsa": change_nsa,
                "change_sa": change_sa,
                "change_sa_significant": sig_sa,
                "minimum_significant_change": min_sig,
                "row_index": len(out) + 1,
                "table_id": "ces-anatab1",
                "table_title": title,
            }
        )
    return {"rows": out, "table_id": "ces-anatab1", "table_title": title}


def parse_table2(content: bytes) -> dict[str, Any]:
    """Parse CES Table 2 — detailed industries ranked by over-the-month change."""
    rows = _load_rows(content)
    start = _find_data_start(rows)
    ref = _reference_date(rows[:start])
    title = _title(rows, CES_TITLES["t2"])
    out: list[dict[str, Any]] = []
    for row in rows[start:]:
        raw0 = row[0] if row else None
        if not _norm(raw0):
            continue
        if _is_footnote(raw0):
            break
        change_sa, sig_sa, _ = _to_value(row[3])
        min_sig, _, _ = _to_value(row[4])
        prior_avg, _, _ = _to_value(row[5])
        out.append(
            {
                "reference_date": ref,
                "rank": _norm(raw0),
                "label": _norm(row[1]),
                "naics_code": _norm(row[2]) or None,
                "change_sa": change_sa,
                "change_sa_significant": sig_sa,
                "minimum_significant_change": min_sig,
                "prior_3month_average": prior_avg,
                "row_index": len(out) + 1,
                "table_id": "ces-anatab2",
                "table_title": title,
            }
        )
    return {"rows": out, "table_id": "ces-anatab2", "table_title": title}


def _parse_multi_change(
    content: bytes, key: str, fields: tuple[str, ...]
) -> dict[str, Any]:
    """Parse Table 3A/3B — six change columns over an industry hierarchy.

    ``fields`` names the six value columns; Table 3A flags significance on every
    column while Table 3B (recent averages) carries no significance markers.
    """
    rows = _load_rows(content)
    start = _find_data_start(rows)
    ref = _reference_date(rows[:start])
    table_id = f"ces-anatab{key[1:]}"
    title = _title(rows, CES_TITLES[key])
    flag_significance = key == "t3a"
    out: list[dict[str, Any]] = []
    for row in rows[start:]:
        raw0 = row[0] if row else None
        if not _norm(raw0):
            continue
        if _is_footnote(raw0):
            break
        level, label = _clean_label(raw0)
        record: dict[str, Any] = {
            "reference_date": ref,
            "indent_level": level,
            "label": label,
            "row_index": len(out) + 1,
            "table_id": table_id,
            "table_title": title,
        }
        for offset, field in enumerate(fields, start=1):
            value, significant, _ = _to_value(row[offset])
            record[field] = value
            if flag_significance:
                record[f"{field}_significant"] = significant
        out.append(record)
    return {"rows": out, "table_id": table_id, "table_title": title}


def parse_table3a(content: bytes) -> dict[str, Any]:
    """Parse CES Table 3A — SA employment changes over multiple horizons."""
    return _parse_multi_change(
        content,
        "t3a",
        (
            "otm_change_latest",
            "otm_change_prior_1",
            "otm_change_prior_2",
            "current_3month_change",
            "current_6month_change",
            "current_12month_change",
        ),
    )


def parse_table3b(content: bytes) -> dict[str, Any]:
    """Parse CES Table 3B — over-the-month changes vs. recent averages."""
    return _parse_multi_change(
        content,
        "t3b",
        (
            "otm_change_latest",
            "otm_change_prior_1",
            "otm_change_prior_2",
            "prior_3month_average",
            "prior_6month_average",
            "prior_12month_average",
        ),
    )


def parse_table4(content: bytes) -> dict[str, Any]:
    """Parse CES Table 4 — over-the-year employment changes and significance."""
    rows = _load_rows(content)
    start = _find_data_start(rows)
    ref = _reference_date(rows[:start])
    title = _title(rows, CES_TITLES["t4"])
    out: list[dict[str, Any]] = []
    for row in rows[start:]:
        raw0 = row[0] if row else None
        if not _norm(raw0):
            continue
        if _is_footnote(raw0):
            break
        level, label = _clean_label(raw0)
        number, sig, _ = _to_value(row[1])
        percent, _, _ = _to_value(row[2])
        min_sig, _, _ = _to_value(row[3])
        out.append(
            {
                "reference_date": ref,
                "indent_level": level,
                "label": label,
                "oty_change_number": number,
                "oty_change_number_significant": sig,
                "oty_change_percent": percent,
                "minimum_significant_change": min_sig,
                "row_index": len(out) + 1,
                "table_id": "ces-anatab4",
                "table_title": title,
            }
        )
    return {"rows": out, "table_id": "ces-anatab4", "table_title": title}


def parse_table5(content: bytes) -> dict[str, Any]:
    """Parse CES Table 5 — average weekly hours + hourly earnings (two blocks)."""
    rows = _load_rows(content)
    start = _find_data_start(rows)
    ref = _reference_date(rows[:start])
    title = _title(rows, CES_TITLES["t5"])
    measure = "Average weekly hours"
    out: list[dict[str, Any]] = []
    for row in rows[start:]:
        raw0 = row[0] if row else None
        label_norm = _norm(raw0)
        if not label_norm:
            continue
        if _is_footnote(raw0):
            break
        lowered = label_norm.lower()
        if lowered in ("average weekly hours", "average hourly earnings"):
            measure = label_norm
            continue
        if lowered == "industry" or _norm(row[1]).lower().startswith("normal"):
            continue
        normal, _, _ = _to_value(row[1])
        change_nsa, _, _ = _to_value(row[2])
        change_sa, sig_sa, _ = _to_value(row[3])
        min_sig, _, _ = _to_value(row[4])
        out.append(
            {
                "reference_date": ref,
                "measure": measure,
                "indent_level": 0,
                "label": label_norm,
                "normal_seasonal_movement": normal,
                "change_nsa": change_nsa,
                "change_sa": change_sa,
                "change_sa_significant": sig_sa,
                "minimum_significant_change": min_sig,
                "row_index": len(out) + 1,
                "table_id": "ces-anatab5",
                "table_title": title,
            }
        )
    return {"rows": out, "table_id": "ces-anatab5", "table_title": title}


def parse_table6(content: bytes) -> dict[str, Any]:
    """Parse CES Table 6 — aggregate weekly hours + payroll changes (two blocks)."""
    rows = _load_rows(content)
    start = _find_data_start(rows) - 1  # include the first "Industry" header row
    ref = _reference_date(rows[: start + 1])
    title = _title(rows, CES_TITLES["t6"])
    measure = "Aggregate weekly hours"
    out: list[dict[str, Any]] = []
    for row in rows[start:]:
        raw0 = row[0] if row else None
        label = _norm(raw0)
        if not label:
            continue
        if _is_footnote(raw0):
            break
        if label.lower() == "industry":
            measure = (
                "Aggregate weekly payrolls"
                if "payroll" in _norm(row[1]).lower()
                else "Aggregate weekly hours"
            )
            continue
        aggregate, _, _ = _to_value(row[1])
        otm_number, _, _ = _to_value(row[2])
        otm_percent, _, _ = _to_value(row[3])
        oty_number, _, _ = _to_value(row[4])
        oty_percent, _, _ = _to_value(row[5])
        out.append(
            {
                "reference_date": ref,
                "measure": measure,
                "indent_level": 0,
                "label": label,
                "aggregate_value": aggregate,
                "otm_change_number": otm_number,
                "otm_change_percent": otm_percent,
                "oty_change_number": oty_number,
                "oty_change_percent": oty_percent,
                "row_index": len(out) + 1,
                "table_id": "ces-anatab6",
                "table_title": title,
            }
        )
    return {"rows": out, "table_id": "ces-anatab6", "table_title": title}


def parse_table7(content: bytes) -> dict[str, Any]:
    """Parse CES Table 7 — most recent employment peak / trough and changes."""
    rows = _load_rows(content)
    start = _find_data_start(rows)
    ref = _reference_date(rows[:start])
    title = _title(rows, CES_TITLES["t7"])
    out: list[dict[str, Any]] = []
    for row in rows[start:]:
        raw0 = row[0] if row else None
        if not _norm(raw0):
            continue
        if _is_footnote(raw0):
            break
        level, label = _clean_label(raw0)
        employment, _, _ = _to_value(row[1])
        peak_employment, _, _ = _to_value(row[3])
        trough_employment, _, _ = _to_value(row[5])
        change_peak, _, _ = _to_value(row[6])
        change_trough, _, _ = _to_value(row[7])
        out.append(
            {
                "reference_date": ref,
                "indent_level": level,
                "label": label,
                "current_employment": employment,
                "peak_date": _to_mon_year(row[2]),
                "peak_employment": peak_employment,
                "trough_date": _to_mon_year(row[4]),
                "trough_employment": trough_employment,
                "change_from_peak": change_peak,
                "change_from_trough": change_trough,
                "row_index": len(out) + 1,
                "table_id": "ces-anatab7",
                "table_title": title,
            }
        )
    return {"rows": out, "table_id": "ces-anatab7", "table_title": title}


def parse_confidence_intervals(content: bytes) -> dict[str, Any]:
    """Parse the CES confidence-interval workbook (Tables A-C2, seven sheets)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    row_index = 0
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue
        title = _title(rows, "")
        ci_match = _CI_TABLE_RE.search(title)
        ci_table = ci_match.group(1).upper() if ci_match else sheet_name
        measure, employee_group = _CI_SHEET_META.get(sheet_name, (sheet_name, None))
        start = _find_data_start(rows)
        ncols = len(rows[start - 1])
        wide = ncols >= 7
        for row in rows[start:]:
            raw0 = row[0] if row else None
            if not _norm(raw0):
                continue
            if _is_footnote(raw0):
                break
            level, label = _clean_label(raw0)
            row_index += 1
            if wide:
                first, _, _ = _to_value(row[1])
                second, _, _ = _to_value(row[2])
                third, _, _ = _to_value(row[3])
                three, _, _ = _to_value(row[4])
                six, _, _ = _to_value(row[5])
                twelve, _, _ = _to_value(row[6])
            else:
                first, _, _ = _to_value(row[1])
                second = third = None
                three, _, _ = _to_value(row[2])
                six, _, _ = _to_value(row[3])
                twelve, _, _ = _to_value(row[4])
            out.append(
                {
                    "ci_table": ci_table,
                    "measure": measure,
                    "employee_group": employee_group,
                    "indent_level": level,
                    "label": label,
                    "ci_1month_first": first,
                    "ci_1month_second": second,
                    "ci_1month_third": third,
                    "ci_3month": three,
                    "ci_6month": six,
                    "ci_12month": twelve,
                    "row_index": row_index,
                    "table_id": "ces-confidence-intervals",
                    "table_title": CES_TITLES["ci"],
                }
            )
    wb.close()
    return {
        "rows": out,
        "table_id": "ces-confidence-intervals",
        "table_title": CES_TITLES["ci"],
    }


_PARSERS: dict[str, Callable[[bytes], dict[str, Any]]] = {
    "t1": parse_table1,
    "t2": parse_table2,
    "t3a": parse_table3a,
    "t3b": parse_table3b,
    "t4": parse_table4,
    "t5": parse_table5,
    "t6": parse_table6,
    "t7": parse_table7,
    "ci": parse_confidence_intervals,
}


def parse_ces(content: bytes, key: str) -> dict[str, Any]:
    """Dispatch to the parser for one CES analytical-table key."""
    return _PARSERS[key](content)
