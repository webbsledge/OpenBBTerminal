"""BLS PPI detailed report tables."""

from __future__ import annotations

import calendar
import re
from datetime import date as dateType
from typing import Any

from openbb_core.app.model.abstract.error import OpenBBError
from openbb_core.provider.abstract.data import Data
from openbb_core.provider.abstract.fetcher import Fetcher
from openbb_core.provider.abstract.query_params import QueryParams
from openbb_core.provider.utils.errors import EmptyDataError
from pydantic import ConfigDict, Field, field_serializer

from openbb_bls.utils.constants import BLS_USER_AGENT

_TABLE_TITLES: dict[int, str] = {
    1: (
        "Producer Price Index percentage changes and weights for Final Demand-"
        "Intermediate Demand (FD-ID) aggregation system, seasonally adjusted."
    ),
    2: (
        "Producer Price Index percentage changes for selected Final Demand-"
        "Intermediate Demand (FD-ID) aggregations, seasonally adjusted."
    ),
    3: (
        "Producer Price Indexes for Final Demand-Intermediate Demand (FD-ID) "
        "aggregations, seasonally adjusted."
    ),
    4: (
        "Producer Price Index percentage changes and weights for Final Demand-"
        "Intermediate Demand (FD-ID) aggregation system, not seasonally adjusted."
    ),
    5: (
        "Producer Price Indexes for Final Demand-Intermediate Demand (FD-ID) "
        "aggregations, not seasonally adjusted."
    ),
    6: (
        "Producer Price Index percentage changes for selected commodity "
        "groupings, seasonally adjusted."
    ),
    7: (
        "Producer Price Index percentage changes for commodity and service "
        "groupings and individual items, not seasonally adjusted."
    ),
    8: (
        "Producer Price Indexes and percentage changes for selected commodity "
        "groupings, seasonally adjusted."
    ),
    9: (
        "Producer Price Indexes for commodity and service groupings and "
        "individual items, not seasonally adjusted."
    ),
    10: (
        "Producer Price Indexes and percentage changes for the net output of "
        "selected NAICS industry sector and 3-digit subsector groups, not "
        "seasonally adjusted."
    ),
    11: (
        "Producer Price Indexes for the net output of selected NAICS industries "
        "and their products, not seasonally adjusted."
    ),
    12: (
        "Producer Price Index percentage changes for the net output of selected "
        "NAICS industries and their products, not seasonally adjusted."
    ),
    13: (
        "Producer Price Indexes and percentage changes for special commodity "
        "groupings, not seasonally adjusted."
    ),
    14: (
        "Producer Price Indexes and percentage changes for inputs to "
        "construction industries and selected construction outputs, not "
        "seasonally adjusted."
    ),
}

_TABLE_GLOSSARY = "\n".join(f"  {n:>2}: {title}" for n, title in _TABLE_TITLES.items())

_XLSX_EARLIEST = dateType(2022, 1, 1)
_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
_HIDE: dict[str, Any] = {"x-widget_config": {"hide": True}}


def _pct(name: str) -> dict[str, Any]:
    """Percent-column config: headerName + numeric type + percent formatter + greenRed render."""
    return {
        "x-widget_config": {
            "headerName": name,
            "cellDataType": "number",
            "formatterFn": "percent",
            "renderFn": "greenRed",
        }
    }


def _num(name: str) -> dict[str, Any]:
    """Numeric (non-percent) column config: headerName + numeric type."""
    return {
        "x-widget_config": {
            "headerName": name,
            "cellDataType": "number",
        }
    }


class BlsPpiDetailedReportQueryParams(QueryParams):
    """BLS PPI Detailed Report Query Parameters."""

    __json_schema_extra__ = {
        "table": {
            "x-widget_config": {
                "options": [
                    {"label": f"Table {n}. {title}", "value": n}
                    for n, title in _TABLE_TITLES.items()
                ],
                "description": "Which of the 14 tables published in the monthly PPI Detailed Report "
                "to return. Each table is a different cut of the same release",
                "style": {"popupWidth": 950},
            }
        },
    }

    date: dateType | None = Field(
        default=None,
        description="Month of the Detailed Report to fetch; None resolves to the most recent release.",
    )
    table: int = Field(
        default=1,
        ge=1,
        le=14,
        description="Which of the 14 tables in the monthly PPI Detailed Report to return.",
    )


class BlsPpiDetailedReportData(Data):
    """One ``(series, date)`` row from a BLS PPI detailed report table."""

    model_config = ConfigDict(
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS PPI Detailed Report Tables",
                "$.description": ("Producer Price Index detailed report tables."),
                "$.gridData": {"w": 30, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "PPI",
            }
        }
    )

    date: dateType = Field(
        description="First day of the calendar month this row's values apply to.",
    )
    code: str | None = Field(
        default=None,
        description="Concatenated series identifier for the row.",
    )
    group_code: str | None = Field(
        default=None,
        description="BLS 'Group code' from the source XLSX.",
        json_schema_extra=_HIDE,
    )
    item_code: str | None = Field(
        default=None,
        description="BLS 'Item code' from the source XLSX.",
        json_schema_extra=_HIDE,
    )
    industry_code: str | None = Field(
        default=None,
        description="NAICS industry code from the source XLSX.",
        json_schema_extra=_HIDE,
    )
    product_code: str | None = Field(
        default=None,
        description="Product breakdown code of an industry's net output.",
        json_schema_extra=_HIDE,
    )
    label: str | None = Field(
        default=None,
        description="Human-readable series title from the source XLSX.",
        json_schema_extra={
            "x-widget_config": {
                "renderFn": "hoverCard",
                "renderFnParams": {
                    "hoverCard": {
                        "cellField": "value",
                        "markdown": "{footnote}",
                    }
                },
            }
        },
    )
    footnote: str | None = Field(
        default=None,
        description="Combined BLS footnote text for any markers on the label.",
        json_schema_extra=_HIDE,
    )

    @field_serializer("label", when_used="json")
    def _serialize_label_with_footnote(self, value: str | None) -> Any:
        """Wrap the label with its footnote text for the hover-card renderer."""
        if value and self.footnote:
            return {"value": value, "footnote": self.footnote}
        return value

    level: int | None = Field(
        default=None,
        description="Row indent depth in the source XLSX hierarchy.",
    )
    index_value: float | None = Field(
        default=None,
        description="Price-index value at the date.",
        json_schema_extra=_num("Index Value"),
    )
    pct_change_1m: float | None = Field(
        default=None,
        description="One-month percent change in the index ending at the date.",
        json_schema_extra=_pct("Pct Change 1M"),
    )
    pct_change_12m: float | None = Field(
        default=None,
        description="Twelve-month percent change in the index ending at the date.",
        json_schema_extra=_pct("Pct Change 12M"),
    )
    relative_importance: float | None = Field(
        default=None,
        description="Relative-importance weight of this series within its parent aggregate, as a percentage.",
        json_schema_extra=_pct("Relative Importance"),
    )
    other_index_base: str | None = Field(
        default=None,
        description="Alternative index base period for this row when BLS overrides the table-wide default.",
        json_schema_extra=_HIDE,
    )
    seasonally_adjusted: bool | None = Field(
        default=None,
        description="Whether the values on this row are seasonally adjusted.",
        json_schema_extra=_HIDE,
    )
    table_id: str = Field(
        description="Source table identifier.",
        json_schema_extra=_HIDE,
    )
    table_name: str = Field(
        description="Full BLS table title.",
        json_schema_extra=_HIDE,
    )
    table_number: int = Field(
        description="Which of the 14 numbered tables in the monthly PPI Detailed Report this row belongs to.",
        json_schema_extra=_HIDE,
    )
    release_period: str = Field(
        description="Reporting month of the source release.",
        json_schema_extra=_HIDE,
    )


class BlsPpiDetailedReportFetcher(
    Fetcher[BlsPpiDetailedReportQueryParams, list[BlsPpiDetailedReportData]]
):
    """BLS PPI Detailed Report Fetcher."""

    require_credentials = False

    @staticmethod
    def transform_query(params: dict[str, Any]) -> BlsPpiDetailedReportQueryParams:
        """Validate and coerce the query."""
        return BlsPpiDetailedReportQueryParams(**params)

    @staticmethod
    def extract_data(
        query: BlsPpiDetailedReportQueryParams,
        credentials: dict[str, str] | None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Download a monthly detailed-report XLSX and parse one table."""
        if query.date is None:
            year, month, content = _discover_latest()
        else:
            if query.date < _XLSX_EARLIEST:
                raise OpenBBError(
                    "BLS publishes the PPI Detailed Report as XLSX only from "
                    f"January 2022 onward; requested date {query.date.isoformat()} "
                    "is too early."
                )
            year, month = query.date.year, query.date.month
            content = _fetch_month_xlsx(year, month)
            if content is None:
                raise OpenBBError(
                    f"BLS has no PPI Detailed Report XLSX for "
                    f"{calendar.month_name[month]} {year}."
                )
        return _parse_table(content, query.table, year, month)

    @staticmethod
    def transform_data(
        query: BlsPpiDetailedReportQueryParams,
        data: dict[str, Any],
        **kwargs: Any,
    ) -> list[BlsPpiDetailedReportData]:
        """Coerce parsed rows into ``BlsPpiDetailedReportData``."""
        rows = data.get("rows", [])
        if not rows:
            raise EmptyDataError(
                f"No rows parsed from BLS PPI Detailed Report "
                f"'{data.get('table_id', '?')}'."
            )
        return [BlsPpiDetailedReportData.model_validate(r) for r in rows]


def _month_url(year: int, month: int) -> str:
    """Build the canonical XLSX URL for a given ``year`` and ``month``."""
    month_name = calendar.month_name[month].lower()
    return (
        f"https://www.bls.gov/ppi/detailed-report/"
        f"ppi-detailed-report-{month_name}-{year}.xlsx"
    )


_XLSX_MAGIC = b"PK\x03\x04"


def _fetch_month_xlsx(year: int, month: int) -> bytes | None:
    """Fetch one monthly XLSX, returning ``None`` when the file isn't published."""
    import requests

    url = _month_url(year, month)
    resp = requests.get(url, headers=_HEADERS, timeout=60, allow_redirects=False)
    if resp.status_code in (301, 302, 303, 307, 308, 404):
        return None
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")
    if not resp.content.startswith(_XLSX_MAGIC):
        return None
    return resp.content


def _discover_latest() -> tuple[int, int, bytes]:
    """Walk back from today's month until a 200 lands, returning the XLSX bytes."""
    today = dateType.today()
    year, month = today.year, today.month
    for _ in range(6):
        content = _fetch_month_xlsx(year, month)
        if content is not None:
            return year, month, content
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    raise OpenBBError(
        "Could not locate a recent BLS PPI Detailed Report XLSX within the last "
        "six months — the publication URL pattern may have changed."
    )


def _parse_table(
    content: bytes, table_number: int, year: int, month: int
) -> dict[str, Any]:
    """Open an XLSX from bytes and emit one row per ``(series, date)``."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_name = f"Table {table_number}"
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise OpenBBError(
            f"BLS PPI Detailed Report for {calendar.month_name[month]} {year} "
            f"does not contain sheet '{sheet_name}'."
        )
    sheet = wb[sheet_name]
    raw_rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    release_period = f"{calendar.month_name[month]} {year}"
    table_id = f"ppi-dr-{year}-{month:02d}-t{table_number}"
    table_name = (
        f"PPI Detailed Report {release_period} — "
        f"Table {table_number}: {_TABLE_TITLES[table_number]}"
    )

    classifications, data_start = _detect_columns(raw_rows)
    if not classifications:  # pragma: no cover -- every BLS Detailed Report has headers
        return {"rows": [], "table_id": table_id, "table_name": table_name}

    footnotes = _parse_table_footnotes(raw_rows[data_start:])

    out: list[dict[str, Any]] = []
    code_stack: dict[int, dict[str, str | None]] = {}

    for source_row in raw_rows[data_start:]:
        raw_records = _row_to_records(source_row, classifications, year, month)
        if not raw_records:
            continue
        indent = raw_records[0]["level"] or 0
        own_industry = raw_records[0].get("industry_code")
        own_group = raw_records[0].get("group_code")

        effective_industry = (
            own_industry
            if own_industry
            else _inherit_from_stack(code_stack, indent, "industry_code")
        )
        effective_group = (
            own_group
            if own_group
            else _inherit_from_stack(code_stack, indent, "group_code")
        )

        code_stack[indent] = {
            "industry_code": effective_industry,
            "group_code": effective_group,
        }
        for lvl in [k for k in code_stack if k > indent]:
            del code_stack[lvl]

        footnote_text = _resolve_label_footnotes(raw_records[0].get("label"), footnotes)

        for record in raw_records:
            record["industry_code"] = effective_industry
            record["group_code"] = effective_group
            record["code"] = _build_code(record)
            record["footnote"] = footnote_text
            record["table_id"] = table_id
            record["table_name"] = table_name
            record["table_number"] = table_number
            record["release_period"] = release_period
            out.append(record)

    out.sort(
        key=lambda r: (
            -r["date"].toordinal(),
            r.get("code") or "",
            r.get("level") or 0,
            r.get("label") or "",
        )
    )

    return {"rows": out, "table_id": table_id, "table_name": table_name}


def _inherit_from_stack(
    stack: dict[int, dict[str, str | None]], indent: int, key: str
) -> str | None:
    """Return the most recent value of ``key`` at any indent strictly shallower than ``indent``."""
    for lvl in range(indent - 1, -1, -1):
        entry = stack.get(lvl)
        if entry and entry.get(key):
            return entry[key]
    return None


_FOOTNOTES_HEADER_RE = re.compile(r"^footnotes?\b", re.IGNORECASE)
_FOOTNOTE_DEF_RE = re.compile(r"^\(([^)]+)\)\s*(.*)$")
_LABEL_REF_RE = re.compile(r"\(([^)]+)\)")


def _parse_table_footnotes(trailing_rows: list[tuple]) -> dict[str, str]:
    """Build ``{'(3)': 'PPI defines …'}`` from the table's footnote legend."""
    found_header = False
    out: dict[str, str] = {}
    for row in trailing_rows:
        if not row or len(row) < 2:
            continue
        cell = row[1]
        if not isinstance(cell, str):
            continue
        text = cell.strip()
        if not text:
            continue
        if _FOOTNOTES_HEADER_RE.match(text):
            found_header = True
            continue
        if not found_header:
            continue
        match = _FOOTNOTE_DEF_RE.match(text)
        if match is None:
            continue
        marker = f"({match.group(1)})"
        body = match.group(2).strip()
        out[marker] = body or text
    return out


def _resolve_label_footnotes(
    label: str | None, footnotes: dict[str, str]
) -> str | None:
    """Resolve any ``(N)`` footnote markers in ``label`` to their text."""
    if not label or not footnotes:
        return None
    matched: list[str] = []
    seen: set[str] = set()

    for match in _LABEL_REF_RE.finditer(label):
        marker = match.group(0)

        if marker not in footnotes or marker in seen:
            continue

        seen.add(marker)
        matched.append(footnotes[marker])

    if not matched:
        return None

    return "\n\n".join(matched)


def _build_code(record: dict[str, Any]) -> str | None:
    """Build the canonical concatenated ``code`` from the row's component codes."""
    product = record.get("product_code")
    industry = record.get("industry_code")
    group = record.get("group_code")
    item = record.get("item_code")

    if product:
        return product
    if group and item:
        return f"{group}{item}"
    if industry:
        return industry
    if group:
        return group
    if item:
        return item

    return None


_FOOTNOTE_MARKER_RE = re.compile(r"\s*\(\w+\)\s*$")


def _clean_header(text: Any) -> str:
    """Collapse newlines and runs of whitespace in a header cell."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\n", " ").split()).strip()


def _strip_footnote_marker(text: str) -> str:
    """Remove a trailing ``(N)``/``(p)`` footnote marker from a BLS header label."""
    return _FOOTNOTE_MARKER_RE.sub("", text).strip()


_MONTH_LOOKUP: dict[str, int] = {}
for _m in range(1, 13):
    _MONTH_LOOKUP[calendar.month_abbr[_m].lower()] = _m
    _MONTH_LOOKUP[calendar.month_name[_m].lower()] = _m
_MONTH_LOOKUP["sept"] = 9


def _month_from_token(token: str) -> int | None:
    """Map a BLS month token (``'Apr'``, ``'Apr.'``, ``'Sept'``, …) to ``1..12``."""
    return _MONTH_LOOKUP.get(token.strip().lower().rstrip("."))


_PERIOD_RE_RANGE = re.compile(r"\bto\s+([A-Za-z]+)\.?\s*(\d{4})?", re.IGNORECASE)
_PERIOD_RE_SINGLE = re.compile(r"\b([A-Za-z]+)\.?\s+(\d{4})\b", re.IGNORECASE)


def _period_label_to_date(  # noqa: PLR0911
    label: str | None,
    report_year: int,
    report_month: int,
    is_range: bool,
) -> dateType | None:
    """Parse a BLS period label into the end-date of the period."""
    if not label:
        return None

    text = _strip_footnote_marker(label)

    if is_range:
        m = _PERIOD_RE_RANGE.search(text)

        if not m:
            return None

        month_token, year_token = m.groups()
        month_num = _month_from_token(month_token)

        if month_num is None:
            return None

        if year_token is not None:
            return dateType(int(year_token), month_num, 1)

        year = report_year if month_num <= report_month else report_year - 1

        return dateType(year, month_num, 1)

    m = _PERIOD_RE_SINGLE.search(text)

    if not m:
        return None

    month_token, year_token = m.groups()
    month_num = _month_from_token(month_token)

    if month_num is None:
        return None

    return dateType(int(year_token), month_num, 1)


def _detect_columns(
    rows: list[tuple],
) -> tuple[list[tuple[str, str | None]], int]:
    """Locate the two-row header block and classify each column."""
    for idx, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        if _clean_header(row[0]).lower() != "indent level":
            continue
        super_row = row
        sub_row = rows[idx + 1] if idx + 1 < len(rows) else ()
        return _classify_header_pair(super_row, sub_row), idx + 2
    return [], 0


def _classify_header_pair(
    super_row: tuple, sub_row: tuple
) -> list[tuple[str, str | None]]:
    """Classify columns from one super/sub header pair."""
    n = max(len(super_row), len(sub_row))
    filled_super: list[str] = []
    last = ""
    for i in range(n):
        cell = _clean_header(super_row[i] if i < len(super_row) else None)
        if cell:
            last = cell
        filled_super.append(last)

    classifications: list[tuple[str, str | None]] = []
    for i in range(n):
        s = filled_super[i]
        sub = _clean_header(sub_row[i] if i < len(sub_row) else None)
        s_lo = s.lower()
        sub_lo = sub.lower()

        kind: str
        period: str | None = None

        if i == 0 or s_lo == "indent level":
            kind = "indent"
        elif s_lo in {"grouping", "title", "title(1)"}:
            kind = "label"
        elif s_lo == "group code" or sub_lo == "group code":
            kind = "group_code"
        elif s_lo == "item code" or sub_lo == "item code":
            kind = "item_code"
        elif "industry code" in s_lo or "industry code" in sub_lo:
            kind = "industry_code"
        elif "product code" in s_lo or "product code" in sub_lo:
            kind = "product_code"
        elif "other index base" in s_lo or s_lo == "index base":
            kind = "other_index_base"
        elif "relative importance" in s_lo:
            kind = "relative_importance"
            period = _strip_footnote_marker(
                _FOOTNOTE_MARKER_RE.sub("", s)
                .replace("Relative Importance", "")
                .replace("relative importance", "")
                .strip()
            )
        elif "12-month percent change" in s_lo:
            kind = "pct_change_12m"
            period = _strip_footnote_marker(sub) if sub else None
        elif "1-month percent change" in s_lo:
            kind = (
                "pct_change_1m_sa"
                if "seasonally adjusted" in s_lo
                else "pct_change_1m_nsa"
            )
            period = _strip_footnote_marker(sub) if sub else None
        elif "index" in s_lo and "code" not in s_lo and "base" not in s_lo:
            kind = "index_sa" if "seasonally adjusted" in s_lo else "index_nsa"
            period = _strip_footnote_marker(sub) if sub else None
        else:
            kind = "skip"

        classifications.append((kind, period))
    return classifications


def _coerce_number(value: Any) -> float | None:
    """Coerce a BLS numeric cell to ``float`` (``'-'`` / blanks become ``None``)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in {"", "-", "(NA)", "–"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


_VALUE_KINDS = {
    "relative_importance",
    "pct_change_12m",
    "pct_change_1m_sa",
    "pct_change_1m_nsa",
    "index_sa",
    "index_nsa",
}


def _row_to_records(
    row: tuple,
    classifications: list[tuple[str, str | None]],
    report_year: int,
    report_month: int,
) -> list[dict[str, Any]]:
    """Pivot one XLSX row into long-form records — one per ``(series, date)``."""
    if not row or row[0] in (None, ""):
        return []
    indent_raw = row[0]
    try:
        indent_level = int(str(indent_raw).strip())
    except (TypeError, ValueError):
        return []

    label: str | None = None
    group_code: str | None = None
    item_code: str | None = None
    industry_code: str | None = None
    product_code: str | None = None
    other_index_base: str | None = None
    per_date: dict[dateType, dict[str, Any]] = {}
    seasonally_adjusted: bool | None = None

    for i, (kind, period_label) in enumerate(classifications):
        if i >= len(row):
            break
        if kind in {"indent", "skip"}:
            continue
        raw = row[i]
        if raw is None:
            continue
        if isinstance(raw, str):
            cleaned: Any = raw.strip()
            if cleaned == "":
                continue
        else:
            cleaned = raw

        if kind == "label" and label is None:
            label = str(cleaned)
            continue
        if kind == "group_code" and isinstance(cleaned, str):
            group_code = cleaned
            continue
        if kind == "item_code" and isinstance(cleaned, str):
            item_code = cleaned
            continue
        if kind == "industry_code" and isinstance(cleaned, str):
            industry_code = cleaned
            continue
        if kind == "product_code" and isinstance(cleaned, str):
            product_code = cleaned
            continue
        if kind == "other_index_base":
            other_index_base = str(cleaned)
            continue

        if kind not in _VALUE_KINDS:
            continue
        is_range = kind in {
            "pct_change_12m",
            "pct_change_1m_sa",
            "pct_change_1m_nsa",
        }
        date_key = _period_label_to_date(
            period_label, report_year, report_month, is_range
        )
        if date_key is None:
            continue
        num = _coerce_number(cleaned)
        if num is None:
            continue

        slot = per_date.setdefault(date_key, {})
        if kind == "relative_importance":
            slot["relative_importance"] = num
        elif kind == "pct_change_12m":
            slot["pct_change_12m"] = num
        elif kind in {"pct_change_1m_sa", "pct_change_1m_nsa"}:
            slot["pct_change_1m"] = num
            seasonally_adjusted = kind == "pct_change_1m_sa"
        elif kind in {"index_sa", "index_nsa"}:
            slot["index_value"] = num
            seasonally_adjusted = kind == "index_sa"

    if not per_date:
        return []

    code_parts = [p for p in (group_code, item_code, industry_code, product_code) if p]
    code = "".join(code_parts) if code_parts else None

    if code is None and label is None:
        return []

    records: list[dict[str, Any]] = []

    for d in sorted(per_date):
        record: dict[str, Any] = {
            "date": d,
            "code": code,
            "label": label,
            "level": indent_level,
            "group_code": group_code,
            "item_code": item_code,
            "industry_code": industry_code,
            "product_code": product_code,
            "other_index_base": other_index_base,
            "seasonally_adjusted": seasonally_adjusted,
        }
        record.update(per_date[d])
        records.append(record)

    return records
