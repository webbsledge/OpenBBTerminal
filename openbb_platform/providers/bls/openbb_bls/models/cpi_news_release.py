"""BLS CPI news release tables."""

from __future__ import annotations

import calendar
import re
from datetime import date as dateType
from typing import Any

from openbb_core.app.model.abstract.error import OpenBBError
from openbb_core.provider.abstract.data import Data
from openbb_core.provider.abstract.fetcher import Fetcher
from openbb_core.provider.abstract.query_params import QueryParams
from openbb_core.provider.utils.errors import EmptyDataError
from pydantic import ConfigDict, Field, field_serializer

from openbb_bls.utils.constants import BLS_USER_AGENT

_TABLE_TITLES: dict[int, str] = {
    1: "CPI-U U.S. city average, by expenditure category",
    2: "CPI-U U.S. city average, by detailed expenditure category",
    3: "CPI-U U.S. city average, special aggregate indexes",
    4: "CPI-U selected areas, all items index",
    5: "Chained CPI (C-CPI-U) and CPI-U U.S. city average, all items 1-month and 12-month percent changes",
    6: "CPI-U U.S. city average, by expenditure category, 1-month analysis",
    7: "CPI-U U.S. city average, by expenditure category, 12-month analysis",
}

_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
_HIDE: dict[str, Any] = {"x-widget_config": {"hide": True}}
_XLSX_MAGIC = b"PK\x03\x04"
_MONTH_LOOKUP: dict[str, int] = {}
for _m in range(1, 13):
    _MONTH_LOOKUP[calendar.month_abbr[_m].lower()] = _m
    _MONTH_LOOKUP[calendar.month_name[_m].lower()] = _m
_MONTH_LOOKUP["sept"] = 9

_SINGLE_PERIOD_RE = re.compile(r"([A-Za-z]+)\.?\s+(\d{4})", re.IGNORECASE)
_RANGE_PERIOD_RE = re.compile(
    r"([A-Za-z]+)\.?\s*(\d{4})?\s*[-–]\s*([A-Za-z]+)\.?\s*(\d{4})", re.IGNORECASE
)


class _CpiNrQueryParams(QueryParams):
    """Shared query params: just the release month."""

    date: dateType | None = Field(
        default=None,
        description="Reporting month of the CPI release; None resolves to the most recent release.",
    )


_LABEL_HOVER_CARD = {
    "x-widget_config": {
        "renderFn": "hoverCard",
        "renderFnParams": {
            "hoverCard": {
                "cellField": "value",
                "markdown": "{footnote}",
            }
        },
    }
}


def _hdr(name: str) -> dict[str, Any]:
    """Override the column header name."""
    return {"x-widget_config": {"headerName": name}}


def _pct(name: str) -> dict[str, Any]:
    """Percent-column config: headerName + numeric type + percent formatter + greenRed render."""
    return {
        "x-widget_config": {
            "headerName": name,
            "cellDataType": "number",
            "formatterFn": "percent",
            "renderFn": "greenRed",
        }
    }


def _num(name: str) -> dict[str, Any]:
    """Numeric (non-percent) column config: headerName + numeric type."""
    return {
        "x-widget_config": {
            "headerName": name,
            "cellDataType": "number",
        }
    }


def _fetch_and_parse(table_number: int, query_date: dateType | None) -> dict[str, Any]:
    """Resolve the month, download the XLSX, and parse one News Release table."""
    if query_date is None:
        year, month, content = _discover_latest_nr(table_number)
    else:
        year, month = query_date.year, query_date.month
        content = _fetch_nr_xlsx(year, month, table_number)
        if content is None:
            raise OpenBBError(
                f"BLS CPI News Release Table {table_number} for "
                f"{calendar.month_name[month]} {year} was not found."
            )
    return _parse_nr_table(content, year, month, table_number)


class _CpiNrBaseData(Data):
    """Common fields shared by every CPI News Release table response."""

    model_config = ConfigDict(extra="ignore")

    date: dateType = Field(
        description="Reference month the row's values apply to (first of month).",
    )
    label: str | None = Field(
        default=None,
        description="BLS expenditure category, item, or area title.",
        json_schema_extra=_LABEL_HOVER_CARD,
    )
    row_index: int = Field(
        default=0,
        description="Sequential ordering index preserving BLS hierarchy.",
        json_schema_extra=_HIDE,
    )
    footnote: str | None = Field(
        default=None,
        description="Resolved footnote text for any (N) markers in label.",
        json_schema_extra=_HIDE,
    )
    table_id: str = Field(
        description="Source table identifier.",
        json_schema_extra=_HIDE,
    )
    table_name: str = Field(
        description="Full BLS table title.",
        json_schema_extra=_HIDE,
    )
    release_period: str = Field(
        description="Reporting month of the source release.",
        json_schema_extra=_HIDE,
    )

    @field_serializer("label", when_used="json")
    def _serialize_label_with_footnote(self, value: str | None) -> Any:
        """Wrap the label with footnote text for the hover-card renderer."""
        if value and self.footnote:
            return {"value": value, "footnote": self.footnote}
        return value


class BlsCpiNrTable1Data(_CpiNrBaseData):
    """One item row from Table 1 — CPI-U expenditure category indexes & changes."""

    model_config = ConfigDict(
        extra="ignore",
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI News Release — Table 1 (Expenditure Category)",
                "$.description": "CPI-U U.S. city average, by expenditure category.",
                "$.gridData": {"w": 40, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        },
    )

    level: int | None = Field(
        default=None, description="Indent depth from the source XLSX."
    )
    relative_importance: float | None = Field(
        default=None,
        description="Weight as a percent of the parent index.",
        json_schema_extra=_pct("Relative Importance"),
    )
    index_value: float | None = Field(
        default=None,
        description="NSA CPI index at the latest reference month (date).",
        json_schema_extra=_num("Index Value"),
    )
    index_value_prior_month: float | None = Field(
        default=None,
        description="NSA CPI index one month before date.",
        json_schema_extra=_num("Index Value (Prior Month)"),
    )
    index_value_year_ago: float | None = Field(
        default=None,
        description="NSA CPI index twelve months before date.",
        json_schema_extra=_num("Index Value (Year Ago)"),
    )
    pct_change_12m: float | None = Field(
        default=None,
        description="Twelve-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 12M"),
    )
    pct_change_1m_nsa: float | None = Field(
        default=None,
        description="One-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M NSA"),
    )
    pct_change_1m_sa: float | None = Field(
        default=None,
        description="One-month SA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M SA"),
    )
    pct_change_1m_sa_prior: float | None = Field(
        default=None,
        description="One-month SA percent change ending one month before date.",
        json_schema_extra=_pct("Pct Change 1M SA (Prior Month)"),
    )
    pct_change_1m_sa_two_back: float | None = Field(
        default=None,
        description="One-month SA percent change ending two months before date.",
        json_schema_extra=_pct("Pct Change 1M SA (Two Months Back)"),
    )


class BlsCpiNrTable2Data(_CpiNrBaseData):
    """One item row from Table 2 — CPI-U detailed expenditure category (% change only)."""

    model_config = ConfigDict(
        extra="ignore",
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI News Release — Table 2 (Detailed Expenditure)",
                "$.description": "CPI-U U.S. city average, by detailed expenditure category.",
                "$.gridData": {"w": 40, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        },
    )

    level: int | None = Field(
        default=None, description="Indent depth from the source XLSX."
    )
    relative_importance: float | None = Field(
        default=None,
        description="Weight as a percent of the parent index.",
        json_schema_extra=_pct("Relative Importance"),
    )
    pct_change_12m: float | None = Field(
        default=None,
        description="Twelve-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 12M"),
    )
    pct_change_1m_nsa: float | None = Field(
        default=None,
        description="One-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M NSA"),
    )
    pct_change_1m_sa: float | None = Field(
        default=None,
        description="One-month SA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M SA"),
    )
    pct_change_1m_sa_prior: float | None = Field(
        default=None,
        description="One-month SA percent change ending one month before date.",
        json_schema_extra=_pct("Pct Change 1M SA (Prior Month)"),
    )
    pct_change_1m_sa_two_back: float | None = Field(
        default=None,
        description="One-month SA percent change ending two months before date.",
        json_schema_extra=_pct("Pct Change 1M SA (Two Months Back)"),
    )


class BlsCpiNrTable3Data(_CpiNrBaseData):
    """One aggregate row from Table 3 — CPI-U special aggregate indexes."""

    model_config = ConfigDict(
        extra="ignore",
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI News Release — Table 3 (Special Aggregates)",
                "$.description": "CPI-U U.S. city average, special aggregate indexes.",
                "$.gridData": {"w": 40, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        },
    )

    level: int | None = Field(
        default=None, description="Indent depth from the source XLSX."
    )
    relative_importance: float | None = Field(
        default=None,
        description="Weight as a percent of the parent index.",
        json_schema_extra=_pct("Relative Importance"),
    )
    index_value: float | None = Field(
        default=None,
        description="NSA CPI index at the latest reference month (date).",
        json_schema_extra=_num("Index Value"),
    )
    index_value_prior_month: float | None = Field(
        default=None,
        description="NSA CPI index one month before date.",
        json_schema_extra=_num("Index Value (Prior Month)"),
    )
    index_value_year_ago: float | None = Field(
        default=None,
        description="NSA CPI index twelve months before date.",
        json_schema_extra=_num("Index Value (Year Ago)"),
    )
    pct_change_12m: float | None = Field(
        default=None,
        description="Twelve-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 12M"),
    )
    pct_change_1m_nsa: float | None = Field(
        default=None,
        description="One-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M NSA"),
    )
    pct_change_1m_sa: float | None = Field(
        default=None,
        description="One-month SA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M SA"),
    )
    pct_change_1m_sa_prior: float | None = Field(
        default=None,
        description="One-month SA percent change ending one month before date.",
        json_schema_extra=_pct("Pct Change 1M SA (Prior Month)"),
    )
    pct_change_1m_sa_two_back: float | None = Field(
        default=None,
        description="One-month SA percent change ending two months before date.",
        json_schema_extra=_pct("Pct Change 1M SA (Two Months Back)"),
    )


class BlsCpiNrTable4Data(_CpiNrBaseData):
    """One area row from Table 4 — CPI-U selected areas, all items index."""

    model_config = ConfigDict(
        extra="ignore",
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI News Release — Table 4 (Selected Areas)",
                "$.description": "CPI-U selected areas, all items index.",
                "$.gridData": {"w": 40, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        },
    )

    level: int | None = Field(
        default=None, description="Indent depth from the source XLSX."
    )
    pricing_schedule: str | None = Field(
        default=None,
        description="Area pricing schedule code (M = monthly, S = semiannual, B = bimonthly).",
    )
    pct_change_12m: float | None = Field(
        default=None,
        description="Twelve-month percent change ending at date.",
        json_schema_extra=_pct("Pct Change 12M"),
    )
    pct_change_1m_nsa: float | None = Field(
        default=None,
        description="One-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M NSA"),
    )
    pct_change_window: float | None = Field(
        default=None,
        description="Two-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 2M NSA"),
    )
    pct_change_window_start_date: dateType | None = Field(
        default=None,
        description="Start-of-month for the 2-month comparison.",
        json_schema_extra=_HIDE,
    )
    pct_change_12m_prior: float | None = Field(
        default=None,
        description="Twelve-month percent change ending one month before date.",
        json_schema_extra=_pct("Pct Change 12M (Prior Month)"),
    )
    pct_change_1m_nsa_prior: float | None = Field(
        default=None,
        description="One-month NSA percent change ending one month before date.",
        json_schema_extra=_pct("Pct Change 1M NSA (Prior Month)"),
    )
    pct_change_window_prior: float | None = Field(
        default=None,
        description="Two-month NSA percent change ending one month before date.",
        json_schema_extra=_pct("Pct Change 2M NSA (Prior Month)"),
    )
    pct_change_window_prior_start_date: dateType | None = Field(
        default=None,
        description="Start-of-month for pct_change_window_prior.",
        json_schema_extra=_HIDE,
    )


class BlsCpiNrTable5Data(_CpiNrBaseData):
    """One (date, basket) row from Table 5 — Chained CPI vs CPI-U time series."""

    model_config = ConfigDict(
        extra="ignore",
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI News Release — Table 5 (Chained vs CPI-U)",
                "$.description": "Chained CPI (C-CPI-U) and CPI-U all-items 1-month and 12-month percent changes.",
                "$.gridData": {"w": 40, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        },
    )

    pct_change_1m_nsa: float | None = Field(
        default=None,
        description="One-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M NSA"),
    )
    pct_change_12m: float | None = Field(
        default=None,
        description="Twelve-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 12M"),
    )


class BlsCpiNrTable6Data(_CpiNrBaseData):
    """One item row from Table 6 — CPI-U 1-month analysis with effect / std error / largest-since."""

    model_config = ConfigDict(
        extra="ignore",
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI News Release — Table 6 (1-Month Analysis)",
                "$.description": "CPI-U U.S. city average, by expenditure category, 1-month analysis.",
                "$.gridData": {"w": 40, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        },
    )

    level: int | None = Field(
        default=None, description="Indent depth from the source XLSX."
    )
    relative_importance: float | None = Field(
        default=None,
        description="Weight as a percent of the parent index.",
        json_schema_extra=_pct("Relative Importance"),
    )
    pct_change_1m_sa: float | None = Field(
        default=None,
        description="One-month SA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 1M SA"),
    )
    effect_on_all_items_sa: float | None = Field(
        default=None,
        description="SA contribution to the 1-month All Items change.",
        json_schema_extra=_pct("Effect on All Items (SA)"),
    )
    standard_error_median_price_change: float | None = Field(
        default=None,
        description="Standard error of the median price change.",
        json_schema_extra=_pct("Std Error (Median Price Change)"),
    )
    largest_or_smallest_since_marker: str | None = Field(
        default=None,
        description="L-/S- marker plus the historical month at which the current SA 1-month change last occurred.",
        json_schema_extra=_hdr("Largest/Smallest Since"),
    )
    largest_or_smallest_since_pct: float | None = Field(
        default=None,
        description="SA 1-month percent change recorded at the historical reference month.",
        json_schema_extra=_pct("Largest/Smallest Since (Pct)"),
    )


class BlsCpiNrTable7Data(_CpiNrBaseData):
    """One item row from Table 7 — CPI-U 12-month analysis with effect / std error / largest-since."""

    model_config = ConfigDict(
        extra="ignore",
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI News Release — Table 7 (12-Month Analysis)",
                "$.description": "CPI-U U.S. city average, by expenditure category, 12-month analysis.",
                "$.gridData": {"w": 40, "h": 27},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        },
    )

    level: int | None = Field(
        default=None, description="Indent depth from the source XLSX."
    )
    relative_importance: float | None = Field(
        default=None,
        description="Weight as a percent of the parent index.",
        json_schema_extra=_pct("Relative Importance"),
    )
    pct_change_12m: float | None = Field(
        default=None,
        description="Twelve-month NSA percent change ending at date.",
        json_schema_extra=_pct("Pct Change 12M"),
    )
    effect_on_all_items_nsa: float | None = Field(
        default=None,
        description="NSA contribution to the 12-month All Items change.",
        json_schema_extra=_pct("Effect on All Items (NSA)"),
    )
    standard_error_median_price_change: float | None = Field(
        default=None,
        description="Standard error of the median price change.",
        json_schema_extra=_pct("Std Error (Median Price Change)"),
    )
    largest_or_smallest_since_marker: str | None = Field(
        default=None,
        description="L-/S- marker plus the historical month at which the current NSA 12-month change last occurred.",
        json_schema_extra=_hdr("Largest/Smallest Since"),
    )
    largest_or_smallest_since_pct: float | None = Field(
        default=None,
        description="NSA 12-month percent change recorded at the historical reference month.",
        json_schema_extra=_pct("Largest/Smallest Since (Pct)"),
    )


def _make_fetcher(table_number: int, data_class: type[Data]) -> type[Fetcher]:
    """Build a Fetcher class bound to one News Release table number and Data model."""

    class _Fetcher(Fetcher[_CpiNrQueryParams, list[Data]]):
        """BLS CPI News Release per-table fetcher."""

        require_credentials = False
        data_type = data_class

        @staticmethod
        def transform_query(params: dict[str, Any]) -> _CpiNrQueryParams:
            """Validate and coerce the query."""
            return _CpiNrQueryParams(**params)

        @staticmethod
        def extract_data(
            query: _CpiNrQueryParams,
            credentials: dict[str, str] | None,
            **kwargs: Any,
        ) -> dict[str, Any]:
            """Download the monthly XLSX and parse the bound table."""
            return _fetch_and_parse(table_number, query.date)

        @staticmethod
        def transform_data(
            query: _CpiNrQueryParams,
            data: dict[str, Any],
            **kwargs: Any,
        ) -> list[Data]:
            """Coerce parsed rows into the table-specific Data model."""
            rows = data.get("rows", [])
            if not rows:
                raise EmptyDataError(
                    f"No rows parsed from CPI News Release table "
                    f"{table_number} ({data.get('table_id', '?')})."
                )
            return [data_class.model_validate(r) for r in rows]

    _Fetcher.__name__ = f"BlsCpiNrTable{table_number}Fetcher"
    _Fetcher.__qualname__ = _Fetcher.__name__
    return _Fetcher


BlsCpiNrTable1Fetcher = _make_fetcher(1, BlsCpiNrTable1Data)
BlsCpiNrTable2Fetcher = _make_fetcher(2, BlsCpiNrTable2Data)
BlsCpiNrTable3Fetcher = _make_fetcher(3, BlsCpiNrTable3Data)
BlsCpiNrTable4Fetcher = _make_fetcher(4, BlsCpiNrTable4Data)
BlsCpiNrTable5Fetcher = _make_fetcher(5, BlsCpiNrTable5Data)
BlsCpiNrTable6Fetcher = _make_fetcher(6, BlsCpiNrTable6Data)
BlsCpiNrTable7Fetcher = _make_fetcher(7, BlsCpiNrTable7Data)


def _nr_url(year: int, month: int, table: int) -> str:
    """Canonical URL for one News Release Table file."""
    return (
        f"https://www.bls.gov/cpi/tables/supplemental-files/"
        f"news-release-table{table}-{year:04d}{month:02d}.xlsx"
    )


def _fetch_nr_xlsx(year: int, month: int, table: int) -> bytes | None:
    """Fetch one News Release Table XLSX; ``None`` on redirect or non-XLSX."""
    import requests

    url = _nr_url(year, month, table)
    resp = requests.get(url, headers=_HEADERS, timeout=60, allow_redirects=False)
    if resp.status_code in (301, 302, 303, 307, 308, 404):
        return None
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")
    if not resp.content.startswith(_XLSX_MAGIC):
        return None
    return resp.content


def _discover_latest_nr(table: int) -> tuple[int, int, bytes]:
    """Walk back from today until a published News Release Table file is found."""
    today = dateType.today()
    year, month = today.year, today.month
    for _ in range(6):
        content = _fetch_nr_xlsx(year, month, table)
        if content is not None:
            return year, month, content
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    raise OpenBBError(
        "Could not locate a recent CPI News Release Table XLSX within the "
        "last six months."
    )


def _to_float(value: Any) -> float | None:
    """Coerce a BLS numeric cell to ``float``; blanks become ``None``."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in ("", "-"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _clean_header(text: Any) -> str:
    """Collapse newlines and whitespace in one header cell."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\n", " ").split()).strip()


def _month_token_to_num(token: str) -> int | None:
    """Map a BLS month token (``'Apr'``, ``'Apr.'``, ``'Sept'``) to ``1..12``."""
    return _MONTH_LOOKUP.get(token.strip().lower().rstrip("."))


def _parse_single_period(text: str, report_year: int) -> dateType | None:
    """Parse a single-period header (``'Apr. 2026'`` or ``'May 2025'``)."""
    match = _SINGLE_PERIOD_RE.search(text)
    if match is None:
        return None
    month_token, year_token = match.groups()
    month = _month_token_to_num(month_token)
    if month is None:
        return None
    return dateType(int(year_token), month, 1)


def _parse_range_period(text: str) -> tuple[dateType, dateType] | None:
    """Parse a range header (``'Apr. 2025-Apr. 2026'``) → ``(start_date, end_date)``."""
    match = _RANGE_PERIOD_RE.search(text)
    if match is None:
        return None
    start_m, start_y, end_m, end_y = match.groups()
    start_month = _month_token_to_num(start_m)
    end_month = _month_token_to_num(end_m)
    if start_month is None or end_month is None or end_y is None:
        return None
    end_year = int(end_y)
    start_year = int(start_y) if start_y is not None else end_year
    return dateType(start_year, start_month, 1), dateType(end_year, end_month, 1)


_PCT_TO_FROM_RE = re.compile(
    r"percent change to\s+([A-Za-z]+)\.?\s*(\d{4})\s*from",
    re.IGNORECASE,
)
_MONTH_YEAR_LABEL_RE = re.compile(r"^([A-Za-z]+)\.?\s+(\d{4})$")
_FOOTNOTE_DEF_RE = re.compile(r"^\(([^)]+)\)\s*(.*)$")
_LABEL_REF_RE = re.compile(r"\(([^)]+)\)")


def _parse_to_from_super(super_text: str) -> dateType | None:
    """Extract the 'to <date>' anchor from a 'Percent change to X from:' super-header."""
    match = _PCT_TO_FROM_RE.search(super_text)
    if match is None:
        return None
    month_token, year_token = match.group(1), match.group(2)
    month = _month_token_to_num(month_token)
    if month is None:
        return None
    return dateType(int(year_token), month, 1)


def _months_between(start: dateType, end: dateType) -> int:
    """Return the absolute number of calendar months between two first-of-month dates."""
    months = (end.year - start.year) * 12 + (end.month - start.month)
    return abs(months)


def _parse_month_year_label(text: str) -> dateType | None:
    """Parse a 'Month YYYY' / 'Mon. YYYY' row label into a first-of-month date."""
    match = _MONTH_YEAR_LABEL_RE.match(text.strip())
    if match is None:
        return None
    month_token, year_token = match.groups()
    month = _month_token_to_num(month_token)
    if month is None:
        return None
    return dateType(int(year_token), month, 1)


def _parse_nr_footnotes(rows: list[tuple]) -> dict[str, str]:
    """Build ``{'(N)': 'definition text'}`` from trailing footnote-definition rows."""
    out: dict[str, str] = {}
    for row in rows:
        if not row:
            continue
        for cell in row:
            if not isinstance(cell, str):
                continue
            text = cell.strip()
            match = _FOOTNOTE_DEF_RE.match(text)
            if match is None:
                continue
            marker = f"({match.group(1)})"
            body = match.group(2).strip() or text
            out[marker] = body
            break
    return out


def _resolve_nr_footnotes(label: str | None, footnotes: dict[str, str]) -> str | None:
    """Concatenate footnote text for any ``(N)`` markers in ``label``."""
    if not label or not footnotes:
        return None
    matched: list[str] = []
    seen: set[str] = set()
    for match in _LABEL_REF_RE.finditer(label):
        marker = match.group(0)
        if marker not in footnotes or marker in seen:
            continue
        seen.add(marker)
        matched.append(footnotes[marker])
    if not matched:
        return None
    return "\n\n".join(matched)


def _parse_nr_table(
    content: bytes, year: int, month: int, table_number: int
) -> dict[str, Any]:
    """Parse one News Release Table XLSX into long-form rows."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    raw_rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    release_period = f"{calendar.month_name[month]} {year}"
    release_date = dateType(year, month, 1)
    table_id = f"cpi-nr-{year:04d}-{month:02d}-t{table_number}"
    table_name = (
        f"CPI News Release {release_period} — Table {table_number}: "
        f"{_TABLE_TITLES[table_number]}"
    )

    super_row: tuple = ()
    sub_row: tuple = ()
    data_start = 0
    for idx, row in enumerate(raw_rows):
        if not row or row[0] is None:
            continue
        if _clean_header(row[0]).lower() == "indent level":
            super_row = row
            sub_row = raw_rows[idx + 1] if idx + 1 < len(raw_rows) else ()
            data_start = idx + 2
            break
    if not super_row:
        return {"rows": [], "table_id": table_id, "table_name": table_name}

    body_rows = raw_rows[data_start:]
    footnotes = _parse_nr_footnotes(body_rows)

    if table_number == 5:
        rows = _parse_nr_table_5(
            body_rows,
            super_row,
            sub_row,
            year,
            month,
            table_id,
            table_name,
            release_period,
            release_date,
            footnotes,
        )
        return {"rows": rows, "table_id": table_id, "table_name": table_name}

    if table_number in (6, 7):
        rows = _parse_nr_table_6_7(
            body_rows,
            table_number,
            year,
            month,
            table_id,
            table_name,
            release_period,
            release_date,
            footnotes,
        )
        return {"rows": rows, "table_id": table_id, "table_name": table_name}

    classifications = _classify_columns(super_row, sub_row, year, month)
    pricing_col = _find_pricing_schedule_col(super_row, sub_row)

    out: list[dict[str, Any]] = []
    source_idx = 0
    for row in body_rows:
        if not row or row[0] in (None, ""):
            continue
        try:
            level = int(str(row[0]).strip())
        except (TypeError, ValueError):
            continue
        label = (
            row[1].strip()
            if len(row) > 1 and isinstance(row[1], str) and row[1].strip()
            else None
        )
        if label is None:
            continue
        source_idx += 1

        ri_value: float | None = None
        pricing_schedule: str | None = None
        if pricing_col is not None and pricing_col < len(row):
            pcell = row[pricing_col]
            if isinstance(pcell, str) and pcell.strip():
                pricing_schedule = pcell.strip()
            elif pcell is not None:
                pricing_schedule = str(pcell).strip() or None
        per_date: dict[dateType, dict[str, Any]] = {}

        for i, classification in enumerate(classifications):
            if i >= len(row) or classification is None:
                continue
            kind, period, extra = classification
            value = row[i]
            if value is None:
                continue
            cleaned: Any = value.strip() if isinstance(value, str) else value
            if isinstance(cleaned, str) and cleaned == "":
                continue
            num = _to_float(cleaned)
            if num is None:
                continue

            if kind == "relative_importance":
                ri_value = num
            elif kind == "index_nsa" and period is not None:
                per_date.setdefault(period, {})["index_value"] = num
            elif kind == "pct_change_12m_nsa" and period is not None:
                per_date.setdefault(period, {})["pct_change_12m"] = num
            elif kind == "pct_change_1m_nsa" and period is not None:
                per_date.setdefault(period, {})["pct_change_1m_nsa"] = num
            elif kind == "pct_change_1m_sa" and period is not None:
                per_date.setdefault(period, {})["pct_change_1m_sa"] = num
            elif kind == "pct_change_window" and period is not None:
                slot = per_date.setdefault(period, {})
                slot["pct_change_window"] = num
                if extra is not None:
                    slot["pct_change_window_start_date"] = extra

        if not per_date and ri_value is None and pricing_schedule is None:
            continue

        record = _pivot_per_item_row(
            per_date,
            label=label,
            level=level,
            source_idx=source_idx,
            ri_value=ri_value,
            pricing_schedule=pricing_schedule,
            footnote=_resolve_nr_footnotes(label, footnotes),
            release_date=release_date,
            release_period=release_period,
            table_id=table_id,
            table_name=table_name,
            table_number=table_number,
        )
        out.append(record)

    out.sort(key=lambda r: r["row_index"])
    return {"rows": out, "table_id": table_id, "table_name": table_name}


def _shift_month(d: dateType, delta_months: int) -> dateType:
    """Shift a first-of-month date by an integer number of months."""
    total = d.year * 12 + (d.month - 1) + delta_months
    return dateType(total // 12, (total % 12) + 1, 1)


def _pivot_per_item_row(
    per_date: dict[dateType, dict[str, Any]],
    *,
    label: str,
    level: int,
    source_idx: int,
    ri_value: float | None,
    pricing_schedule: str | None,
    footnote: str | None,
    release_date: dateType,
    release_period: str,
    table_id: str,
    table_name: str,
    table_number: int,
) -> dict[str, Any]:
    """Collapse a per-date measure dict into one item row with offset-named columns."""
    anchor = max(per_date) if per_date else release_date
    prior = _shift_month(anchor, -1)
    two_back = _shift_month(anchor, -2)
    year_ago = _shift_month(anchor, -12)

    latest = per_date.get(anchor, {})
    prior_slot = per_date.get(prior, {})
    two_back_slot = per_date.get(two_back, {})
    year_ago_slot = per_date.get(year_ago, {})

    return {
        "date": anchor,
        "label": label,
        "level": level,
        "row_index": source_idx,
        "relative_importance": ri_value,
        "pricing_schedule": pricing_schedule,
        "footnote": footnote,
        "index_value": latest.get("index_value"),
        "index_value_prior_month": prior_slot.get("index_value"),
        "index_value_year_ago": year_ago_slot.get("index_value"),
        "pct_change_12m": latest.get("pct_change_12m"),
        "pct_change_1m_nsa": latest.get("pct_change_1m_nsa"),
        "pct_change_1m_sa": latest.get("pct_change_1m_sa"),
        "pct_change_1m_sa_prior": prior_slot.get("pct_change_1m_sa"),
        "pct_change_1m_sa_two_back": two_back_slot.get("pct_change_1m_sa"),
        "pct_change_window": latest.get("pct_change_window"),
        "pct_change_window_start_date": latest.get("pct_change_window_start_date"),
        "pct_change_12m_prior": prior_slot.get("pct_change_12m"),
        "pct_change_1m_nsa_prior": prior_slot.get("pct_change_1m_nsa"),
        "pct_change_window_prior": prior_slot.get("pct_change_window"),
        "pct_change_window_prior_start_date": prior_slot.get(
            "pct_change_window_start_date"
        ),
        "table_id": table_id,
        "table_name": table_name,
        "table_number": table_number,
        "release_period": release_period,
    }


def _parse_nr_table_5(
    body_rows: list[tuple],
    super_row: tuple,
    sub_row: tuple,
    year: int,
    month: int,
    table_id: str,
    table_name: str,
    release_period: str,
    release_date: dateType,
    footnotes: dict[str, str],
) -> list[dict[str, Any]]:
    """Parse Table 5 (Month Year × C-CPI-U/CPI-U 1m/12m % change) into long-form rows."""
    n = max(len(super_row), len(sub_row))
    filled_super: list[str] = []
    last = ""
    for i in range(n):
        cell = _clean_header(super_row[i] if i < len(super_row) else None)
        if cell:
            last = cell
        filled_super.append(last)

    columns: list[tuple[str, str] | None] = []
    for i in range(n):
        s = filled_super[i]
        sub = _clean_header(sub_row[i] if i < len(sub_row) else None)
        s_lo = s.lower()
        sub_norm = sub.upper().replace("(1)", "").strip()
        if i <= 1 or not sub:
            columns.append(None)
            continue
        if "12-month" in s_lo or "twelve" in s_lo:
            kind = "pct_change_12m"
        elif "1-month" in s_lo or "one" in s_lo:
            kind = "pct_change_1m_nsa"
        else:
            columns.append(None)
            continue
        basket = (
            "C-CPI-U"
            if "C-CPI-U" in sub_norm
            else "CPI-U"
            if "CPI-U" in sub_norm
            else None
        )
        if basket is None:
            columns.append(None)
            continue
        columns.append((kind, basket))

    grouped: dict[tuple[dateType, str], dict[str, Any]] = {}
    for row in body_rows:
        if not row or row[1] in (None, ""):
            continue
        if not isinstance(row[1], str):
            continue
        row_label = row[1].strip()
        period_date = _parse_month_year_label(row_label)
        if period_date is None:
            continue
        footnote_text = _resolve_nr_footnotes(row_label, footnotes)
        for i, classification in enumerate(columns):
            if classification is None or i >= len(row):
                continue
            kind, basket = classification
            num = _to_float(row[i])
            if num is None:
                continue
            key = (period_date, basket)
            slot = grouped.setdefault(
                key,
                {
                    "date": period_date,
                    "label": basket,
                    "level": 0,
                    "relative_importance": None,
                    "pricing_schedule": None,
                    "footnote": footnote_text,
                    "table_id": table_id,
                    "table_name": table_name,
                    "table_number": 5,
                    "release_period": release_period,
                },
            )
            slot[kind] = num

    out = list(grouped.values())
    out.sort(key=lambda r: (-(r["date"]).toordinal(), r["label"]))
    for i, entry in enumerate(out, start=1):
        entry["row_index"] = i
    return out


def _parse_nr_table_6_7(
    body_rows: list[tuple],
    table_number: int,
    year: int,
    month: int,
    table_id: str,
    table_name: str,
    release_period: str,
    release_date: dateType,
    footnotes: dict[str, str],
) -> list[dict[str, Any]]:
    """Parse Tables 6 and 7 positionally — main change + effect + std error + largest-since marker."""
    main_kind = "pct_change_1m_sa" if table_number == 6 else "pct_change_12m"
    effect_kind = (
        "effect_on_all_items_sa" if table_number == 6 else "effect_on_all_items_nsa"
    )
    out: list[dict[str, Any]] = []
    source_idx = 0
    for row in body_rows:
        if not row or row[0] in (None, ""):
            continue
        try:
            level = int(str(row[0]).strip())
        except (TypeError, ValueError):
            continue
        if len(row) < 2 or not isinstance(row[1], str) or not row[1].strip():
            continue
        label = row[1].strip()
        source_idx += 1
        record: dict[str, Any] = {
            "date": release_date,
            "label": label,
            "level": level,
            "row_index": source_idx,
            "footnote": _resolve_nr_footnotes(label, footnotes),
            "table_id": table_id,
            "table_name": table_name,
            "table_number": table_number,
            "release_period": release_period,
            "relative_importance": _to_float(row[2]) if len(row) > 2 else None,
            main_kind: _to_float(row[3]) if len(row) > 3 else None,
            effect_kind: _to_float(row[4]) if len(row) > 4 else None,
            "standard_error_median_price_change": _to_float(row[5])
            if len(row) > 5
            else None,
            "largest_or_smallest_since_marker": (
                row[6].strip()
                if len(row) > 6 and isinstance(row[6], str) and row[6].strip()
                else None
            ),
            "largest_or_smallest_since_pct": _to_float(row[7])
            if len(row) > 7
            else None,
        }
        out.append(record)
    return out


def _find_pricing_schedule_col(super_row: tuple, sub_row: tuple) -> int | None:
    """Locate the column index of the 'Pricing Schedule' column, if present."""
    n = max(len(super_row), len(sub_row))
    for i in range(n):
        cell = _clean_header(super_row[i] if i < len(super_row) else None)
        if "pricing schedule" in cell.lower():
            return i
    return None


def _classify_columns(
    super_row: tuple, sub_row: tuple, report_year: int, report_month: int
) -> list[tuple[str, dateType | None, dateType | None] | None]:
    """Classify each column into ``(kind, period_end, period_start | None)`` triples."""
    n = max(len(super_row), len(sub_row))
    filled_super: list[str] = []
    last = ""
    for i in range(n):
        cell = _clean_header(super_row[i] if i < len(super_row) else None)
        if cell:
            last = cell
        filled_super.append(last)

    out: list[tuple[str, dateType | None, dateType | None] | None] = []
    for i in range(n):
        s = filled_super[i]
        sub = _clean_header(sub_row[i] if i < len(sub_row) else None)
        s_lo = s.lower()

        if i == 0 or s_lo == "indent level":
            out.append(None)
            continue
        if s_lo.startswith(
            (
                "expenditure category",
                "special aggregate",
                "title",
                "item",
                "area",
                "pricing schedule",
            )
        ):
            out.append(None)
            continue
        if "relative importance" in s_lo:
            out.append(("relative_importance", None, None))
            continue
        if "unadjusted index" in s_lo or s_lo.endswith("indexes"):
            d = _parse_single_period(sub, report_year)
            out.append(("index_nsa", d, None))
            continue
        if "seasonally adjusted percent change" in s_lo:
            rng = _parse_range_period(sub)
            d = rng[1] if rng else None
            out.append(("pct_change_1m_sa", d, None))
            continue
        if "percent change to" in s_lo:
            end_date = _parse_to_from_super(s)
            start_date = _parse_single_period(sub, report_year) if sub else None
            if end_date is None or start_date is None:
                out.append(None)
                continue
            months = _months_between(start_date, end_date)
            if months >= 12:
                out.append(("pct_change_12m_nsa", end_date, None))
            elif months == 1:
                out.append(("pct_change_1m_nsa", end_date, None))
            else:
                out.append(("pct_change_window", end_date, start_date))
            continue
        if "twelve month" in s_lo or s_lo == "twelve":
            rng = _parse_range_period(sub)
            if rng is not None:
                _, end = rng
                out.append(("pct_change_12m_nsa", end, None))
                continue
            d = _parse_single_period(sub, report_year)
            if d is not None:
                out.append(("pct_change_12m_nsa", d, None))
                continue
            release_anchor = dateType(report_year, report_month, 1)
            out.append(("pct_change_12m_nsa", release_anchor, None))
            continue
        if "one month" in s_lo or s_lo == "month":
            rng = _parse_range_period(sub)
            sub_lo = sub.lower()
            kind = (
                "pct_change_1m_sa"
                if "seasonally adjusted" in sub_lo
                else "pct_change_1m_nsa"
            )
            if rng is not None:
                _, end = rng
                out.append((kind, end, None))
                continue
            d = _parse_single_period(sub, report_year)
            if d is not None:
                out.append((kind, d, None))
                continue
            release_anchor = dateType(report_year, report_month, 1)
            out.append((kind, release_anchor, None))
            continue
        if "unadjusted percent change" in s_lo or "percent change" in s_lo:
            rng = _parse_range_period(sub)
            if rng is None:
                out.append(None)
                continue
            start, end = rng
            months = _months_between(start, end)
            if months >= 12:
                out.append(("pct_change_12m_nsa", end, None))
            else:
                out.append(("pct_change_1m_nsa", end, None))
            continue
        out.append(None)
    return out
