"""BLS CPI seasonal adjustment factors."""

from __future__ import annotations

from datetime import date as dateType
from typing import Any

from openbb_core.app.model.abstract.error import OpenBBError
from openbb_core.provider.abstract.data import Data
from openbb_core.provider.abstract.fetcher import Fetcher
from openbb_core.provider.abstract.query_params import QueryParams
from openbb_core.provider.utils.errors import EmptyDataError
from pydantic import ConfigDict, Field

from openbb_bls.utils.constants import BLS_USER_AGENT

_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
_HIDE: dict[str, Any] = {"x-widget_config": {"hide": True}}
_XLSX_MAGIC = b"PK\x03\x04"
_MONTH_NUM = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


class BlsCpiSeasonalFactorsQueryParams(QueryParams):
    """BLS CPI Seasonal Factors Query Parameters."""

    year: int | None = Field(
        default=None,
        description="Release year of the revised seasonal-adjustment file; None walks back until a published file is found.",
    )


class BlsCpiSeasonalFactorsData(Data):
    """One ``(CPI series, month)`` seasonally-adjusted observation."""

    model_config = ConfigDict(
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI Seasonal Factors",
                "$.description": (
                    "Consumer Price Index revised seasonally adjusted "
                    "indexes and the underlying seasonal factors."
                ),
                "$.gridData": {"w": 40, "h": 20},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        }
    )

    date: dateType = Field(
        description="First day of the calendar month the values apply to.",
    )
    item_code: str = Field(
        description="Short BLS CPI item code.",
    )
    series_id: str = Field(
        description="Full BLS time-series ID.",
    )
    title: str | None = Field(
        default=None,
        description="Human-readable CPI item title.",
    )
    seasonally_adjusted_index: float | None = Field(
        default=None,
        description="Revised seasonally-adjusted index level.",
    )
    seasonal_factor: float | None = Field(
        default=None,
        description="Multiplicative seasonal factor (100.000 = no seasonality).",
    )
    table_id: str = Field(
        description="Source table identifier.",
        json_schema_extra=_HIDE,
    )
    table_name: str = Field(
        description="Full BLS table title.",
        json_schema_extra=_HIDE,
    )


class BlsCpiSeasonalFactorsFetcher(
    Fetcher[
        BlsCpiSeasonalFactorsQueryParams,
        list[BlsCpiSeasonalFactorsData],
    ]
):
    """BLS CPI Seasonal Factors Fetcher."""

    require_credentials = False

    @staticmethod
    def transform_query(
        params: dict[str, Any],
    ) -> BlsCpiSeasonalFactorsQueryParams:
        """Validate and coerce the query."""
        return BlsCpiSeasonalFactorsQueryParams(**params)

    @staticmethod
    def extract_data(
        query: BlsCpiSeasonalFactorsQueryParams,
        credentials: dict[str, str] | None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Fetch the requested yearly SA XLSX and pivot to long form."""
        if query.year is None:
            year, content = _discover_latest_sa()
        else:
            content = _fetch_sa_xlsx(query.year)
            if content is None:
                raise OpenBBError(
                    f"BLS CPI Seasonal Adjustment file for {query.year} not found."
                )
            year = query.year
        return _parse_sa_xlsx(content, year)

    @staticmethod
    def transform_data(
        query: BlsCpiSeasonalFactorsQueryParams,
        data: dict[str, Any],
        **kwargs: Any,
    ) -> list[BlsCpiSeasonalFactorsData]:
        """Coerce parsed rows into ``BlsCpiSeasonalFactorsData``."""
        rows = data.get("rows", [])
        if not rows:
            raise EmptyDataError(
                f"No rows parsed from CPI seasonal-factor table "
                f"'{data.get('table_id', '?')}'."
            )
        return [BlsCpiSeasonalFactorsData.model_validate(r) for r in rows]


def _sa_url(year: int) -> str:
    """Canonical SA-factors URL for ``year``."""
    return (
        f"https://www.bls.gov/cpi/tables/seasonal-adjustment/"
        f"revised-seasonally-adjusted-indexes-{year}.xlsx"
    )


def _fetch_sa_xlsx(year: int) -> bytes | None:
    """Fetch one yearly SA XLSX; ``None`` if BLS redirects or returns non-XLSX."""
    import requests

    resp = requests.get(
        _sa_url(year), headers=_HEADERS, timeout=60, allow_redirects=False
    )
    if resp.status_code in (301, 302, 303, 307, 308, 404):
        return None
    if resp.status_code != 200:
        raise OpenBBError(
            f"BLS returned HTTP {resp.status_code} fetching {_sa_url(year)}."
        )
    if not resp.content.startswith(_XLSX_MAGIC):
        return None
    return resp.content


def _discover_latest_sa() -> tuple[int, bytes]:
    """Walk back from the current year to find the most recent SA XLSX."""
    today = dateType.today()
    for year in range(today.year, today.year - 6, -1):
        content = _fetch_sa_xlsx(year)
        if content is not None:
            return year, content
    raise OpenBBError(
        "Could not locate a recent CPI Seasonal Adjustment XLSX — the BLS "
        "publication URL pattern may have changed."
    )


def _to_float(value: Any) -> float | None:
    """Coerce a BLS numeric cell to ``float``; blanks become ``None``."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in ("", "-"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_sa_xlsx(content: bytes, year: int) -> dict[str, Any]:
    """Open the XLSX and emit long-form rows pivoted by month."""
    import io
    import warnings

    import openpyxl

    # Some BLS workbooks carry header/footer XML that openpyxl cannot parse and
    # warns about; the warning is irrelevant to the cell data we read here.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        raw_rows = list(sheet.iter_rows(values_only=True))
        wb.close()

    table_id = f"cpi-sa-{year}"
    table_name = f"CPI Revised Seasonally Adjusted Indexes and Factors — file {year}"

    header_idx = -1
    for idx, row in enumerate(raw_rows):
        if row and isinstance(row[0], str) and row[0].strip().upper() == "ITEM":
            header_idx = idx
            break
    if header_idx < 0:
        return {"rows": [], "table_id": table_id, "table_name": table_name}

    headers = [
        str(c).strip().upper() if c is not None else "" for c in raw_rows[header_idx]
    ]
    col_idx = {name: headers.index(name) for name in headers if name}
    if not all(
        k in col_idx for k in ("ITEM", "TITLE", "SERIESID", "DATA_TYPE", "YEAR")
    ):
        return {"rows": [], "table_id": table_id, "table_name": table_name}

    month_cols = [(m, col_idx[m]) for m in _MONTH_NUM if m in col_idx]
    if len(month_cols) != 12:
        return {"rows": [], "table_id": table_id, "table_name": table_name}

    pivots: dict[tuple[str, int], dict[str, Any]] = {}
    for row in raw_rows[header_idx + 1 :]:
        if not row or row[col_idx["ITEM"]] is None:
            continue
        item = str(row[col_idx["ITEM"]]).strip()
        title = row[col_idx["TITLE"]]
        series_id = row[col_idx["SERIESID"]]
        data_type = row[col_idx["DATA_TYPE"]]
        year_cell = row[col_idx["YEAR"]]
        try:
            row_year = int(str(year_cell).strip())
        except (TypeError, ValueError):
            continue
        if data_type is None:
            continue
        data_type_norm = str(data_type).strip().upper()

        key = (item, row_year)
        slot = pivots.setdefault(
            key,
            {
                "item_code": item,
                "title": str(title).strip()
                if isinstance(title, str) and title.strip()
                else None,
                "series_id": (
                    str(series_id).strip()
                    if isinstance(series_id, str) and series_id.strip()
                    else None
                ),
                "year": row_year,
                "monthly_index": {},
                "monthly_factor": {},
            },
        )
        for month_name, mcol in month_cols:
            month_num = _MONTH_NUM[month_name]
            value = _to_float(row[mcol]) if mcol < len(row) else None
            if value is None:
                continue
            if "FACTOR" in data_type_norm:
                slot["monthly_factor"][month_num] = value
            else:
                slot["monthly_index"][month_num] = value

    out: list[dict[str, Any]] = []
    for (item, row_year), slot in pivots.items():
        series_id = slot.get("series_id") or ""
        title = slot.get("title")
        for month in range(1, 13):
            index_value = slot["monthly_index"].get(month)
            factor_value = slot["monthly_factor"].get(month)
            if index_value is None and factor_value is None:
                continue
            out.append(
                {
                    "date": dateType(row_year, month, 1),
                    "item_code": item,
                    "series_id": series_id,
                    "title": title,
                    "seasonally_adjusted_index": index_value,
                    "seasonal_factor": factor_value,
                    "table_id": table_id,
                    "table_name": table_name,
                }
            )
    out.sort(key=lambda r: (r["date"], r["item_code"]))
    return {"rows": out, "table_id": table_id, "table_name": table_name}
