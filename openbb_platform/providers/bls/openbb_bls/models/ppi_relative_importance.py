"""BLS PPI relative importance tables."""

import re
from datetime import date as dateType
from typing import Any, Literal

from openbb_core.app.model.abstract.error import OpenBBError
from openbb_core.provider.abstract.data import Data
from openbb_core.provider.abstract.fetcher import Fetcher
from openbb_core.provider.abstract.query_params import QueryParams
from openbb_core.provider.utils.errors import EmptyDataError
from pydantic import ConfigDict, Field

from openbb_bls.utils.constants import BLS_USER_AGENT
from openbb_bls.utils.tables import (
    PPI_RELATIVE_IMPORTANCE_CANONICAL,
    PPI_RELATIVE_IMPORTANCE_TABLES,
)

PpiRelativeImportanceCategory = Literal[
    "final_demand",
    "intermediate_demand_commodity",
    "intermediate_demand_flow",
    "commodity",
    "service_construction",
]

PpiRelativeImportanceTableId = Literal[
    "ppi-fdallrel",
    "ppi-fdgrouprel",
    "ppi-idcallrel",
    "ppi-idcgrouprel",
    "ppi-idpallrel",
    "ppi-idpgrouprel",
    "ppi-comrlp",
    "ppi-weprel",
]

_HIDE: dict[str, Any] = {"x-widget_config": {"hide": True}}
_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
_YEAR_RE = re.compile(r"(20\d{2})")


class BlsPpiRelativeImportanceQueryParams(QueryParams):
    """BLS PPI Relative Importance Table Query Parameters."""

    __json_schema_extra__ = {
        "category": {
            "x-widget_config": {
                "options": [
                    {"label": "Final Demand", "value": "final_demand"},
                    {
                        "label": "Intermediate Demand by Commodity Type",
                        "value": "intermediate_demand_commodity",
                    },
                    {
                        "label": "Intermediate Demand by Production Flow",
                        "value": "intermediate_demand_flow",
                    },
                    {"label": "Commodity", "value": "commodity"},
                    {
                        "label": "Wherever-Provided Services & Construction",
                        "value": "service_construction",
                    },
                ]
            }
        },
        "table_id": {
            "x-widget_config": {
                "options": [
                    {"label": entry["label"], "value": key}
                    for key, entry in PPI_RELATIVE_IMPORTANCE_TABLES.items()
                ],
                "style": {"popupWidth": 750},
            }
        },
    }

    category: PpiRelativeImportanceCategory = Field(
        default="final_demand",
        description="Which Relative Importance Table section to fetch.",
    )
    table_id: PpiRelativeImportanceTableId | None = Field(
        default=None,
        description="Specific BLS PPI Relative Importance table identifier; takes precedence over category.",
    )


class BlsPpiRelativeImportanceData(Data):
    """One ``(series, as-of date)`` relative-importance weight row."""

    model_config = ConfigDict(
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS PPI Relative Importance Tables",
                "$.description": (
                    "Producer Price Index relative importance (weights)."
                ),
                "$.gridData": {"w": 40, "h": 20},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "PPI",
            }
        }
    )

    date: dateType = Field(
        description="December reference date for which BLS computed this row's relative-importance weight.",
    )
    code: str | None = Field(
        default=None,
        description="Hierarchical PPI, commodity, or industry code.",
    )
    label: str | None = Field(
        default=None,
        description="Human-readable row label.",
    )
    relative_importance: float | None = Field(
        default=None,
        description="Relative-importance weight, as a percentage of the parent index.",
    )
    table_id: str = Field(
        description="Source BLS table id.",
        json_schema_extra=_HIDE,
    )
    table_name: str = Field(
        description="Full BLS table title.",
        json_schema_extra=_HIDE,
    )


class BlsPpiRelativeImportanceFetcher(
    Fetcher[
        BlsPpiRelativeImportanceQueryParams,
        list[BlsPpiRelativeImportanceData],
    ]
):
    """BLS PPI Relative Importance Table Fetcher."""

    require_credentials = False

    @staticmethod
    def transform_query(
        params: dict[str, Any],
    ) -> BlsPpiRelativeImportanceQueryParams:
        """Validate and coerce the query."""
        return BlsPpiRelativeImportanceQueryParams(**params)

    @staticmethod
    def extract_data(
        query: BlsPpiRelativeImportanceQueryParams,
        credentials: dict[str, str] | None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Fetch the relative-importance XLSX and parse it into long-form rows."""
        table_id = query.table_id or PPI_RELATIVE_IMPORTANCE_CANONICAL[query.category]
        entry = PPI_RELATIVE_IMPORTANCE_TABLES.get(table_id)
        if entry is None:  # pragma: no cover -- ``Literal`` constrains the input
            raise OpenBBError(f"Unknown relative-importance table_id '{table_id}'.")
        return _fetch_xlsx_table(table_id, entry["url"], entry["label"])

    @staticmethod
    def transform_data(
        query: BlsPpiRelativeImportanceQueryParams,
        data: dict[str, Any],
        **kwargs: Any,
    ) -> list[BlsPpiRelativeImportanceData]:
        """Coerce parsed rows into ``BlsPpiRelativeImportanceData``."""
        rows = data.get("rows", [])
        if not rows:
            raise EmptyDataError(
                f"No rows parsed from BLS PPI table '{data.get('table_id', '?')}'."
            )
        return [BlsPpiRelativeImportanceData.model_validate(r) for r in rows]


def _fetch_xlsx_table(table_id: str, url: str, label: str) -> dict[str, Any]:
    """Download a relative-importance XLSX and return long-form rows."""
    import io

    import openpyxl
    import requests

    resp = requests.get(url, headers=_HEADERS, timeout=60)
    if resp.status_code != 200:
        raise OpenBBError(f"BLS returned HTTP {resp.status_code} fetching {url}.")

    wb = openpyxl.load_workbook(
        io.BytesIO(resp.content), read_only=True, data_only=True
    )
    sheet = wb[wb.sheetnames[0]]
    raw_rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    headers, data_start = _detect_header_row(raw_rows)
    if not headers:
        return {"rows": [], "table_id": table_id, "table_name": label}

    value_cols = _detect_value_cols(headers)
    code_col_indices = {i for i, h in enumerate(headers) if h.lower().endswith("code")}
    label_col = _first_label_col(headers)

    out: list[dict[str, Any]] = []
    for row in raw_rows[data_start:]:
        for record in _row_to_records(
            row,
            headers,
            code_col_indices,
            label_col,
            value_cols,
            table_id,
            label,
        ):
            out.append(record)

    out.sort(
        key=lambda r: (r["date"], r.get("code") or "", r.get("label") or ""),
        reverse=False,
    )

    return {"rows": out, "table_id": table_id, "table_name": label}


def _detect_header_row(rows: list[tuple]) -> tuple[list[str], int]:
    """Locate the header row and return ``(column_names, next_data_index)``."""
    for idx, row in enumerate(rows):
        if len(row) < 2 or row[1] is None:
            continue
        cell = str(row[1]).strip().lower()
        if cell.endswith("code"):
            headers = [
                str(c).strip() if c is not None else f"col_{i}"
                for i, c in enumerate(row)
            ]
            return headers, idx + 1
    return [], 0


def _detect_value_cols(headers: list[str]) -> list[tuple[int, dateType]]:
    """Find columns whose header carries a December reference year."""
    out: list[tuple[int, dateType]] = []
    for i, h in enumerate(headers):
        match = _YEAR_RE.search(h)
        if match is None:
            continue
        year = int(match.group(1))
        out.append((i, dateType(year, 12, 1)))
    return out


def _first_label_col(headers: list[str]) -> int | None:
    """Locate the BLS ``Index`` column — the row title in the XLSX."""
    for i, h in enumerate(headers):
        if h.lower() in {"index", "indexes"}:
            return i
    return None


def _row_to_records(
    row: tuple,
    headers: list[str],
    code_col_indices: set[int],
    label_col: int | None,
    value_cols: list[tuple[int, dateType]],
    table_id: str,
    table_name: str,
) -> list[dict[str, Any]]:
    """Pivot one XLSX row into one record per (series, December reference year)."""
    if not row or not value_cols:
        return []

    code: str | None = None
    for i in code_col_indices:
        if i >= len(row):
            continue
        cell = row[i]
        if cell is None:
            continue
        cleaned = cell.strip() if isinstance(cell, str) else cell
        if cleaned == "" or not isinstance(cleaned, str):
            continue
        code = cleaned
        break

    label: str | None = None
    if label_col is not None and label_col < len(row):
        raw = row[label_col]
        if isinstance(raw, str):
            stripped = raw.strip()
            if stripped:
                label = stripped

    has_any_value = False
    parsed: list[tuple[dateType, float]] = []
    for i, as_of in value_cols:
        if i >= len(row):
            continue
        raw = row[i]
        if raw is None:
            continue
        if isinstance(raw, str):
            cleaned: Any = raw.strip()
            if cleaned in ("", "-"):
                continue
            try:
                value = float(cleaned)
            except ValueError:
                continue
        elif isinstance(raw, (int, float)):
            value = float(raw)
        else:
            continue
        has_any_value = True
        parsed.append((as_of, value))

    if not has_any_value:
        return []
    if code is None and label is None:
        return []

    return [
        {
            "date": as_of,
            "code": code,
            "label": label,
            "relative_importance": value,
            "table_id": table_id,
            "table_name": table_name,
        }
        for as_of, value in parsed
    ]
