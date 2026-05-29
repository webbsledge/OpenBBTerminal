"""BLS CPI relative importance and weights."""

from __future__ import annotations

import re
from datetime import date as dateType
from typing import Any, Literal

from openbb_core.app.model.abstract.error import OpenBBError
from openbb_core.provider.abstract.data import Data
from openbb_core.provider.abstract.fetcher import Fetcher
from openbb_core.provider.abstract.query_params import QueryParams
from openbb_core.provider.utils.errors import EmptyDataError
from pydantic import ConfigDict, Field, field_serializer

from openbb_bls.utils.constants import BLS_USER_AGENT

CpiRelativeImportanceTable = Literal[1, 2, 3, 4, 5, 6, 7]

_HEADERS = {
    "User-Agent": BLS_USER_AGENT,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
_HIDE: dict[str, Any] = {"x-widget_config": {"hide": True}}
_XLSX_MAGIC = b"PK\x03\x04"


class BlsCpiRelativeImportanceQueryParams(QueryParams):
    """BLS CPI Relative Importance Query Parameters."""

    __json_schema_extra__ = {
        "table": {
            "x-widget_config": {
                "options": [
                    {
                        "label": "Table 1 — U.S. City Average (CPI-U & CPI-W)",
                        "value": 1,
                    },
                    {
                        "label": "Table 2 — Selected local areas (CPI-U & CPI-W)",
                        "value": 2,
                    },
                    {
                        "label": "Table 3 — Selected local areas (CPI-U & CPI-W)",
                        "value": 3,
                    },
                    {
                        "label": "Table 4 — Selected local areas (CPI-U & CPI-W)",
                        "value": 4,
                    },
                    {
                        "label": "Table 5 — Selected local areas (CPI-U & CPI-W)",
                        "value": 5,
                    },
                    {
                        "label": "Table 6 — Selected local areas (CPI-U & CPI-W)",
                        "value": 6,
                    },
                    {
                        "label": "Table 7 — Selected local areas (CPI-U & CPI-W)",
                        "value": 7,
                    },
                ]
            }
        },
    }

    year: int | None = Field(
        default=None,
        description="Year of the relative-importance file to fetch; None walks back until a published file is found.",
    )
    table: int = Field(
        default=1,
        ge=1,
        le=7,
        description="Which of the 7 sheets in the relative-importance XLSX to return.",
    )


class BlsCpiRelativeImportanceData(Data):
    """One ``(item, basket)`` relative-importance row from the CPI weight file."""

    model_config = ConfigDict(
        json_schema_extra={
            "x-widget_config": {
                "$.name": "BLS CPI Relative Importance & Weights",
                "$.description": "Consumer Price Index relative importance (weights).",
                "$.gridData": {"w": 40, "h": 20},
                "$.refetchInterval": False,
                "$.source": ["BLS"],
                "$.category": "Economy",
                "$.subCategory": "CPI",
            }
        }
    )

    date: dateType = Field(
        description="December reference date for the weights.",
    )
    label: str | None = Field(
        default=None,
        description="BLS item or expenditure-category label.",
        json_schema_extra={
            "x-widget_config": {
                "renderFn": "hoverCard",
                "renderFnParams": {
                    "hoverCard": {
                        "cellField": "value",
                        "markdown": "{footnote}",
                    }
                },
            }
        },
    )
    level: int | None = Field(
        default=None,
        description="Indent depth from the source XLSX.",
    )
    row_index: int = Field(
        description="Sequential ordering index assigned during parsing.",
        json_schema_extra=_HIDE,
    )
    section: str | None = Field(
        default=None,
        description="Section the row appears under in the source XLSX (Expenditure category or Special aggregate indexes).",
    )
    area: str | None = Field(
        default=None,
        description="Geographic area the weight refers to (e.g., U.S. City Average or a selected local area).",
    )
    footnote: str | None = Field(
        default=None,
        description="Resolved footnote text for any (N) markers attached to the row's label.",
        json_schema_extra=_HIDE,
    )

    @field_serializer("label", when_used="json")
    def _serialize_label_with_footnote(self, value: str | None) -> Any:
        """Wrap the label with footnote text for the hover-card renderer."""
        if value and self.footnote:
            return {"value": value, "footnote": self.footnote}
        return value

    basket: str = Field(
        description="Which CPI basket the weight refers to (CPI-U or CPI-W).",
    )
    relative_importance: float | None = Field(
        default=None,
        description="Weight as a percentage of the parent index.",
        json_schema_extra={
            "x-widget_config": {
                "headerName": "Relative Importance",
                "cellDataType": "number",
                "formatterFn": "percent",
                "renderFn": "greenRed",
            }
        },
    )
    table_id: str = Field(
        description="Source table identifier.",
        json_schema_extra=_HIDE,
    )
    table_name: str = Field(
        description="Full BLS table title.",
        json_schema_extra=_HIDE,
    )


class BlsCpiRelativeImportanceFetcher(
    Fetcher[
        BlsCpiRelativeImportanceQueryParams,
        list[BlsCpiRelativeImportanceData],
    ]
):
    """BLS CPI Relative Importance Fetcher."""

    require_credentials = False

    @staticmethod
    def transform_query(
        params: dict[str, Any],
    ) -> BlsCpiRelativeImportanceQueryParams:
        """Validate and coerce the query."""
        return BlsCpiRelativeImportanceQueryParams(**params)

    @staticmethod
    def extract_data(
        query: BlsCpiRelativeImportanceQueryParams,
        credentials: dict[str, str] | None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Fetch the requested yearly XLSX and parse one of its 7 sheets."""
        if query.year is None:
            year, content = _discover_latest_ri()
        else:
            content = _fetch_ri_xlsx(query.year)
            if content is None:
                raise OpenBBError(
                    f"BLS CPI Relative Importance file for {query.year} not found."
                )
            year = query.year
        return _parse_ri_table(content, year, query.table)

    @staticmethod
    def transform_data(
        query: BlsCpiRelativeImportanceQueryParams,
        data: dict[str, Any],
        **kwargs: Any,
    ) -> list[BlsCpiRelativeImportanceData]:
        """Coerce parsed rows into ``BlsCpiRelativeImportanceData``."""
        rows = data.get("rows", [])
        if not rows:
            raise EmptyDataError(
                f"No rows parsed from CPI RI table '{data.get('table_id', '?')}'."
            )
        return [BlsCpiRelativeImportanceData.model_validate(r) for r in rows]


def _ri_url(year: int) -> str:
    """Canonical CPI relative-importance XLSX URL for ``year``."""
    return f"https://www.bls.gov/cpi/tables/relative-importance/{year}.xlsx"


def _fetch_ri_xlsx(year: int) -> bytes | None:
    """Fetch one yearly RI XLSX; ``None`` if BLS redirects or returns non-XLSX."""
    import requests

    resp = requests.get(
        _ri_url(year), headers=_HEADERS, timeout=60, allow_redirects=False
    )
    if resp.status_code in (301, 302, 303, 307, 308, 404):
        return None
    if resp.status_code != 200:
        raise OpenBBError(
            f"BLS returned HTTP {resp.status_code} fetching {_ri_url(year)}."
        )
    if not resp.content.startswith(_XLSX_MAGIC):
        return None
    return resp.content


def _discover_latest_ri() -> tuple[int, bytes]:
    """Walk back from the current calendar year to find the most recent RI XLSX."""
    today = dateType.today()
    for year in range(today.year, today.year - 6, -1):
        content = _fetch_ri_xlsx(year)
        if content is not None:
            return year, content
    raise OpenBBError(
        "Could not locate a recent CPI Relative Importance XLSX — the BLS "
        "publication URL pattern may have changed."
    )


def _parse_ri_table(content: bytes, year: int, table_number: int) -> dict[str, Any]:
    """Open the XLSX and emit long-form rows for the chosen table sheet."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_name = f"Table {table_number}"
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise OpenBBError(
            f"CPI Relative Importance {year} does not contain sheet '{sheet_name}'."
        )
    sheet = wb[sheet_name]
    raw_rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    table_id = f"cpi-ri-{year}-t{table_number}"
    table_name = f"CPI Relative Importance {year} — Table {table_number}"

    super_headers: tuple = ()
    sub_headers: tuple = ()
    data_start = 0
    for idx, row in enumerate(raw_rows):
        if row and isinstance(row[0], str) and row[0].strip().lower() == "indent level":
            super_headers = row
            sub_headers = raw_rows[idx + 1] if idx + 1 < len(raw_rows) else ()
            data_start = idx + 2
            break
    if not super_headers:
        return {"rows": [], "table_id": table_id, "table_name": table_name}

    basket_cols = _classify_ri_columns(super_headers, sub_headers)
    if not basket_cols:
        return {"rows": [], "table_id": table_id, "table_name": table_name}

    body_rows = raw_rows[data_start:]
    footnotes = _parse_footnotes_block(body_rows)

    reference_date = dateType(year - 1, 12, 1)
    per_series: dict[tuple[str | None, str], list[dict[str, Any]]] = {}
    current_section: str | None = None
    for source_idx, row in enumerate(body_rows):
        if _is_section_header_row(row):
            label_cell = row[1] if len(row) > 1 and isinstance(row[1], str) else None
            if label_cell:
                current_section = label_cell.strip()
            continue
        if row[0] in (None, ""):
            continue
        try:
            level = int(str(row[0]).strip())
        except (TypeError, ValueError):
            continue
        label = row[1].strip() if isinstance(row[1], str) and row[1].strip() else None
        if label is None:
            continue
        footnote_text = _resolve_label_footnotes(label, footnotes)
        for col_idx, area, basket in basket_cols:
            cell = row[col_idx]
            value: float | None
            if cell is None:
                value = None
            elif isinstance(cell, (int, float)):
                value = float(cell)
            else:
                text = str(cell).strip()
                if text in ("", "-"):
                    value = None
                else:
                    try:
                        value = float(text)
                    except ValueError:
                        value = None
            per_series.setdefault((area, basket), []).append(
                {
                    "date": reference_date,
                    "label": label,
                    "level": level,
                    "section": current_section,
                    "area": area,
                    "basket": basket,
                    "relative_importance": value,
                    "footnote": footnote_text,
                    "_source_idx": source_idx,
                    "table_id": table_id,
                    "table_name": table_name,
                }
            )

    out: list[dict[str, Any]] = []
    row_index = 0
    for bucket in per_series.values():
        bucket.sort(key=lambda r: r["_source_idx"])
        for entry in bucket:
            row_index += 1
            entry.pop("_source_idx", None)
            entry["row_index"] = row_index
            out.append(entry)

    return {"rows": out, "table_id": table_id, "table_name": table_name}


_FOOTNOTE_DEF_RE = re.compile(r"^\(([^)]+)\)\s*(.*)$")
_LABEL_REF_RE = re.compile(r"\(([^)]+)\)")


def _parse_footnotes_block(rows: list[tuple]) -> dict[str, str]:
    """Build ``{'(N)': 'text'}`` from trailing footnote-definition rows."""
    out: dict[str, str] = {}
    for row in rows:
        if not row:
            continue
        for cell in row:
            if not isinstance(cell, str):
                continue
            text = cell.strip()
            match = _FOOTNOTE_DEF_RE.match(text)
            if match is None:
                continue
            marker = f"({match.group(1)})"
            body = match.group(2).strip() or text
            out[marker] = body
            break
    return out


def _resolve_label_footnotes(
    label: str | None, footnotes: dict[str, str]
) -> str | None:
    """Concatenate footnote text for any ``(N)`` markers in ``label``."""
    if not label or not footnotes:
        return None
    matched: list[str] = []
    seen: set[str] = set()
    for match in _LABEL_REF_RE.finditer(label):
        marker = match.group(0)
        if marker not in footnotes or marker in seen:
            continue
        seen.add(marker)
        matched.append(footnotes[marker])
    if not matched:
        return None
    return "\n\n".join(matched)


def _is_section_header_row(row: tuple) -> bool:
    """Detect a 'Expenditure category' / 'Special aggregate indexes' section header."""
    if not row or len(row) < 2:
        return False
    label_cell = row[1]
    if not isinstance(label_cell, str):
        return False
    label_norm = label_cell.strip().lower()
    return label_norm in (
        "expenditure category",
        "special aggregate indexes",
        "items",
    )


def _classify_ri_columns(
    super_headers: tuple, sub_headers: tuple
) -> list[tuple[int, str | None, str]]:
    """Locate the CPI-U / CPI-W value columns and the area each belongs to.

    Table 1 carries a single area (U.S. City Average); Tables 2-7 repeat the
    CPI-U / CPI-W pair once per selected local area, so the area is read from
    the super-header to keep each ``(area, basket)`` series distinct.
    """
    out: list[tuple[int, str | None, str]] = []
    n = max(len(super_headers), len(sub_headers))
    for i in range(n):
        sub = sub_headers[i] if i < len(sub_headers) else None
        if not isinstance(sub, str):
            continue
        token = sub.strip().upper()
        if token not in ("CPI-U", "CPI-W"):
            continue
        sup = super_headers[i] if i < len(super_headers) else None
        area = (
            " ".join(str(sup).split()) if isinstance(sup, str) and sup.strip() else None
        )
        out.append((i, area, token))
    return out
