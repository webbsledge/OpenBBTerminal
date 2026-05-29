"""Unit tests for ``openbb_bls.utils.jolts_tables`` parser internals."""

from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path

import pytest
from openbb_core.app.model.abstract.error import OpenBBError
from openpyxl import Workbook

from openbb_bls.utils import jolts_tables as jt

_FIXTURES = Path(__file__).parent / "fixtures"


class _FakeResp:
    """Minimal ``requests.Response`` stand-in for fetcher unit tests."""

    def __init__(self, content: bytes = b"", status_code: int = 200):
        self.content = content
        self.status_code = status_code


def test_list_change_analysis_tables_returns_all_entries():
    """Returns one entry per national + state table with the correct URL stems."""
    rows = jt.list_change_analysis_tables()
    national = [r for r in rows if r["scope"] == "national"]
    state = [r for r in rows if r["scope"] == "state"]
    assert len(national) == 12
    assert len(state) == 10
    assert any(r["url"].endswith("/jlt_table1.txt") for r in national)
    assert any(r["url"].endswith("/jltst_table1.txt") for r in state)
    otm_national = [r for r in national if r["period"] == "over-the-month"]
    oty_national = [r for r in national if r["period"] == "over-the-year"]
    assert len(otm_national) == 6
    assert len(oty_national) == 6


def test_fetch_change_analysis_txt_national_success(monkeypatch):
    """National fetch hits the jolts URL and returns decoded text."""
    captured: dict = {}

    def _fake_get(url, **kwargs):
        captured["url"] = url
        return _FakeResp(content=b"hello", status_code=200)

    import requests

    monkeypatch.setattr(requests, "get", _fake_get)
    out = jt.fetch_change_analysis_txt("national", 1)
    assert out == "hello"
    assert captured["url"].endswith("/jlt_table1.txt")


def test_fetch_change_analysis_txt_state_uses_jltst_url(monkeypatch):
    """State fetch hits the jltst URL stem."""
    captured: dict = {}

    def _fake_get(url, **kwargs):
        captured["url"] = url
        return _FakeResp(content=b"x", status_code=200)

    import requests

    monkeypatch.setattr(requests, "get", _fake_get)
    jt.fetch_change_analysis_txt("state", 2)
    assert captured["url"].endswith("/jltst_table2.txt")


def test_fetch_change_analysis_txt_unknown_scope_raises():
    """An unknown scope value raises an OpenBBError."""
    with pytest.raises(OpenBBError, match="Unknown JOLTS scope"):
        jt.fetch_change_analysis_txt("planetary", 1)


def test_fetch_change_analysis_txt_non_200_raises(monkeypatch):
    """Non-200 responses surface as OpenBBError including the status code."""

    def _fake_get(url, **kwargs):
        return _FakeResp(status_code=503)

    import requests

    monkeypatch.setattr(requests, "get", _fake_get)
    with pytest.raises(OpenBBError, match="HTTP 503"):
        jt.fetch_change_analysis_txt("national", 1)


def test_fetch_change_analysis_txt_latin1_decode(monkeypatch):
    """Latin-1-only bytes still decode via the fallback path."""

    class _Resp:
        status_code = 200
        content = "café".encode("latin-1")

    import requests

    monkeypatch.setattr(requests, "get", lambda *a, **k: _Resp())
    text = jt.fetch_change_analysis_txt("national", 1)
    assert "caf" in text


def test_month_to_int_short_and_full():
    """Short and full month spellings both resolve."""
    assert jt._month_to_int("Jan") == 1
    assert jt._month_to_int("january") == 1
    assert jt._month_to_int("DEC.") == 12


def test_month_to_int_unknown_returns_none():
    """Unknown month token returns ``None``."""
    assert jt._month_to_int("blargle") is None


def test_parse_period_invalid_year_returns_none_pair():
    """Bad year text triggers the ValueError fallback path."""
    s, e = jt._parse_period("Jan", "not-year", "Feb", "2025")
    assert s is None and e is None


def test_parse_period_partial_unknown_month_returns_none():
    """An unknown end-month name gives a ``None`` end-date but valid start."""
    s, e = jt._parse_period("Jan", "2025", "Blarg", "2025")
    assert s == date(2025, 1, 1)
    assert e is None


def test_parse_change_analysis_empty_input_raises():
    """Empty TXT input raises an OpenBBError."""
    with pytest.raises(OpenBBError, match="Empty"):
        jt.parse_change_analysis("", "national", 1)


def test_parse_change_analysis_missing_title_raises():
    """First line missing the TABLE header raises an OpenBBError."""
    with pytest.raises(OpenBBError, match="missing the"):
        jt.parse_change_analysis("not a header\nbody\n", "national", 1)


def test_parse_change_analysis_national_basic():
    """National TXT fixture parses, captures sections + SA flag + significance."""
    text = (_FIXTURES / "jolts_national_t1.txt").read_text()
    out = jt.parse_change_analysis(text, "national", 1)
    assert out["table_id"] == "jolts-national-t1"
    rows = out["rows"]
    assert rows
    first = rows[0]
    assert first["seasonally_adjusted"] is True
    assert first["period"] == "over-the-month"
    assert first["measure"].lower().startswith("job openings")
    assert first["period_start"] == date(2026, 2, 1)
    assert first["period_end"] == date(2026, 3, 1)
    assert first["source_date"] == date(2026, 5, 5)
    sections = {r["section"] for r in rows if r["section"]}
    assert "Industry" in sections
    assert "Region" in sections
    sig = [r for r in rows if r["rate_passes_significance"]]
    assert sig


def test_parse_change_analysis_state_fixture_runs():
    """State TXT fixture parses cleanly."""
    text = (_FIXTURES / "jolts_state_t1.txt").read_text()
    out = jt.parse_change_analysis(text, "state", 1)
    assert out["rows"]


def test_parse_change_analysis_not_seasonally_adjusted_flag():
    """``not seasonally adjusted`` token on line 2 flips the SA flag to False."""
    txt = (
        "TABLE 1: Job openings estimated rate and level changes "
        "between February 2026 and March 2026, and test of significance\n"
        "         not seasonally adjusted\n"
        "\n"
        "Total nonfarm                                        -0.1           0.3                      -56            517\n"
        "  SOURCE: BLS, May 5, 2026.\n"
    )
    out = jt.parse_change_analysis(txt, "national", 1)
    assert out["rows"][0]["seasonally_adjusted"] is False


def test_parse_change_analysis_period_kind_year():
    """A start/end > 11 months apart classifies as over-the-year."""
    txt = (
        "TABLE 7: Job openings estimated rate and level changes "
        "between March 2025 and March 2026, and test of significance, seasonally adjusted\n"
        "\n"
        "\n"
        "Total nonfarm                                        -0.1           0.3                      -56            517\n"
    )
    out = jt.parse_change_analysis(txt, "national", 7)
    assert out["rows"][0]["period"] == "over-the-year"


def test_parse_change_analysis_sa_field_none_when_unspecified():
    """Missing SA hint in line 2 yields ``seasonally_adjusted=None``."""
    txt = (
        "TABLE 1: Job openings estimated rate and level changes "
        "between February 2026 and March 2026, and test of significance\n"
        "\n"
        "\n"
        "Total nonfarm                                        -0.1           0.3                      -56            517\n"
    )
    out = jt.parse_change_analysis(txt, "national", 1)
    assert out["rows"][0]["seasonally_adjusted"] is None


def test_parse_change_analysis_skips_note_lines_and_break_on_source():
    """NOTE / asterisk / 'is used' lines skip; '*NOTE' and SOURCE terminate parsing."""
    txt = (
        "TABLE 2: Hires estimated rate and level changes between "
        "February 2026 and March 2026, and test of significance, seasonally adjusted\n"
        "\n"
        "\n"
        "  Total private                                      -0.1           0.3                      -54            490\n"
        "NOTE: bare note kept inline\n"
        "* a stray asterisk\n"
        "is used in testing the over-the-month change.\n"
        "  Construction                                       0.2           1.0                       23             91\n"
        "*NOTE: this terminates\n"
        "  After break                                        0.4           1.0                        5             10\n"
        "  SOURCE: BLS, May 5, 2026.\n"
    )
    out = jt.parse_change_analysis(txt, "national", 2)
    labels = [r["label"] for r in out["rows"]]
    assert "Total private" in labels
    assert "Construction" in labels
    assert "After break" not in labels


def test_parse_change_analysis_skips_unparseable_source_date():
    """A SOURCE line with an unrecognised month yields ``source_date=None``."""
    txt = (
        "TABLE 1: Job openings estimated rate and level changes "
        "between February 2026 and March 2026, and test of significance, "
        "seasonally adjusted\n"
        "\n"
        "\n"
        "Total nonfarm                                        -0.1           0.3                      -56            517\n"
        "  SOURCE: BLS, NotAMonth 5, 2026.\n"
    )
    out = jt.parse_change_analysis(txt, "national", 1)
    assert out["rows"][0]["source_date"] is None


def test_parse_change_analysis_skips_invalid_source_day():
    """A SOURCE line with an out-of-range day yields ``source_date=None``."""
    txt = (
        "TABLE 1: Job openings estimated rate and level changes "
        "between February 2026 and March 2026, and test of significance, "
        "seasonally adjusted\n"
        "\n"
        "\n"
        "Total nonfarm                                        -0.1           0.3                      -56            517\n"
        "  SOURCE: BLS, May 99, 2026.\n"
    )
    out = jt.parse_change_analysis(txt, "national", 1)
    assert out["rows"][0]["source_date"] is None


def test_parse_change_analysis_blank_label_line_skipped():
    """A data line whose label strips to empty is skipped (no record)."""
    txt = (
        "TABLE 1: Job openings estimated rate and level changes "
        "between February 2026 and March 2026, and test of significance, "
        "seasonally adjusted\n"
        "\n"
        "\n"
        # Leading-whitespace-only label: the regex captures a lone space, which
        # strips to '' and triggers the ``if not label: continue`` guard.
        "          -0.1           0.3                      -56            517\n"
        "Total nonfarm                                        -0.1           0.3                      -56            517\n"
    )
    out = jt.parse_change_analysis(txt, "national", 1)
    labels = [r["label"] for r in out["rows"]]
    assert labels == ["Total nonfarm"]


def test_cell_to_float_variants():
    """``_cell_to_float`` collapses every supported input type."""
    assert jt._cell_to_float(None) is None
    assert jt._cell_to_float(True) is None
    assert jt._cell_to_float(3) == 3.0
    assert jt._cell_to_float(2.5) == 2.5
    assert jt._cell_to_float("1,234.5") == 1234.5
    assert jt._cell_to_float("  ") is None
    assert jt._cell_to_float("abc") is None
    assert jt._cell_to_float(b"3.0") is None


def test_fetch_revision_xlsx_sa(monkeypatch):
    """SA fetcher builds the sa-revision URL and returns content bytes."""
    captured: dict = {}
    payload = b"PK\x03\x04data"

    def _fake_get(url, **kwargs):
        captured["url"] = url
        return _FakeResp(content=payload, status_code=200)

    import requests

    monkeypatch.setattr(requests, "get", _fake_get)
    out = jt.fetch_revision_xlsx(seasonally_adjusted=True)
    assert out == payload
    assert "sa-revision-tables.xlsx" in captured["url"]


def test_fetch_revision_xlsx_nsa(monkeypatch):
    """NSA fetcher builds the nsa-revision URL."""
    captured: dict = {}

    def _fake_get(url, **kwargs):
        captured["url"] = url
        return _FakeResp(content=b"PK\x03\x04ok", status_code=200)

    import requests

    monkeypatch.setattr(requests, "get", _fake_get)
    jt.fetch_revision_xlsx(seasonally_adjusted=False)
    assert "nsa-revision-tables.xlsx" in captured["url"]


def test_fetch_revision_xlsx_non_200_raises(monkeypatch):
    """Non-200 HTTP surfaces as OpenBBError."""

    def _fake_get(url, **kwargs):
        return _FakeResp(status_code=404)

    import requests

    monkeypatch.setattr(requests, "get", _fake_get)
    with pytest.raises(OpenBBError, match="HTTP 404"):
        jt.fetch_revision_xlsx(seasonally_adjusted=True)


def test_fetch_revision_xlsx_non_xlsx_payload_raises(monkeypatch):
    """A 200 response that lacks the XLSX magic bytes raises OpenBBError."""

    def _fake_get(url, **kwargs):
        return _FakeResp(content=b"<html>", status_code=200)

    import requests

    monkeypatch.setattr(requests, "get", _fake_get)
    with pytest.raises(OpenBBError, match="non-XLSX content"):
        jt.fetch_revision_xlsx(seasonally_adjusted=False)


def test_parse_revision_xlsx_existing_fixture():
    """Bundled mini SA workbook parses into wide rows."""
    content = (_FIXTURES / "jolts_sa_rev_mini.xlsx").read_bytes()
    rows = jt.parse_revision_xlsx(content, seasonally_adjusted=True)
    assert rows
    sample = rows[0]
    assert sample["industry_code"] == "00"
    assert sample["seasonally_adjusted"] is True
    assert "level_1st" in sample
    assert sample["table_id"] == "jolts-revisions-sa"


def test_parse_revision_xlsx_nsa_title_uses_nsa_token():
    """NSA-token title yields the NSA table_id + table_title."""
    content = (_FIXTURES / "jolts_nsa_rev_mini.xlsx").read_bytes()
    rows = jt.parse_revision_xlsx(content, seasonally_adjusted=False)
    assert rows
    assert rows[0]["table_id"] == "jolts-revisions-nsa"
    assert "not seasonally" in rows[0]["table_title"]


def test_parse_revision_xlsx_skips_short_sheet():
    """Sheets with fewer than 5 rows are skipped without raising."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("short")
    ws.append(["TOTAL NONFARM, seasonally adjusted, in thousands"])
    ws.append(["only two rows"])
    buf = io.BytesIO()
    wb.save(buf)
    rows = jt.parse_revision_xlsx(buf.getvalue(), seasonally_adjusted=True)
    assert rows == []


def test_parse_revision_xlsx_industry_name_no_marker():
    """Sheet title without an SA marker falls through to the raw title."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("99")
    ws.append(["INDUSTRY MYSTERY"] + [None] * 53)
    measures = (
        "JOB OPENINGS",
        "HIRES",
        "TOTAL SEPARATIONS",
        "QUITS",
        "LAYOFFS & DISCHARGES",
        "OTHER SEPARATIONS",
    )
    row1: list = []
    for measure in measures:
        row1.extend(["Month", measure, None, None, None, None, None, None, " "])
    ws.append(row1)
    ws.append([None] * 54)
    ws.append([None] * 54)
    data_row: list = []
    date_val = datetime(2024, 1, 1)
    for i, _ in enumerate(measures):
        data_row.extend([date_val, 1, 2, 3, 4, 5, 6, 7, " "])
    ws.append(data_row)
    buf = io.BytesIO()
    wb.save(buf)
    rows = jt.parse_revision_xlsx(buf.getvalue(), seasonally_adjusted=True)
    assert rows and rows[0]["industry_name"] == "INDUSTRY MYSTERY"


def test_parse_revision_xlsx_skips_non_string_title():
    """A non-string title cell yields industry_name=''."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("01")
    ws.append([12345] + [None] * 53)
    measures = (
        "JOB OPENINGS",
        "HIRES",
        "TOTAL SEPARATIONS",
        "QUITS",
        "LAYOFFS & DISCHARGES",
        "OTHER SEPARATIONS",
    )
    row1: list = []
    for measure in measures:
        row1.extend(["Month", measure, None, None, None, None, None, None, " "])
    ws.append(row1)
    ws.append([None] * 54)
    ws.append([None] * 54)
    data_row: list = []
    for _ in measures:
        data_row.extend([datetime(2024, 1, 1), 1, 2, 3, 4, 5, 6, 7, " "])
    ws.append(data_row)
    buf = io.BytesIO()
    wb.save(buf)
    rows = jt.parse_revision_xlsx(buf.getvalue(), seasonally_adjusted=True)
    assert rows and rows[0]["industry_name"] == ""


def test_parse_revision_xlsx_drops_all_none_data_row():
    """Rows where every revision cell is None are skipped."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("00")
    ws.append(["TOTAL NONFARM, seasonally adjusted, in thousands"] + [None] * 53)
    measures = (
        "JOB OPENINGS",
        "HIRES",
        "TOTAL SEPARATIONS",
        "QUITS",
        "LAYOFFS & DISCHARGES",
        "OTHER SEPARATIONS",
    )
    row1: list = []
    for measure in measures:
        row1.extend(["Month", measure, None, None, None, None, None, None, " "])
    ws.append(row1)
    ws.append([None] * 54)
    ws.append([None] * 54)
    data_row: list = []
    for _ in measures:
        data_row.extend(
            [datetime(2024, 1, 1), None, None, None, None, None, None, None, " "]
        )
    ws.append(data_row)
    buf = io.BytesIO()
    wb.save(buf)
    rows = jt.parse_revision_xlsx(buf.getvalue(), seasonally_adjusted=True)
    assert rows == []


def test_parse_revision_xlsx_skips_non_date_month_cell():
    """Month cells that aren't dates/datetimes are skipped."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("00")
    ws.append(["TOTAL NONFARM, seasonally adjusted, in thousands"] + [None] * 53)
    measures = (
        "JOB OPENINGS",
        "HIRES",
        "TOTAL SEPARATIONS",
        "QUITS",
        "LAYOFFS & DISCHARGES",
        "OTHER SEPARATIONS",
    )
    row1: list = []
    for measure in measures:
        row1.extend(["Month", measure, None, None, None, None, None, None, " "])
    ws.append(row1)
    ws.append([None] * 54)
    ws.append([None] * 54)
    bad_row: list = []
    for _ in measures:
        bad_row.extend(["not a date", 1, 2, 3, 4, 5, 6, 7, " "])
    ws.append(bad_row)
    good_row: list = []
    for _ in measures:
        good_row.extend([datetime(2024, 1, 1), 1, 2, 3, 4, 5, 6, 7, " "])
    ws.append(good_row)
    buf = io.BytesIO()
    wb.save(buf)
    rows = jt.parse_revision_xlsx(buf.getvalue(), seasonally_adjusted=True)
    assert rows and all(r["date"] == date(2024, 1, 1) for r in rows)


class _FakeWS:
    """In-memory worksheet for parse_revision_xlsx tests."""

    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, values_only=False):
        """Yield raw tuple rows."""
        for r in self._rows:
            yield tuple(r)


class _FakeWB:
    """Workbook stand-in dispatching iter_rows from a dict of sheets."""

    def __init__(self, sheets):
        self._sheets = sheets
        self.sheetnames = list(sheets.keys())

    def __getitem__(self, name):
        return self._sheets[name]

    def close(self):
        """No-op."""


def _patch_openpyxl(monkeypatch, sheets):
    """Replace ``openpyxl.load_workbook`` so parse_revision_xlsx sees our sheets."""
    import openpyxl

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *a, **kwargs: _FakeWB(sheets),
    )


def _measure_header():
    """Build the 6-block measure header row used by JOLTS revision sheets."""
    measures = (
        "JOB OPENINGS",
        "HIRES",
        "TOTAL SEPARATIONS",
        "QUITS",
        "LAYOFFS & DISCHARGES",
        "OTHER SEPARATIONS",
    )
    row: list = []
    for m in measures:
        row.extend(["Month", m, None, None, None, None, None, None, " "])
    return row


def test_parse_revision_xlsx_short_row_breaks_loop(monkeypatch):
    """A row that ends before ``col_start - 1`` breaks the data loop early."""
    rows = [
        ("TOTAL NONFARM, seasonally adjusted, in thousands",) + (None,) * 53,
        tuple(_measure_header()),
        (None,) * 54,
        (None,) * 54,
        # Short row: only 2 cells — the 2nd block_start (col=10) triggers the
        # ``col_start - 1 >= len(row)`` break path.
        ("first", "second"),
    ]
    _patch_openpyxl(monkeypatch, {"00": _FakeWS(rows)})
    result = jt.parse_revision_xlsx(b"PK\x03\x04ignored", seasonally_adjusted=True)
    assert result == []


def test_parse_revision_xlsx_none_month_cell_skips(monkeypatch):
    """A row whose month-column cell is ``None`` is silently skipped."""
    header = _measure_header()
    row_template: list = []
    for _ in range(6):
        row_template.extend([None, 1, 2, 3, 4, 5, 6, 7, " "])
    rows = [
        ("TOTAL NONFARM, seasonally adjusted, in thousands",) + (None,) * 53,
        tuple(header),
        (None,) * 54,
        (None,) * 54,
        tuple(row_template),
    ]
    _patch_openpyxl(monkeypatch, {"00": _FakeWS(rows)})
    result = jt.parse_revision_xlsx(b"PK\x03\x04ignored", seasonally_adjusted=True)
    assert result == []


def test_parse_revision_xlsx_date_only_month_cell(monkeypatch):
    """``datetime.date`` (not datetime) month cells exercise the ``elif date`` branch."""
    header = _measure_header()
    row_template: list = []
    for _ in range(6):
        row_template.extend([date(2024, 1, 1), 1, 2, 3, 4, 5, 6, 7, " "])
    rows = [
        ("TOTAL NONFARM, seasonally adjusted, in thousands",) + (None,) * 53,
        tuple(header),
        (None,) * 54,
        (None,) * 54,
        tuple(row_template),
    ]
    _patch_openpyxl(monkeypatch, {"00": _FakeWS(rows)})
    result = jt.parse_revision_xlsx(b"PK\x03\x04ignored", seasonally_adjusted=True)
    assert result and result[0]["date"] == date(2024, 1, 1)


def test_parse_revision_xlsx_short_cell_block_assigns_none(monkeypatch):
    """A short trailing block forces the final revision cells to ``None``."""
    header = _measure_header()
    # 5 full blocks of (month, 7 cells, separator) + a sixth truncated block
    # that ends after the month + 2 cells, so 5 of the 7 field keys hit the
    # ``cell_col >= len(row)`` branch.
    row_template: list = []
    for _ in range(5):
        row_template.extend([date(2024, 1, 1), 1, 2, 3, 4, 5, 6, 7, None])
    row_template.extend([date(2024, 1, 1), 1, 2])
    rows = [
        ("TOTAL NONFARM, seasonally adjusted, in thousands",) + (None,) * 53,
        tuple(header),
        (None,) * 54,
        (None,) * 54,
        tuple(row_template),
    ]
    _patch_openpyxl(monkeypatch, {"00": _FakeWS(rows)})
    result = jt.parse_revision_xlsx(b"PK\x03\x04ignored", seasonally_adjusted=True)
    assert result
    last = [r for r in result if r["measure"] == "Other Separations"][0]
    assert last["level_benchmark"] is None
