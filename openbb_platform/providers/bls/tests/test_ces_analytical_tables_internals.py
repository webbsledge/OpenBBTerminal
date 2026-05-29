"""Unit tests for ``openbb_bls.utils.ces_analytical_tables`` internals."""

from __future__ import annotations

import io
from datetime import date

import pytest
from openbb_core.app.model.abstract.error import OpenBBError
from openbb_core.provider.utils.errors import EmptyDataError
from openpyxl import Workbook

import openbb_bls.models.ces_analytical_tables as cesm
from openbb_bls.utils import ces_analytical_tables as ces


class _FakeResp:
    """Minimal ``requests.Response`` stand-in."""

    def __init__(self, content: bytes = b"", status_code: int = 200):
        self.content = content
        self.status_code = status_code


def _wb(rows: list[list], sheet: str = "Sheet1") -> bytes:
    """Build a single-sheet XLSX from a list of row lists."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _wb_sheets(sheets: dict[str, list[list]]) -> bytes:
    """Build a multi-sheet XLSX from a mapping of sheet name to row lists."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# fetch_table_xlsx
# --------------------------------------------------------------------------


def test_fetch_table_xlsx_non_200_raises(monkeypatch):
    """A non-200 response raises OpenBBError."""
    import requests

    monkeypatch.setattr(requests, "get", lambda *a, **k: _FakeResp(status_code=503))
    with pytest.raises(OpenBBError, match="HTTP 503"):
        ces.fetch_table_xlsx("cesanatab1")


def test_fetch_table_xlsx_non_xlsx_raises(monkeypatch):
    """A 200 response without XLSX magic bytes raises OpenBBError."""
    import requests

    monkeypatch.setattr(requests, "get", lambda *a, **k: _FakeResp(content=b"<html>"))
    with pytest.raises(OpenBBError, match="non-XLSX content"):
        ces.fetch_table_xlsx("cesanatab1")


def test_fetch_table_xlsx_success(monkeypatch):
    """A valid XLSX payload is returned verbatim."""
    import requests

    payload = ces._XLSX_MAGIC + b"rest"
    monkeypatch.setattr(requests, "get", lambda *a, **k: _FakeResp(content=payload))
    assert ces.fetch_table_xlsx("cesanatab1") == payload


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------


def test_norm():
    """``_norm`` collapses nbsp/newlines and maps None to ''."""
    assert ces._norm(None) == ""
    assert ces._norm("a\xa0\xa0b\n c") == "a b c"


def test_to_value_all_branches():
    """``_to_value`` handles None, bool, numeric, padded/significant, N/A, text."""
    assert ces._to_value(None) == (None, False, None)
    assert ces._to_value(True) == (None, False, None)
    assert ces._to_value(5) == (5.0, False, None)
    assert ces._to_value(2.5) == (2.5, False, None)
    assert ces._to_value("\xa0\xa0123.0*") == (123.0, True, None)
    assert ces._to_value(".") == (None, False, None)
    assert ces._to_value("-") == (None, False, None)
    assert ces._to_value("   ") == (None, False, None)
    assert ces._to_value("1,234.5") == (1234.5, False, None)
    assert ces._to_value("foo") == (None, False, "foo")
    # significance marker on an otherwise-empty cell
    assert ces._to_value(".*") == (None, True, None)


def test_to_mon_year_branches():
    """``_to_mon_year`` parses ``MON-YYYY`` and rejects blanks / bad months."""
    assert ces._to_mon_year(None) is None
    assert ces._to_mon_year(".") is None
    assert ces._to_mon_year("JUL-2024") == date(2024, 7, 1)
    assert ces._to_mon_year("XXX-2024") is None
    assert ces._to_mon_year("not a date") is None


def test_clean_label_branches():
    """``_clean_label`` resolves dot- and nbsp-indented hierarchy depths."""
    assert ces._clean_label(".Total nonfarm") == (1, "Total nonfarm")
    assert ces._clean_label("...Goods-producing") == (3, "Goods-producing")
    assert ces._clean_label("\xa0\xa0Mining and logging") == (2, "Mining and logging")
    assert ces._clean_label("Total private") == (0, "Total private")
    assert ces._clean_label(None) == (0, "")


def test_reference_date_branches():
    """``_reference_date`` returns the latest month, skipping bad tokens / non-strings."""
    rows = [
        (None, "Change from Mar. 2026 to Apr. 2026", 12345),
        ("Xyz 2026 should not match",),
    ]
    assert ces._reference_date(rows) == date(2026, 4, 1)
    assert ces._reference_date([(None, "no dates here")]) is None


def test_find_data_start_missing_raises():
    """``_find_data_start`` raises when no Industry/Rank header row exists."""
    with pytest.raises(OpenBBError, match="Industry"):
        ces._find_data_start([("foo", "bar"), ("baz", "qux")])


# --------------------------------------------------------------------------
# per-table parsers — synthetic workbooks exercising blank-skip + footnote-break
# --------------------------------------------------------------------------


def test_parse_table1_synthetic():
    """Table 1 parses values, significance, hierarchy; skips blank, stops at footnote."""
    rows = [
        ["TABLE 1. Employment..."],
        [None, "Not seasonally adjusted", None, "Seasonally adjusted", None],
        [
            "Industry",
            "Normal seasonal movement",
            "Change from Mar. 2026 to Apr. 2026",
            "Change from Mar. 2026 to Apr. 2026",
            "Minimum significant change",
        ],
        [".Total nonfarm", "\xa0811.0", "\xa0926.0", "\xa0115.0*", "\xa0122.3"],
        [None, None, None, None, None],
        ["* Passed test of significance", None, None, None, None],
        ["...Goods-producing", "1.0", "2.0", "3.0", "4.0"],
    ]
    res = ces.parse_table1(_wb(rows))
    assert res["table_id"] == "ces-anatab1"
    assert len(res["rows"]) == 1
    rec = res["rows"][0]
    assert rec["label"] == "Total nonfarm"
    assert rec["indent_level"] == 1
    assert rec["change_sa"] == 115.0
    assert rec["change_sa_significant"] is True
    assert rec["reference_date"] == date(2026, 4, 1)


def test_parse_table2_synthetic():
    """Table 2 parses rank/NAICS/change; uses the fallback title when row 0 blank."""
    rows = [
        [None],
        [
            "Rank",
            "Industry",
            "NAICS",
            "Change from Mar. 2026 to Apr. 2026",
            "Minimum significant change",
            "Prior 3-month average",
        ],
        ["1", "Couriers and messengers", "492", "\xa037.9*", "\xa05.6", "\xa02.4"],
        [None, None, None, None, None, None],
        ["* Passed test of significance"],
    ]
    res = ces.parse_table2(_wb(rows))
    assert res["table_title"] == ces.CES_TITLES["t2"]
    rec = res["rows"][0]
    assert rec["rank"] == "1"
    assert rec["naics_code"] == "492"
    assert rec["change_sa_significant"] is True
    assert len(res["rows"]) == 1


def test_parse_table3a_significance_flags():
    """Table 3A flags significance on each of the six change columns."""
    rows = [
        ["TABLE 3A..."],
        [None],
        [
            "Industry",
            "Change from Mar. 2026 to Apr. 2026",
            "Change from Feb. 2026 to Mar. 2026",
            "Change from Jan. 2026 to Feb. 2026",
            "Current 3-month change",
            "Current 6-month change",
            "Current 12-month change",
        ],
        [".Total nonfarm", "115.0", "185.0*", "-156.0", "144.0", "328.0*", "251.0"],
        ["\xa0", None, None, None, None, None, None],
        ["* Passed test of significance"],
    ]
    res = ces.parse_table3a(_wb(rows))
    rec = res["rows"][0]
    assert rec["otm_change_prior_1"] == 185.0
    assert rec["otm_change_prior_1_significant"] is True
    assert rec["otm_change_latest_significant"] is False
    assert rec["current_6month_change_significant"] is True


def test_parse_table3b_no_significance():
    """Table 3B carries the six average columns without significance flags."""
    rows = [
        ["TABLE 3B..."],
        [None],
        [
            "Industry",
            "Change from Mar. 2026 to Apr. 2026",
            "Change from Feb. 2026 to Mar. 2026",
            "Change from Jan. 2026 to Feb. 2026",
            "Prior 3-month average",
            "Prior 6-month average",
            "Prior 12-month average",
        ],
        [".Total nonfarm", "115.0", "185.0", "-156.0", "63.0", "12.2", "20.3"],
    ]
    res = ces.parse_table3b(_wb(rows))
    rec = res["rows"][0]
    assert rec["prior_12month_average"] == 20.3
    assert "otm_change_latest_significant" not in rec


def test_parse_table4_synthetic():
    """Table 4 parses OTY number/percent + significance + minimum significant."""
    rows = [
        ["TABLE 4..."],
        [None, "Change from", None, "Minimum"],
        [None, "Apr. 2025 to Apr. 2026", None, "significant"],
        ["Industry", "Number", "Percent", "change"],
        [".Total private", "511.0*", "0.4", "260.4"],
        ["\xa0", None, None, None],
        ["* Passed test of significance"],
    ]
    res = ces.parse_table4(_wb(rows))
    rec = res["rows"][0]
    assert rec["oty_change_number"] == 511.0
    assert rec["oty_change_number_significant"] is True
    assert rec["oty_change_percent"] == 0.4
    assert rec["reference_date"] == date(2026, 4, 1)


def test_parse_table5_two_blocks():
    """Table 5 parses both the hours and earnings blocks and skips repeated headers."""
    header = [
        "Industry",
        "Normal seasonal movement",
        "Change from Mar. 2026 to Apr. 2026",
        "Change from Mar. 2026 to Apr. 2026",
        "Minimum significant change",
    ]
    rows = [
        ["TABLE 5..."],
        [
            "Average weekly hours",
            "Not seasonally adjusted",
            None,
            "Seasonally adjusted",
            None,
        ],
        header,
        ["Total private", "-0.5", "-0.4", "0.1", "0.1"],
        [None, None, None, None, None],
        [
            "Average hourly earnings",
            "Not seasonally adjusted",
            None,
            "Seasonally adjusted",
            None,
        ],
        header,
        ["Total private", "-0.09", "-0.03", "0.06*", "0.05"],
        ["* Passed test of significance"],
    ]
    res = ces.parse_table5(_wb(rows))
    measures = {r["measure"] for r in res["rows"]}
    assert measures == {"Average weekly hours", "Average hourly earnings"}
    earnings = [r for r in res["rows"] if r["measure"] == "Average hourly earnings"][0]
    assert earnings["change_sa"] == 0.06
    assert earnings["change_sa_significant"] is True


def test_parse_table6_two_blocks():
    """Table 6 switches measure between weekly hours and payroll blocks."""
    rows = [
        ["TABLE 6..."],
        [None, None, "Change from", None, "Change from", None],
        [
            None,
            "Aggregate",
            "Mar. 2026 to Apr. 2026",
            None,
            "Apr. 2025 to Apr. 2026",
            None,
        ],
        ["Industry", "weekly hours", "Number", "Percent", "Number", "Percent"],
        ["Total private", "4645180", "17749", "0.4", "17527", "0.4"],
        [None, None, None, None, None, None],
        ["Industry", "weekly payrolls", "Number", "Percent", "Number", "Percent"],
        ["Total private", "173776184", "941636", "0.5", "6625358", "4.0"],
        ["* Passed test of significance", None, None, None, None, None],
    ]
    res = ces.parse_table6(_wb(rows))
    measures = {r["measure"] for r in res["rows"]}
    assert measures == {"Aggregate weekly hours", "Aggregate weekly payrolls"}
    payroll = [r for r in res["rows"] if r["measure"] == "Aggregate weekly payrolls"][0]
    assert payroll["aggregate_value"] == 173776184.0
    assert payroll["oty_change_percent"] == 4.0


def test_parse_table7_peak_trough():
    """Table 7 parses peak/trough dates, employment levels, and changes."""
    rows = [
        ["TABLE 7..."],
        [None, None, None, None, None, None, "Change to Apr. 2026", None],
        [
            "Industry",
            "Employment Apr. 2026",
            "Most recent peak",
            "Peak employment",
            "Most recent trough",
            "Trough employment",
            "Most recent peak",
            "Most recent trough",
        ],
        [".Total nonfarm", "158736.0", ".", ".", ".", ".", ".", "."],
        [
            "...Goods-producing",
            "21523.0",
            "JUL-2024",
            "21636.0",
            "APR-2020",
            "18523.0",
            "-113.0",
            "3000.0",
        ],
        ["\xa0", None, None, None, None, None, None, None],
        ["* The CES peak/trough program..."],
    ]
    res = ces.parse_table7(_wb(rows))
    nonfarm = res["rows"][0]
    assert nonfarm["peak_date"] is None
    goods = res["rows"][1]
    assert goods["peak_date"] == date(2024, 7, 1)
    assert goods["peak_employment"] == 21636.0
    assert goods["change_from_peak"] == -113.0


def test_parse_confidence_intervals_wide_and_narrow():
    """The CI parser handles Table A (wide) + B/C (narrow), skips blanks, stops at footnote."""
    sheets = {
        "CI Industry Employment Chang": [
            ["TABLE A. 90 percent confidence intervals..."],
            [None],
            [
                "Industry",
                "1-month change, first release",
                "1-month change, second release",
                "1-month change, third release",
                "3-month change",
                "6-month change",
                "12-month change",
            ],
            [
                "Total nonfarm",
                "122.28",
                "100.88",
                "98.01",
                "168.92",
                "215.20",
                "280.77",
            ],
            [
                "\xa0Goods-producing",
                "37.70",
                "32.81",
                "32.70",
                "56.71",
                "76.99",
                "100.73",
            ],
            ["\xa0", None, None, None, None, None, None],
        ],
        "CI AOT_AE": [
            ["TABLE C1. 90 percent confidence intervals..."],
            [None],
            [
                "Industry",
                "1-month change, first release",
                "3-month change",
                "6-month change",
                "12-month change",
            ],
            ["Manufacturing", "0.07", "0.08", "0.12", "0.12"],
            ["*Significant changes are calculated at..."],
        ],
    }
    res = ces.parse_confidence_intervals(_wb_sheets(sheets))
    rows = res["rows"]
    table_a = [r for r in rows if r["ci_table"] == "A"]
    table_c1 = [r for r in rows if r["ci_table"] == "C1"]
    assert len(table_a) == 2
    assert table_a[0]["ci_1month_second"] == 100.88
    assert table_a[0]["measure"] == "Employment"
    assert table_a[0]["employee_group"] is None
    assert len(table_c1) == 1
    assert table_c1[0]["ci_1month_second"] is None
    assert table_c1[0]["measure"] == "Average overtime hours"
    assert table_c1[0]["employee_group"] == "All employees"


def test_parse_confidence_intervals_unknown_sheet_falls_back():
    """A sheet outside the known map uses the sheet name for ci_table / measure."""
    sheets = {
        "Mystery Sheet": [
            ["Some heading without a table letter"],
            [None],
            [
                "Industry",
                "1-month change, first release",
                "3-month change",
                "6-month change",
                "12-month change",
            ],
            ["Total private", "0.05", "0.05", "0.07", "0.07"],
        ]
    }
    res = ces.parse_confidence_intervals(_wb_sheets(sheets))
    rec = res["rows"][0]
    assert rec["ci_table"] == "Mystery Sheet"
    assert rec["measure"] == "Mystery Sheet"


def test_parse_confidence_intervals_skips_empty_sheet():
    """A sheet with no rows is skipped without raising."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("CI AWH_AE")  # empty
    real = wb.create_sheet("CI AHE_AE")
    real.append(["TABLE B1. ..."])
    real.append([None])
    real.append(
        [
            "Industry",
            "1-month change, first release",
            "3-month change",
            "6-month change",
            "12-month change",
        ]
    )
    real.append(["Total private", "0.05", "0.08", "0.1", "0.12"])
    buf = io.BytesIO()
    wb.save(buf)
    res = ces.parse_confidence_intervals(buf.getvalue())
    assert len(res["rows"]) == 1
    assert res["rows"][0]["ci_table"] == "B1"


def test_parse_ces_dispatch():
    """``parse_ces`` routes a key to its parser."""
    rows = [
        ["TABLE 1..."],
        [None, "Not seasonally adjusted", None, "Seasonally adjusted", None],
        ["Industry", "Normal", "Change", "Change", "Min sig"],
        [".Total nonfarm", "1.0", "2.0", "3.0", "4.0"],
    ]
    res = ces.parse_ces(_wb(rows), "t1")
    assert res["table_id"] == "ces-anatab1"


def test_single_table_fetcher_empty_rows_raises():
    """A single-table fetcher raises EmptyDataError when no rows are parsed."""
    query = cesm.BlsCesTable1Fetcher.transform_query({})
    with pytest.raises(EmptyDataError):
        cesm.BlsCesTable1Fetcher.transform_data(
            query, {"rows": [], "table_id": "ces-anatab1"}
        )
