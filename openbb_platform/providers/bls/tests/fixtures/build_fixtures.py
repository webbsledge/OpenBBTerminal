"""One-shot generator for minimal BLS XLSX test fixtures.

Run from the providers/bls directory:

    /Users/darrenlee/miniconda3/envs/obb5/bin/python tests/fixtures/build_fixtures.py

Produces small reproducible XLSX files under ``tests/fixtures/`` that exercise
the BLS parsers without paying the cost of downloading the full multi-MB
workbooks. Run only when the parser expectations change.
"""

from __future__ import annotations

import io
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook

FIXTURES = Path(__file__).parent

_PPI_DETAILED_REPORT_URL = (
    "https://www.bls.gov/ppi/detailed-report/ppi-detailed-report-april-2026.xlsx"
)
_PPI_USER_AGENT = "openbb-platform-bls/2.0 (hello@openbb.co)"


def build_productivity_lps_mini() -> bytes:
    """Trim labor-productivity-major-sectors.xlsx to MachineReadable + a few rows."""
    wb = Workbook()
    wb.remove(wb.active)

    # Quarterly sheet — minimal header to satisfy release-date parsing
    quarterly = wb.create_sheet("Quarterly")
    quarterly.append(
        [
            "Quarterly labor productivity and related measures for major sectors, "
            "N.A. = data not available"
        ]
    )
    quarterly.append(
        [
            "Data released May 7, 2026; Bureau of Labor Statistics, "
            "Office of Productivity and Technology"
        ]
    )
    quarterly.append(["Sector", "Basis", "Measure", "Units", "2026 Q1"])

    # Annual placeholder (parser checks sheet existence for release date)
    wb.create_sheet("Annual")

    # MachineReadable — the actual parsed payload
    mr = wb.create_sheet("MachineReadable")
    mr.append(["Sector", "Basis", "Measure", "Units", "Year", "Qtr", "Value"])
    rows = [
        (
            "Nonfarm business sector",
            "All workers",
            "Labor productivity",
            "% Change from previous quarter",
            2025,
            4,
            1.8,
        ),
        (
            "Nonfarm business sector",
            "All workers",
            "Labor productivity",
            "% Change from previous quarter",
            2026,
            1,
            2.1,
        ),
        (
            "Nonfarm business sector",
            "All workers",
            "Labor productivity",
            "Index (2017=100)",
            2025,
            4,
            110.8,
        ),
        (
            "Nonfarm business sector",
            "All workers",
            "Labor productivity",
            "Index (2017=100)",
            2026,
            1,
            112.0,
        ),
        # Placeholder "Level - not available" row (always blank) — must be dropped.
        (
            "Nonfarm business sector",
            "All workers",
            "Labor productivity",
            "Level - not available",
            2026,
            1,
            None,
        ),
        (
            "Nonfarm business sector",
            "All workers",
            "Hours worked",
            "% Change from previous quarter",
            2026,
            1,
            0.5,
        ),
        (
            "Business sector",
            "All workers",
            "Labor productivity",
            "% Change from previous quarter",
            2026,
            1,
            2.0,
        ),
        (
            "Nonfarm business sector",
            "All workers",
            "Labor productivity",
            "Index (2017=100)",
            2025,
            "Annual",
            109.5,
        ),
    ]
    for r in rows:
        mr.append(r)

    # BusinessCycles — minimal header so dataset='major-sectors-business-cycles' parses
    bc = wb.create_sheet("BusinessCycles")
    bc.append(
        [
            "Labor productivity and related measures for major sectors by "
            "approximate business cycle period, N.A. = data not available"
        ]
    )
    bc.append(
        [
            "Data released May 7, 2026; Bureau of Labor Statistics, "
            "Office of Productivity and Technology"
        ]
    )
    bc.append(["Sector", "Basis", "Measure", "Units", "2020 Q1 - 2026 Q1"])
    bc.append(
        [
            "Nonfarm business sector",
            "All workers",
            "Labor productivity",
            "Compound annual growth rate",
            1.9,
        ]
    )

    wb.create_sheet("ReadMe")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_productivity_teh_mini() -> bytes:
    """Trim total-economy-hours-employment.xlsx to MachineReadable + a few rows."""
    wb = Workbook()
    wb.remove(wb.active)
    quarterly = wb.create_sheet("Quarterly")
    quarterly.append(
        [
            "Quarterly hours worked and employment in total U.S. economy and "
            "subsectors, N.A. = data not available"
        ]
    )
    quarterly.append(
        [
            "Data released May 7, 2026; Bureau of Labor Statistics, "
            "Office of Productivity and Technology"
        ]
    )
    quarterly.append(["Sector", "Basis", "Component", "Measure", "Units", "2026 Q1"])

    mr = wb.create_sheet("MachineReadable")
    mr.append(
        ["Sector", "Basis", "Component", "Measure", "Units", "Year", "Qtr", "Value"]
    )
    rows = [
        (
            "Total economy",
            "All workers",
            "Total U.S. economy",
            "Hours worked",
            "Billions of hours",
            2025,
            4,
            67.2,
        ),
        (
            "Total economy",
            "All workers",
            "Total U.S. economy",
            "Hours worked",
            "Billions of hours",
            2026,
            1,
            67.5,
        ),
        (
            "Total economy",
            "All workers",
            "Nonfarm business: total",
            "Employment",
            "Millions of jobs",
            2026,
            1,
            132.1,
        ),
    ]
    for r in rows:
        mr.append(r)

    wb.create_sheet("ReadMe")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_jolts_revisions_mini(seasonally_adjusted: bool) -> bytes:
    """Build a minimal JOLTS revisions workbook (one industry sheet, 3 months)."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("00")
    sa_token = (
        "seasonally adjusted" if seasonally_adjusted else "not seasonally adjusted"
    )
    ws.append([f"TOTAL NONFARM, {sa_token}, in thousands"] + [None] * 53)
    measures = (
        "JOB OPENINGS",
        "HIRES",
        "TOTAL SEPARATIONS",
        "QUITS",
        "LAYOFFS & DISCHARGES",
        "OTHER SEPARATIONS",
    )
    row1: list = []
    for measure in measures:
        row1.extend(["Month", measure, None, None, None, None, None, None, " "])
    ws.append(row1)
    row2: list = []
    for _ in measures:
        row2.extend(
            [
                None,
                "Level",
                None,
                None,
                "Revision 1st to 2nd",
                None,
                "Revision 2nd to Benchmark",
                None,
                " ",
            ]
        )
    ws.append(row2)
    row3: list = []
    for _ in measures:
        row3.extend(
            [
                None,
                "1st",
                "2nd",
                "Benchmark",
                "Level change",
                "% change",
                "Level change",
                "% change",
                " ",
            ]
        )
    ws.append(row3)
    for month_offset in range(3):
        data_row: list = []
        date_val = datetime(2024, month_offset + 1, 1)
        base = 5000 + month_offset * 10
        for i, _ in enumerate(measures):
            data_row.extend(
                [
                    date_val,
                    base + i * 100,
                    base + i * 100 + 5,
                    base + i * 100 + 12,
                    5,
                    0.001 * (i + 1),
                    7,
                    0.0015 * (i + 1),
                    " ",
                ]
            )
        ws.append(data_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _append_cpi_main_sheet(ws, title, label_header, section, data_rows) -> None:
    """Append the standard CPI supplemental layout (3 notes + 2 headers + data)."""
    ws.append(["", title, None, None, None])
    ws.append(["", "[1982-84=100, unless otherwise noted]", None, None, None])
    ws.append(["", "", None, None, None])
    ws.append(
        [
            "Indent Level",
            label_header,
            "Relative importance 2025",
            "Unadjusted indexes",
            "Unadjusted indexes",
        ]
    )
    ws.append([None, None, None, "Mar. 2026", "Apr. 2026"])
    ws.append(["", "", None, None, None])
    ws.append([None, None, None, None, None])
    ws.append([0, section, None, None, None])
    for row in data_rows:
        ws.append(row)


def build_cpi_supp_cpi_u_mini() -> bytes:
    """Build the cpi-u supplemental workbook with both US and Regional sheets."""
    wb = Workbook()
    wb.remove(wb.active)
    us = wb.create_sheet("US")
    _append_cpi_main_sheet(
        us,
        "Table 1. CPI-U U.S. city average, by expenditure category",
        "Expenditure category",
        "Expenditure category",
        [
            [0, "All items", 100.0, 318.0, 319.5],
            [1, "Food", 13.5, 290.0, 291.2],
            [1, "Energy", 6.5, 280.4, 282.0],
        ],
    )
    regional = wb.create_sheet("Regional")
    _append_cpi_main_sheet(
        regional,
        "Table 2. CPI-U selected areas, all items index",
        "Area",
        "Region",
        [
            [0, "Northeast", None, 332.0, 333.4],
            [0, "Midwest", None, 305.1, 306.0],
            [0, "South", None, 312.7, 313.9],
        ],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cpi_supp_cpi_w_mini() -> bytes:
    """Build the single-sheet cpi-w supplemental workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Sheet0")
    _append_cpi_main_sheet(
        ws,
        "Table 1. CPI-W U.S. city average, by expenditure category",
        "Expenditure category",
        "Expenditure category",
        [
            [0, "All items", 100.0, 311.0, 312.3],
            [1, "Housing", 39.8, 333.2, 334.1],
            [1, "Transportation", 18.2, 271.5, 272.8],
        ],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cpi_supp_historical_mini() -> bytes:
    """Build the historical-cpi-u workbook with Index values + Index averages sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    values = wb.create_sheet("Index values")
    values.append(["Historical CPI-U U.S. city average, all items"] + [None] * 13)
    values.append(
        [
            "Indent Level",
            "Year",
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]
    )
    values.append(
        [
            None,
            2025,
            300.0,
            300.5,
            301.0,
            301.6,
            302.1,
            302.7,
            303.2,
            303.8,
            304.3,
            304.9,
            305.4,
            306.0,
        ]
    )
    values.append(
        [
            None,
            2026,
            306.6,
            307.1,
            307.7,
            308.2,
            308.8,
            309.3,
            309.9,
            310.4,
            311.0,
            311.5,
            312.1,
            312.6,
        ]
    )

    averages = wb.create_sheet("Index averages")
    averages.append(["Historical CPI-U U.S. city average, averages"] + [None] * 6)
    averages.append(
        [
            "Indent Level",
            "Year",
            "Semiannual",
            "Semiannual",
            "Annual avg",
            "Percent change from previous",
            "Percent change from previous",
        ]
    )
    averages.append(
        [None, None, "1st half", "2nd half", None, "Dec. to Dec.", "Annual avg"]
    )
    averages.append([None, 2025, 301.0, 304.0, 302.5, 2.9, 2.7])
    averages.append([None, 2026, 307.5, 311.0, 309.2, 3.1, 2.8])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_ppi_detailed_report_trimmed() -> bytes:
    """Download the April 2026 PPI Detailed Report XLSX and trim every sheet.

    Parameters
    ----------
    None

    Returns
    -------
    bytes
        Trimmed XLSX bytes — header block + ~12 data rows + footnote tail
        per sheet, preserving the parser's expected layout.

    Notes
    -----
    Source URL is ``_PPI_DETAILED_REPORT_URL``. We pull the full ~1.2 MB
    workbook into memory, then for each ``Table N`` sheet we keep the rows
    up to and including the two-row header pair (the parser locates this
    by detecting an ``Indent Level`` cell in column A), 12 data rows after
    that, and the trailing ~12 footnote-legend rows. Result is ~35 KB.
    """
    req = urllib.request.Request(
        _PPI_DETAILED_REPORT_URL, headers={"User-Agent": _PPI_USER_AGENT}
    )
    with urllib.request.urlopen(req, timeout=60) as resp:  # noqa: S310
        src_bytes = resp.read()

    src = openpyxl.load_workbook(io.BytesIO(src_bytes), read_only=True, data_only=True)
    dst = Workbook()
    dst.remove(dst.active)

    keep_data = 12
    keep_footnotes = 12

    for sheet_name in src.sheetnames:
        if sheet_name == "Table of Contents":
            continue
        src_ws = src[sheet_name]
        rows = list(src_ws.iter_rows(values_only=True))
        dst_ws = dst.create_sheet(sheet_name)

        indent_idx = None
        for i, r in enumerate(rows):
            if not r or r[0] is None:
                continue
            first = str(r[0]).replace("\n", " ").strip().lower()
            if first == "indent level":
                indent_idx = i
                break
        if indent_idx is None:
            for r in rows[: 6 + keep_data]:
                dst_ws.append(list(r))
            continue

        for i in range(0, indent_idx + 2):
            dst_ws.append(list(rows[i]))

        data_start = indent_idx + 2
        end_data = min(data_start + keep_data, len(rows))
        for i in range(data_start, end_data):
            dst_ws.append(list(rows[i]))

        ncols = max((len(r) for r in rows), default=1)
        dst_ws.append([None] * ncols)

        trailing_start = max(len(rows) - keep_footnotes, end_data)
        for i in range(trailing_start, len(rows)):
            dst_ws.append(list(rows[i]))

    src.close()
    buf = io.BytesIO()
    dst.save(buf)
    return buf.getvalue()


_CES_STEMS = (
    "cesanatab1",
    "cesanatab2",
    "cesanatab3a",
    "cesanatab3b",
    "cesanatab4",
    "cesanatab5",
    "cesanatab6",
    "cesanatab7",
    "cesconfidenceintervals",
)


def _ces_norm(value) -> str:
    """Collapse non-breaking spaces / newlines in a CES cell."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def _trim_ces_rows(rows: list, per_block: int = 5) -> list:
    """Keep the header block, a few data rows per section, and the footnote.

    Section breaks (the repeated ``Average hourly earnings`` block in Table 5
    and the per-sheet ``Industry`` sub-headers) are always retained so the
    parser still sees the structure it keys on.
    """
    header_idx = None
    for i, row in enumerate(rows):
        if row and _ces_norm(row[0]).lower() in ("industry", "rank"):
            header_idx = i
            break
    if header_idx is None:
        return [list(r) for r in rows]
    out = [list(r) for r in rows[: header_idx + 1]]
    kept = 0
    for row in rows[header_idx + 1 :]:
        first = _ces_norm(row[0])
        if not first:
            continue
        if first.startswith("*"):
            out.append(list(row))
            break
        is_section = first.lower() in (
            "average weekly hours",
            "average hourly earnings",
            "industry",
        ) or _ces_norm(row[1]).lower().startswith("normal")
        if is_section:
            out.append(list(row))
            kept = 0
            continue
        if kept < per_block:
            out.append(list(row))
            kept += 1
    return out


def build_ces_trimmed(stem: str) -> bytes:
    """Download one CES analytical-table workbook and trim every sheet."""
    url = f"https://www.bls.gov/web/empsit/{stem}.xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": _PPI_USER_AGENT})
    with urllib.request.urlopen(req, timeout=90) as resp:  # noqa: S310
        src_bytes = resp.read()
    return _trim_ces_bytes(src_bytes)


def _trim_ces_bytes(src_bytes: bytes) -> bytes:
    """Trim every sheet of a CES workbook held in memory."""
    src = openpyxl.load_workbook(io.BytesIO(src_bytes), read_only=True, data_only=True)
    dst = Workbook()
    dst.remove(dst.active)
    for sheet_name in src.sheetnames:
        rows = list(src[sheet_name].iter_rows(values_only=True))
        dst_ws = dst.create_sheet(sheet_name[:31])
        for row in _trim_ces_rows(rows):
            dst_ws.append(row)
    src.close()
    buf = io.BytesIO()
    dst.save(buf)
    return buf.getvalue()


_XIMPIM_CHART_SLUGS = {
    "import-export": "us-import-and-export-price-indexes-12-month-percent-change",
    "imports-by-category": "us-import-price-indexes-by-category-12-month-percent-change",
    "exports-by-category": "us-export-price-indexes-by-category-12-month-percent-change",
    "imports-by-origin": "us-import-price-indexes-by-origin-12-month-percent-change",
    "exports-by-grains": "us-export-price-indexes-by-selected-grains-12-month-percent-change",
    "air-passenger-fares": "air-passenger-fares-12-month-percent-change",
}


def _trim_chart_page(base: str, slug: str, keep: int = 4) -> bytes:
    """Download a BLS news-release chart page and keep only its trimmed data table."""
    from bs4 import BeautifulSoup  # ty: ignore[unresolved-import]

    url = f"https://www.bls.gov/charts/{base}/{slug}.htm"
    req = urllib.request.Request(
        url, headers={"User-Agent": _PPI_USER_AGENT, "Accept": "text/html,*/*"}
    )
    with urllib.request.urlopen(req, timeout=60) as resp:  # noqa: S310
        html = resp.read().decode("utf-8", "replace")
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", class_="regular")
    body = table.find("tbody")
    trs = body.find_all("tr")
    keep_trs = set(trs[:keep])
    if len(trs) > keep:
        keep_trs.add(trs[-1])
    for tr in trs:
        if tr not in keep_trs:
            tr.decompose()
    return ("<!DOCTYPE html><html><body>" + str(table) + "</body></html>").encode(
        "utf-8"
    )


def build_ximpim_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one import-export chart page and keep only its trimmed data table."""
    return _trim_chart_page("import-export", slug, keep)


def build_empsit_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one Employment Situation chart page and keep only its trimmed table."""
    return _trim_chart_page("employment-situation", slug, keep)


def build_productivity_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one Productivity and Costs chart page and keep only its trimmed table."""
    return _trim_chart_page("productivity-and-costs", slug, keep)


def build_cpi_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one Consumer Price Index chart page and keep only its trimmed table."""
    return _trim_chart_page("consumer-price-index", slug, keep)


def build_ppi_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one Producer Price Index chart page and keep only its trimmed table."""
    return _trim_chart_page("producer-price-index", slug, keep)


def build_tfp_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one Total Factor Productivity chart page and keep only its trimmed table."""
    return _trim_chart_page("total-factor-productivity", slug, keep)


def build_wholesale_retail_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one wholesale/retail productivity chart page and trim its table."""
    return _trim_chart_page("productivity-wholesale-retail", slug, keep)


def build_mining_manufacturing_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one mining/manufacturing productivity chart page and trim its table."""
    return _trim_chart_page("productivity-mining-manufacturing", slug, keep)


def build_jolts_chart_trimmed(slug: str, keep: int = 4) -> bytes:
    """Download one JOLTS chart page and keep only its trimmed data table."""
    return _trim_chart_page("job-openings-and-labor-turnover", slug, keep)


def build_empsit_summary_trimmed(slug: str, keep_per_section: int = 3) -> bytes:
    """Download an Employment Situation summary table, keeping section headers
    and a few data rows per section."""
    from bs4 import BeautifulSoup  # ty: ignore[unresolved-import]

    url = f"https://www.bls.gov/news.release/{slug}.htm"
    req = urllib.request.Request(
        url, headers={"User-Agent": _PPI_USER_AGENT, "Accept": "text/html,*/*"}
    )
    with urllib.request.urlopen(req, timeout=60) as resp:  # noqa: S310
        html = resp.read().decode("utf-8", "replace")
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", class_="regular")
    body = table.find("tbody") or table
    trs = body.find_all("tr")
    keep_ids: set = set()
    in_section = 0
    for tr in trs:
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if not cells:
            continue
        if not any(v.strip() for v in cells[1:]):  # section / blank row
            keep_ids.add(id(tr))
            in_section = 0
        elif in_section < keep_per_section:
            keep_ids.add(id(tr))
            in_section += 1
    for tr in trs:
        if id(tr) not in keep_ids:
            tr.decompose()
    return ("<!DOCTYPE html><html><body>" + str(table) + "</body></html>").encode(
        "utf-8"
    )


def main() -> None:
    """Write all fixture XLSX files into the fixtures directory."""
    written: list[tuple[str, int]] = []

    builders: list[tuple[str, callable]] = [
        ("prod2_lps_mini.xlsx", build_productivity_lps_mini),
        ("prod2_teh_mini.xlsx", build_productivity_teh_mini),
        (
            "jolts_sa_rev_mini.xlsx",
            lambda: build_jolts_revisions_mini(seasonally_adjusted=True),
        ),
        (
            "jolts_nsa_rev_mini.xlsx",
            lambda: build_jolts_revisions_mini(seasonally_adjusted=False),
        ),
        (
            "ppi_detailed_report_trimmed.xlsx",
            build_ppi_detailed_report_trimmed,
        ),
        ("cpi_supp_cpi_u.xlsx", build_cpi_supp_cpi_u_mini),
        ("cpi_supp_cpi_w.xlsx", build_cpi_supp_cpi_w_mini),
        ("cpi_supp_historical_cpi_u.xlsx", build_cpi_supp_historical_mini),
        *(
            (f"{stem}.xlsx", (lambda s=stem: build_ces_trimmed(s)))
            for stem in _CES_STEMS
        ),
    ]
    for name, fn in builders:
        path = FIXTURES / name
        path.write_bytes(fn())
        written.append((name, path.stat().st_size))

    for name, size in written:
        pass


if __name__ == "__main__":
    main()
